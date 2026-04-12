#!/usr/bin/env python3
"""
One workbook: load the empty family template once, apply JPM PDF fills + Goldman (via QB
extraction) fills, save a single output .xlsx.
"""

from __future__ import annotations

import tempfile
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


def populate_family_template(
    template_path: str | Path,
    output_path: str | Path,
    *,
    jpm_pdf: str | Path | None = None,
    jpm_accounts: Iterable[str] | None = None,
    gs_pdf_account_pairs: list[tuple[str | Path, str | None]] | None = None,
) -> str:
    """
    gs_pdf_account_pairs: list of (pdf_path, account_id or None).
    If account_id is None, infer from QB workbook (same as populate-template).
    """
    from pdf_to_qb import pdf_to_qb_excel
    from template_populator import apply_template_from_qb_workbook

    from scripts.populate_jpm_template_from_pdf import apply_jpm_pdf_to_workbook

    wb = load_workbook(str(template_path), keep_links=False)

    if jpm_pdf:
        apply_jpm_pdf_to_workbook(wb, jpm_pdf, accounts=jpm_accounts)

    for pdf_p, acct in gs_pdf_account_pairs or []:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            qb_tmp = f.name
        wb_src = None
        try:
            pdf_to_qb_excel(str(pdf_p), qb_tmp, overwrite=True)
            wb_src = load_workbook(qb_tmp, data_only=True)
            apply_template_from_qb_workbook(wb, wb_src, account_id=acct)
        finally:
            if wb_src is not None:
                wb_src.close()
            Path(qb_tmp).unlink(missing_ok=True)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return str(out)
