#!/usr/bin/env python3
"""
Populate the JPM blocks in the REC demo template directly from a JPM PDF.

This is intentionally a small, high-signal first step:
- detect which JPM account blocks exist in the template (e.g. 1004, 9004, 843, 8008)
- extract a few key MTD values from the PDF text layer (PyMuPDF)
- write those values into the correct account blocks

It does not try to fully replicate every manual workbook nuance yet (BOM/EOM balances,
change-in-accrual deltas, Unrealized sheet, etc.). Those can be layered on once the
core mapping is proven.
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Any, Iterable

import fitz  # PyMuPDF
from openpyxl import load_workbook


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", s).strip()


def _parse_money(s: str) -> float | None:
    s = s.strip()
    if not s:
        return None
    s = s.replace("$", "").replace(",", "")
    # accounting negative
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1].strip()
    try:
        return float(s)
    except Exception:
        return None


_MONEY_RE = re.compile(
    r"(?P<money>\(?-?\$?\s*\d{1,3}(?:,\d{3})*(?:\.\d{2})\)?|\(?-?\$?\s*\d+(?:\.\d{2})\)?)"
)


def _find_first_money_after_label(lines: list[str], label: str) -> float | None:
    """
    Find the first money-like token that occurs on a line after a label line.
    Handles layouts where the label and value are on separate lines.
    """
    label_l = label.lower()
    for i, ln in enumerate(lines):
        if label_l in ln.lower():
            # Look on same line first
            m = _MONEY_RE.search(ln)
            if m:
                v = _parse_money(m.group("money"))
                if v is not None:
                    return v
            # Then look ahead a few lines
            for j in range(i + 1, min(i + 6, len(lines))):
                m2 = _MONEY_RE.search(lines[j])
                if not m2:
                    continue
                v2 = _parse_money(m2.group("money"))
                if v2 is not None:
                    return v2
    return None


def _extract_jpm_mtd_values_from_pdf(pdf_path: str | Path) -> dict[str, float]:
    """
    Extract a minimal set of MTD values.

    For the provided JPM sample PDF, these appear on the page that contains:
    - Domestic Dividends/Distributions
    - Foreign Dividends
    - Interest Income
    - LT Realized Gain/Loss
    """
    doc = fitz.open(str(pdf_path))
    try:
        pages: list[str] = []
        for i in range(doc.page_count):
            pages.append(doc.load_page(i).get_text("text") or "")

        # Find the "income summary" page by keywords.
        income_page_text = None
        for t in pages:
            tl = t.lower()
            if "domestic dividends" in tl and "foreign dividends" in tl and "interest income" in tl:
                income_page_text = t
                break
        if income_page_text is None:
            # fallback: any page with interest income + realized
            for t in pages:
                tl = t.lower()
                if "interest income" in tl and "realized" in tl:
                    income_page_text = t
                    break
        if income_page_text is None:
            return {}

        lines = [l for l in (_norm(x) for x in income_page_text.splitlines()) if l]
        domestic = _find_first_money_after_label(lines, "Domestic Dividends/Distributions")
        foreign = _find_first_money_after_label(lines, "Foreign Dividends")
        interest = _find_first_money_after_label(lines, "Interest Income")
        lt_realized = _find_first_money_after_label(lines, "LT Realized Gain/Loss")

        out: dict[str, float] = {}
        if domestic is not None or foreign is not None:
            out["Dividend Income"] = float((domestic or 0.0) + (foreign or 0.0))
        if interest is not None:
            out["Interest Income"] = float(interest)
        if lt_realized is not None:
            out["Long term realized gain/loss"] = float(lt_realized)
        return out
    finally:
        doc.close()


def _find_account_id_cols(ws) -> dict[str, int]:
    """
    In our REC template, account ids are in row 1 on the JPM sheet.
    Returns mapping like {"1004": col_idx, "9004": col_idx, ...}
    """
    out: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is None:
            continue
        s = str(v).strip()
        if s.isdigit() and 2 <= len(s) <= 6:
            out[s] = c
    return out


def _find_label_row(ws, label_col: int, label: str) -> int | None:
    target = label.strip().lower()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, label_col).value
        if isinstance(v, str) and v.strip().lower() == target:
            return r
    return None


def apply_jpm_pdf_to_workbook(wb, pdf_path: str | Path, *, accounts: Iterable[str] | None = None) -> None:
    """
    Fill JPM account blocks on an open template workbook (mutates in place).
    """
    if "PLSummary JP Morgan Chase" not in wb.sheetnames:
        raise ValueError("Template missing sheet: 'PLSummary JP Morgan Chase'")

    ws = wb["PLSummary JP Morgan Chase"]
    acct_cols = _find_account_id_cols(ws)
    if not acct_cols:
        raise ValueError("Could not find any JPM account blocks on row 1 of template sheet")

    extracted = _extract_jpm_mtd_values_from_pdf(pdf_path)
    label_to_value: dict[str, float] = extracted
    target_accounts = list(accounts) if accounts else sorted(acct_cols.keys())

    for acct in target_accounts:
        if acct not in acct_cols:
            continue
        id_col = acct_cols[acct]
        block_start = id_col + 2
        value_col = block_start + 1

        for label, val in label_to_value.items():
            rr = _find_label_row(ws, block_start, label)
            if rr is None:
                continue
            ws.cell(rr, value_col).value = float(val)


def populate_jpm_template_from_pdf(
    template_path: str | Path,
    pdf_path: str | Path,
    output_path: str | Path,
    *,
    accounts: Iterable[str] | None = None,
) -> str:
    tpl = load_workbook(str(template_path), keep_links=False)
    apply_jpm_pdf_to_workbook(tpl, pdf_path, accounts=accounts)
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    tpl.save(out)
    return str(out)


def main() -> int:
    ap = argparse.ArgumentParser(description="Populate JPM blocks in the REC demo template from a JPM PDF")
    ap.add_argument("--template", required=True, help="Path to template .xlsx (e.g. REC_demo_family_template_populated.xlsx)")
    ap.add_argument("--pdf", required=True, help="Path to JPM PDF (e.g. 9004-...Combined Statement.pdf)")
    ap.add_argument("-o", "--output", required=True, help="Output filled .xlsx path")
    ap.add_argument(
        "--accounts",
        default="",
        help="Comma-separated list of account blocks to fill (e.g. 1004,9004). Default: fill all blocks present in template.",
    )
    args = ap.parse_args()

    accounts = [a.strip() for a in args.accounts.split(",") if a.strip()] if args.accounts else None
    populate_jpm_template_from_pdf(args.template, args.pdf, args.output, accounts=accounts)
    print(f"Saved: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

