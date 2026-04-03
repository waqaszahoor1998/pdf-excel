#!/usr/bin/env python3
"""
Web UI for PDF → Excel (local only, no cloud LLM).

Run: flask --app app run
Or:  python app.py

Then open http://127.0.0.1:5000 — upload a PDF. Primary output is canonical JSON;
optional ZIP adds Excel (QuickBooks-style). Use output_package json_only for JSON
without building Excel. Hybrid/vision use the local VL stack; Ask AI uses Anthropic.
After extraction, pages are audited vs the PDF; summary is in JSON meta and the ZIP
may include a full audit report. No raw data is sent to any cloud service except
Ask AI when that mode is used.
"""

import io
import json
import logging
import os
import tempfile
import zipfile
from pathlib import Path

from dotenv import load_dotenv
from flask import Flask, render_template, request, send_file, flash, redirect, url_for
from werkzeug.exceptions import RequestEntityTooLarge

_env_path = Path(__file__).resolve().parent / ".env"
load_dotenv(_env_path)
# Set CUDA before any VL/llama_cpp use so GPU is used
try:
    from extract_vl import _ensure_cuda_path
    _ensure_cuda_path()
except ImportError:
    pass
log = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(levelname)s [%(name)s] %(message)s")

from tables_to_excel import (
    extract_sections_from_pdf,
    _write_json_from_sections,
    load_sections_from_json,
    _write_sections_to_workbook,
    write_sections_to_workbook_by_page,
    _normalize_page_to_sheet,
)
from pdf_to_qb import pdf_to_qb_excel, transform_extracted_to_qb
from pdf_json_audit import apply_audit_to_extraction_file, audit_pdf_vs_extraction_json
from template_populator import populate_template_from_fields_json, populate_template_from_qb_output

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-secret-change-in-production")

# Max upload size for PDFs (from env or default)
APP_ROOT = Path(__file__).resolve().parent
_val = (os.environ.get("MAX_UPLOAD_MB") or "").strip() or "40"
_default_mb = max(1, min(100, int(_val) if _val.isdigit() else 40))
MAX_CONTENT_LENGTH = _default_mb * 1024 * 1024
app.config["MAX_CONTENT_LENGTH"] = MAX_CONTENT_LENGTH

# Web UI: cap VL/hybrid page count; audit runs on full PDF by default (see AUDIT_MAX_PAGES)
WEB_MAX_VL_PAGES = 20


def _web_audit_page_cap():
    """None = audit all pages. Set AUDIT_MAX_PAGES=N to audit first N only (faster)."""
    v = (os.environ.get("AUDIT_MAX_PAGES") or "").strip()
    if not v:
        return None
    try:
        n = int(v)
        return n if n > 0 else None
    except ValueError:
        return None


def _get_upload_limit_mb():
    return MAX_CONTENT_LENGTH // (1024 * 1024)


def _ai_available():
    return bool(os.environ.get("ANTHROPIC_API_KEY", "").strip())


def _vl_available():
    """True if VL extraction (Qwen2.5-VL) is installed and model is present."""
    try:
        from extract_vl import _model_paths
        model_path, mmproj_path = _model_paths()
        return model_path and Path(model_path).exists() and mmproj_path and Path(mmproj_path).exists()
    except Exception:
        return False


def _extract_method_from_form(form) -> str:
    """library | hybrid | vl — default library."""
    m = (form.get("extract_method") or "").strip().lower()
    if m in ("library", "hybrid", "vl"):
        return m
    return "vl" if (form.get("use_vl") or "").strip().lower() in ("on", "1", "yes", "true") else "library"


def _flash_audit_summary(json_path: Path) -> None:
    """Flash a short message from meta.audit_summary after apply_audit_to_extraction_file."""
    try:
        payload = json.loads(json_path.read_text(encoding="utf-8"))
        meta = payload.get("meta") or {}
        aud = meta.get("audit_summary") or {}
        if int(aud.get("pages_audited") or 0) == 0:
            return
        conf = aud.get("confidence_pct")
        if aud.get("passed_automation"):
            flash(
                f"Audit: automation pass (confidence {conf}%, scope {aud.get('audit_scope', '?')}).",
                "success",
            )
        else:
            flash(
                (aud.get("automation_reason") or "Audit: automation failed.")
                + f" — confidence {conf}%.",
                "warning",
            )
    except Exception:
        pass


@app.errorhandler(RequestEntityTooLarge)
def too_large(e):
    flash(f"File too large. Maximum size is {_get_upload_limit_mb()} MB.")
    return redirect(url_for("index"))


@app.route("/")
def index():
    return render_template(
        "index.html",
        max_mb=_get_upload_limit_mb(),
        ai_available=_ai_available(),
        vl_available=_vl_available(),
    )


@app.route("/extract", methods=["POST"])
def extract():
    if "pdf" not in request.files:
        log.warning("extract: no file in request")
        flash("No file selected.")
        return redirect(url_for("index"))

    file = request.files["pdf"]
    if not file or file.filename == "":
        log.warning("extract: empty filename")
        flash("No file selected.")
        return redirect(url_for("index"))

    if not file.filename.lower().endswith(".pdf"):
        flash("Please upload a PDF file.")
        return redirect(url_for("index"))

    mode = (request.form.get("mode") or "offline").strip().lower()
    query = (request.form.get("query") or "").strip()
    if mode == "ai":
        if not query:
            flash("Enter a query for Ask AI (e.g. 'taxes for January').")
            return redirect(url_for("index"))
        if not _ai_available():
            flash("Ask AI requires ANTHROPIC_API_KEY in .env.")
            return redirect(url_for("index"))

    tmp_dir = None
    try:
        tmp_dir = tempfile.mkdtemp()
        pdf_path = Path(tmp_dir) / "upload.pdf"
        file.save(str(pdf_path))
        use_ai = mode == "ai" and query and _ai_available()
        log.info("extract: saved upload (mode=%s)", "ai" if use_ai else "offline")

        out_path = Path(tmp_dir) / "output.xlsx"
        json_path = Path(tmp_dir) / "output.json"
        audit_report_path = Path(tmp_dir) / "audit_report.json"
        extract_method = _extract_method_from_form(request.form)
        output_package = (request.form.get("output_package") or "full").strip().lower()
        json_only = (not use_ai) and output_package == "json_only"
        if use_ai:
            from config import load_config
            from extract import extract_pdf_to_excel
            cfg = load_config()
            result = extract_pdf_to_excel(str(pdf_path), query, str(out_path), config=cfg)
            json_path = None  # AI path does not produce JSON in same format
            audit_report_path = None
        elif extract_method in ("hybrid", "vl") and not _vl_available():
            flash(
                "Hybrid and vision-only extraction need the local vision model. "
                "Install requirements-vl.txt and run scripts/download_qwen2vl.py.",
            )
            return redirect(url_for("index"))
        elif extract_method == "hybrid":
            try:
                from hybrid_extract import hybrid_pdf_to_json

                hybrid_pdf_to_json(
                    str(pdf_path),
                    str(json_path),
                    max_pages=WEB_MAX_VL_PAGES,
                    overwrite=True,
                )
                if json_only:
                    result = str(json_path)
                else:
                    sections = load_sections_from_json(json_path)
                    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
                        temp_xlsx = f.name
                    try:
                        _write_sections_to_workbook(sections, Path(temp_xlsx))
                        transform_extracted_to_qb(temp_xlsx, str(out_path))
                    finally:
                        Path(temp_xlsx).unlink(missing_ok=True)
                    result = str(out_path)
            except ImportError:
                log.exception("Hybrid extract: import failed")
                flash("Hybrid extraction requires the vision stack (requirements-vl.txt).")
                return redirect(url_for("index"))
            except FileNotFoundError:
                log.exception("Hybrid extract: model or file not found")
                flash("Vision model files not found. Run scripts/download_qwen2vl.py to download the model.")
                return redirect(url_for("index"))
            except Exception as e:
                log.exception("Hybrid extract: %s", e)
                flash(f"Hybrid extraction failed: {e}")
                return redirect(url_for("index"))
        elif extract_method == "vl":
            # Vision model path: PDF → VL → JSON → Excel (for scanned PDFs)
            try:
                from extract_vl import pdf_to_json_vl

                pdf_to_json_vl(str(pdf_path), str(json_path), max_pages=WEB_MAX_VL_PAGES)
                if json_only:
                    result = str(json_path)
                else:
                    sections = load_sections_from_json(json_path)
                    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
                        temp_xlsx = f.name
                    try:
                        _write_sections_to_workbook(sections, Path(temp_xlsx))
                        transform_extracted_to_qb(temp_xlsx, str(out_path))
                    finally:
                        Path(temp_xlsx).unlink(missing_ok=True)
                    result = str(out_path)
            except ImportError:
                log.exception("VL extract: import failed")
                flash("Vision model not available. Install requirements-vl.txt and run scripts/download_qwen2vl.py.")
                return redirect(url_for("index"))
            except FileNotFoundError:
                log.exception("VL extract: model or file not found")
                flash("Vision model files not found. Run scripts/download_qwen2vl.py to download the model.")
                return redirect(url_for("index"))
            except Exception as e:
                log.exception("VL extract: %s", e)
                flash(f"Vision extraction failed: {e}")
                return redirect(url_for("index"))
        else:
            result = pdf_to_qb_excel(
                str(pdf_path),
                str(out_path),
                overwrite=True,
                json_path_out=str(json_path),
                write_excel=not json_only,
            )

        if json_path and json_path.exists() and not use_ai:
            apply_audit_to_extraction_file(
                pdf_path,
                json_path,
                audit_pages=_web_audit_page_cap(),
                report_path=audit_report_path,
                silent=True,
            )
            _flash_audit_summary(json_path)

        if not Path(result).exists():
            log.error("extract: result file missing %s", result)
            flash("Conversion produced no file.")
            return redirect(url_for("index"))

        base_name = Path(file.filename).stem
        if json_path and json_path.exists():
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
                if json_only:
                    zf.write(json_path, f"{base_name}.json")
                    if audit_report_path and audit_report_path.exists():
                        zf.write(audit_report_path, f"{base_name}_audit_report.json")
                else:
                    zf.write(result, f"{base_name}.xlsx")
                    zf.write(json_path, f"{base_name}.json")
                    if audit_report_path and audit_report_path.exists():
                        zf.write(audit_report_path, f"{base_name}_audit_report.json")
            buf.seek(0)
            log.info(
                "extract: sending zip (%s) for %s",
                "json+audit" if json_only else "xlsx+json+audit",
                base_name,
            )
            flash(
                "Download started. If nothing appears, check blocked downloads or your Downloads folder.",
                "success",
            )
            return send_file(
                buf,
                as_attachment=True,
                download_name=f"{base_name}.zip",
                mimetype="application/zip",
            )
        download_name = f"{base_name}.xlsx"
        log.info("extract: sending file %s", download_name)
        return send_file(
            result,
            as_attachment=True,
            download_name=download_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except FileNotFoundError as e:
        log.exception("extract: FileNotFoundError")
        flash(str(e))
        return redirect(url_for("index"))
    except ValueError as e:
        log.exception("extract: ValueError")
        flash(str(e))
        return redirect(url_for("index"))
    except Exception as e:
        log.exception("extract: %s", e)
        flash(f"Error: {e}")
        return redirect(url_for("index"))
    finally:
        if tmp_dir and Path(tmp_dir).exists():
            try:
                for f in Path(tmp_dir).iterdir():
                    f.unlink(missing_ok=True)
                Path(tmp_dir).rmdir()
            except OSError:
                pass


@app.route("/pdf-to-json", methods=["POST"])
def pdf_to_json_route():
    """Step 1: PDF → JSON only. Download JSON (meta includes audit summary on sampled pages)."""
    if "pdf" not in request.files:
        flash("No file selected.")
        return redirect(url_for("index"))
    file = request.files["pdf"]
    if not file or file.filename == "":
        flash("No file selected.")
        return redirect(url_for("index"))
    if not file.filename.lower().endswith(".pdf"):
        flash("Please upload a PDF file.")
        return redirect(url_for("index"))
    extract_method = _extract_method_from_form(request.form)
    if extract_method in ("hybrid", "vl") and not _vl_available():
        flash(
            "Hybrid and vision-only need the local vision model. "
            "Install requirements-vl.txt and run scripts/download_qwen2vl.py.",
        )
        return redirect(url_for("index"))
    tmp_dir = None
    try:
        tmp_dir = tempfile.mkdtemp()
        pdf_path = Path(tmp_dir) / "upload.pdf"
        file.save(str(pdf_path))
        json_path = Path(tmp_dir) / "output.json"
        audit_report_path = Path(tmp_dir) / "audit_report.json"
        if extract_method == "hybrid":
            from hybrid_extract import hybrid_pdf_to_json

            hybrid_pdf_to_json(
                str(pdf_path),
                str(json_path),
                max_pages=WEB_MAX_VL_PAGES,
                overwrite=True,
            )
        elif extract_method == "vl":
            from extract_vl import pdf_to_json_vl

            pdf_to_json_vl(str(pdf_path), str(json_path), max_pages=WEB_MAX_VL_PAGES)
        else:
            sections = extract_sections_from_pdf(str(pdf_path))
            _write_json_from_sections(sections, json_path, overwrite=True)
        if not json_path.exists():
            flash("Conversion produced no JSON.")
            return redirect(url_for("index"))
        apply_audit_to_extraction_file(
            pdf_path,
            json_path,
            audit_pages=_web_audit_page_cap(),
            report_path=audit_report_path,
            silent=True,
        )
        _flash_audit_summary(json_path)
        base_name = Path(file.filename).stem
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.write(json_path, f"{base_name}.json")
            if audit_report_path.exists():
                zf.write(audit_report_path, f"{base_name}_audit_report.json")
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name=f"{base_name}_json.zip",
            mimetype="application/zip",
        )
    except Exception as e:
        log.exception("pdf-to-json: %s", e)
        flash(f"Error: {e}")
        return redirect(url_for("index"))
    finally:
        if tmp_dir and Path(tmp_dir).exists():
            try:
                for f in Path(tmp_dir).iterdir():
                    f.unlink(missing_ok=True)
                Path(tmp_dir).rmdir()
            except OSError:
                pass


@app.route("/json-to-excel", methods=["POST"])
def json_to_excel_route():
    """Step 2: JSON → Excel. Upload JSON (from PDF→JSON or edited) to get structured xlsx.
    If JSON meta or config has page_to_sheet, sections are grouped by page into sheets.
    """
    import json as _json
    if "json_file" not in request.files:
        flash("No file selected.")
        return redirect(url_for("index"))
    file = request.files["json_file"]
    if not file or file.filename == "":
        flash("No file selected.")
        return redirect(url_for("index"))
    if not file.filename.lower().endswith(".json"):
        flash("Please upload a JSON file (from PDF→JSON step).")
        return redirect(url_for("index"))
    tmp_dir = None
    try:
        tmp_dir = tempfile.mkdtemp()
        json_path = Path(tmp_dir) / "input.json"
        file.save(str(json_path))
        out_xlsx = Path(tmp_dir) / "output.xlsx"
        sections = load_sections_from_json(json_path)
        with open(json_path, encoding="utf-8") as f:
            payload = _json.load(f)
        meta = payload.get("meta") or {}
        page_to_sheet = _normalize_page_to_sheet(meta.get("page_to_sheet") or {})
        if not page_to_sheet:
            vl_config_path = Path(__file__).resolve().parent / "config" / "vl.json"
            if vl_config_path.exists():
                with open(vl_config_path, encoding="utf-8") as f:
                    vl_cfg = _json.load(f)
                page_to_sheet = _normalize_page_to_sheet(vl_cfg.get("page_to_sheet") or {})
        has_page = any(len(s) >= 4 and s[3] is not None for s in sections)
        if page_to_sheet and has_page:
            write_sections_to_workbook_by_page(sections, page_to_sheet, out_xlsx)
        else:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
                temp_xlsx = f.name
            try:
                _write_sections_to_workbook(sections, Path(temp_xlsx))
                transform_extracted_to_qb(temp_xlsx, str(out_xlsx))
            finally:
                Path(temp_xlsx).unlink(missing_ok=True)
        if not out_xlsx.exists():
            flash("Conversion produced no Excel file.")
            return redirect(url_for("index"))
        base_name = Path(file.filename).stem
        download_name = f"{base_name}.xlsx"
        xlsx_bytes = out_xlsx.read_bytes()
        return send_file(
            io.BytesIO(xlsx_bytes),
            as_attachment=True,
            download_name=download_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        log.exception("json-to-excel: %s", e)
        flash(f"Error: {e}")
        return redirect(url_for("index"))
    finally:
        if tmp_dir and Path(tmp_dir).exists():
            try:
                for fp in Path(tmp_dir).iterdir():
                    fp.unlink(missing_ok=True)
                Path(tmp_dir).rmdir()
            except OSError:
                pass


@app.route("/workflows")
def workflows_page():
    """Reference + forms for every pipeline exposed in the web UI (mirrors CLI where possible)."""
    return render_template(
        "workflows.html",
        max_mb=_get_upload_limit_mb(),
        ai_available=_ai_available(),
        vl_available=_vl_available(),
    )


@app.route("/populate-template", methods=["POST"])
def populate_template_route():
    """PDF + user template .xlsx → filled template (same as: run.py populate-template)."""
    if "pdf" not in request.files or "template" not in request.files:
        flash("Upload a PDF and a template .xlsx.")
        return redirect(url_for("workflows_page"))
    pdf_f = request.files["pdf"]
    tpl_f = request.files["template"]
    if not pdf_f.filename or not tpl_f.filename:
        flash("PDF and template are required.")
        return redirect(url_for("workflows_page"))
    if not pdf_f.filename.lower().endswith(".pdf"):
        flash("First file must be a PDF.")
        return redirect(url_for("workflows_page"))
    if not tpl_f.filename.lower().endswith(".xlsx"):
        flash("Template must be a .xlsx file.")
        return redirect(url_for("workflows_page"))
    account_id = (request.form.get("account_id") or "").strip() or None
    tmp_dir = None
    try:
        tmp_dir = tempfile.mkdtemp()
        pdf_path = Path(tmp_dir) / "upload.pdf"
        tpl_path = Path(tmp_dir) / "template.xlsx"
        out_path = Path(tmp_dir) / "populated.xlsx"
        pdf_f.save(str(pdf_path))
        tpl_f.save(str(tpl_path))
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            qb_tmp = f.name
        try:
            pdf_to_qb_excel(str(pdf_path), qb_tmp, overwrite=True)
            populate_template_from_qb_output(
                template_path=str(tpl_path),
                qb_output_xlsx=qb_tmp,
                output_path=str(out_path),
                account_id=account_id,
            )
        finally:
            Path(qb_tmp).unlink(missing_ok=True)
        if not out_path.exists():
            flash("Template population produced no file.")
            return redirect(url_for("workflows_page"))
        base = Path(pdf_f.filename).stem
        xlsx_bytes = out_path.read_bytes()
        return send_file(
            io.BytesIO(xlsx_bytes),
            as_attachment=True,
            download_name=f"{base}_populated_template.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        log.exception("populate-template: %s", e)
        flash(f"Error: {e}")
        return redirect(url_for("workflows_page"))
    finally:
        if tmp_dir and Path(tmp_dir).exists():
            try:
                for fp in Path(tmp_dir).iterdir():
                    fp.unlink(missing_ok=True)
                Path(tmp_dir).rmdir()
            except OSError:
                pass


@app.route("/populate-template-from-fields", methods=["POST"])
def populate_template_from_fields_route():
    """fields.json + template .xlsx → filled template (same as: run.py populate-template-from-fields)."""
    if "fields_json" not in request.files or "template" not in request.files:
        flash("Upload fields.json and a template .xlsx.")
        return redirect(url_for("workflows_page"))
    fj = request.files["fields_json"]
    tpl_f = request.files["template"]
    if not fj.filename or not tpl_f.filename:
        flash("fields.json and template are required.")
        return redirect(url_for("workflows_page"))
    if not fj.filename.lower().endswith(".json"):
        flash("fields file must be .json.")
        return redirect(url_for("workflows_page"))
    if not tpl_f.filename.lower().endswith(".xlsx"):
        flash("Template must be a .xlsx file.")
        return redirect(url_for("workflows_page"))
    account_id = (request.form.get("account_id_fields") or "").strip() or None
    tmp_dir = None
    try:
        tmp_dir = tempfile.mkdtemp()
        fields_path = Path(tmp_dir) / "fields.json"
        tpl_path = Path(tmp_dir) / "template.xlsx"
        out_path = Path(tmp_dir) / "populated.xlsx"
        fj.save(str(fields_path))
        tpl_f.save(str(tpl_path))
        populate_template_from_fields_json(
            template_path=str(tpl_path),
            fields_json_path=str(fields_path),
            output_path=str(out_path),
            account_id=account_id,
        )
        if not out_path.exists():
            flash("Template population produced no file.")
            return redirect(url_for("workflows_page"))
        base = Path(tpl_f.filename).stem
        xlsx_bytes = out_path.read_bytes()
        return send_file(
            io.BytesIO(xlsx_bytes),
            as_attachment=True,
            download_name=f"{base}_from_fields.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        log.exception("populate-template-from-fields: %s", e)
        flash(f"Error: {e}")
        return redirect(url_for("workflows_page"))
    finally:
        if tmp_dir and Path(tmp_dir).exists():
            try:
                for f in Path(tmp_dir).iterdir():
                    f.unlink(missing_ok=True)
                Path(tmp_dir).rmdir()
            except OSError:
                pass


@app.route("/extract-fields", methods=["POST"])
def extract_fields_route():
    """PDF → fields.json (same as: run.py fields)."""
    from openpyxl import load_workbook
    from fields_from_qb_output import extract_fields as extract_fields_from_wb

    if "pdf" not in request.files:
        flash("No PDF uploaded.")
        return redirect(url_for("workflows_page"))
    pdf_f = request.files["pdf"]
    if not pdf_f.filename or not pdf_f.filename.lower().endswith(".pdf"):
        flash("Please upload a PDF.")
        return redirect(url_for("workflows_page"))
    tmp_dir = None
    try:
        tmp_dir = tempfile.mkdtemp()
        pdf_path = Path(tmp_dir) / "upload.pdf"
        pdf_f.save(str(pdf_path))
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            qb_tmp = f.name
        try:
            pdf_to_qb_excel(str(pdf_path), qb_tmp, overwrite=True)
            wb = load_workbook(qb_tmp, data_only=True, read_only=True)
            by_target = {
                name: [(name, [list(r) for r in wb[name].iter_rows(values_only=True)])]
                for name in wb.sheetnames
            }
            wb.close()
            fields = extract_fields_from_wb(by_target)
            payload = {"fields": [f.to_json() for f in fields]}
        finally:
            Path(qb_tmp).unlink(missing_ok=True)
        buf = io.BytesIO(json.dumps(payload, indent=2).encode("utf-8"))
        buf.seek(0)
        base = Path(pdf_f.filename).stem
        return send_file(
            buf,
            as_attachment=True,
            download_name=f"{base}_fields.json",
            mimetype="application/json",
        )
    except Exception as e:
        log.exception("extract-fields: %s", e)
        flash(f"Error: {e}")
        return redirect(url_for("workflows_page"))
    finally:
        if tmp_dir and Path(tmp_dir).exists():
            try:
                for f in Path(tmp_dir).iterdir():
                    f.unlink(missing_ok=True)
                Path(tmp_dir).rmdir()
            except OSError:
                pass


@app.route("/clean-json", methods=["POST"])
def clean_json_route():
    """Collapse repetitive sections in extraction JSON (same as: run.py clean-json)."""
    from extract_vl import _drop_repetitive_sections, pdf_phrase_count_for_file

    if "json_file" not in request.files:
        flash("No JSON file.")
        return redirect(url_for("workflows_page"))
    jf = request.files["json_file"]
    if not jf.filename or not jf.filename.lower().endswith(".json"):
        flash("Please upload a .json extraction file.")
        return redirect(url_for("workflows_page"))
    tmp_dir = None
    try:
        tmp_dir = tempfile.mkdtemp()
        json_path = Path(tmp_dir) / "input.json"
        jf.save(str(json_path))
        with open(json_path, encoding="utf-8") as f:
            payload = json.load(f)
        pdf_path = None
        if "pdf" in request.files and request.files["pdf"].filename:
            pdf_path = Path(tmp_dir) / "source.pdf"
            request.files["pdf"].save(str(pdf_path))
        if not pdf_path or not pdf_path.exists():
            meta = payload.get("meta") or {}
            if meta.get("pdf_path") and Path(meta["pdf_path"]).exists():
                pdf_path = Path(meta["pdf_path"])
            elif meta.get("pdf_name"):
                candidate = json_path.parent / meta["pdf_name"]
                if candidate.exists():
                    pdf_path = candidate
        pdf_phrase_count = None
        if pdf_path and pdf_path.exists():
            pdf_phrase_count = pdf_phrase_count_for_file(pdf_path)
        sections = payload.get("sections") or []
        if not sections:
            flash("No sections in JSON to clean.")
            return redirect(url_for("workflows_page"))
        payload["sections"] = _drop_repetitive_sections(sections, pdf_phrase_count=pdf_phrase_count)
        out_buf = io.BytesIO(json.dumps(payload, indent=2, ensure_ascii=False).encode("utf-8"))
        out_buf.seek(0)
        base = Path(jf.filename).stem
        return send_file(
            out_buf,
            as_attachment=True,
            download_name=f"{base}_cleaned.json",
            mimetype="application/json",
        )
    except Exception as e:
        log.exception("clean-json: %s", e)
        flash(f"Error: {e}")
        return redirect(url_for("workflows_page"))
    finally:
        if tmp_dir and Path(tmp_dir).exists():
            try:
                for f in Path(tmp_dir).iterdir():
                    f.unlink(missing_ok=True)
                Path(tmp_dir).rmdir()
            except OSError:
                pass


@app.route("/audit-standalone", methods=["POST"])
def audit_standalone_route():
    """PDF + extraction JSON → full audit report JSON (same as: run.py audit-json)."""
    if "pdf" not in request.files or "json_file" not in request.files:
        flash("Upload a PDF and the extraction JSON.")
        return redirect(url_for("workflows_page"))
    pdf_f = request.files["pdf"]
    jf = request.files["json_file"]
    if not pdf_f.filename or not jf.filename:
        flash("PDF and JSON are required.")
        return redirect(url_for("workflows_page"))
    if not pdf_f.filename.lower().endswith(".pdf"):
        flash("First file must be a PDF.")
        return redirect(url_for("workflows_page"))
    max_pages_raw = (request.form.get("audit_max_pages") or "").strip()
    max_pages = None
    if max_pages_raw.isdigit():
        max_pages = int(max_pages_raw)
    tmp_dir = None
    try:
        tmp_dir = tempfile.mkdtemp()
        pdf_path = Path(tmp_dir) / "source.pdf"
        json_path = Path(tmp_dir) / "extraction.json"
        pdf_f.save(str(pdf_path))
        jf.save(str(json_path))
        report = audit_pdf_vs_extraction_json(pdf_path, json_path, max_pages=max_pages)
        buf = io.BytesIO(json.dumps(report, indent=2, ensure_ascii=False).encode("utf-8"))
        buf.seek(0)
        base = Path(jf.filename).stem
        return send_file(
            buf,
            as_attachment=True,
            download_name=f"{base}_audit_report.json",
            mimetype="application/json",
        )
    except Exception as e:
        log.exception("audit-standalone: %s", e)
        flash(f"Error: {e}")
        return redirect(url_for("workflows_page"))
    finally:
        if tmp_dir and Path(tmp_dir).exists():
            try:
                for f in Path(tmp_dir).iterdir():
                    f.unlink(missing_ok=True)
                Path(tmp_dir).rmdir()
            except OSError:
                pass


if __name__ == "__main__":
    app.run(debug=True, port=5000)
