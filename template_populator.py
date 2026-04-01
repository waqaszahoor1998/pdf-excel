#!/usr/bin/env python3
"""
Populate a user-provided Excel template from QB-extracted workbook values.

Template-first flow:
  PDF -> qb workbook (existing pipeline) -> write values into template sheets.
"""

from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from qb_posting_rules import map_interest_from_purchases_sales


_ACCOUNT_RE = re.compile(r"\b(\d{3}-\d)\b")


def _num(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("$", "")
    if not s:
        return None
    if s.startswith("(") and s.endswith(")"):
        try:
            return -float(s[1:-1].strip())
        except Exception:
            return None
    try:
        return float(s)
    except Exception:
        return None


def _extract_source_label_map(ws) -> dict[str, list[Any]]:
    """From source PLSummary sheet: label in col A -> [B,C,D,E,F]."""
    out: dict[str, list[Any]] = {}
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if isinstance(label, str) and label.strip():
            out[label.strip()] = [ws.cell(r, c).value for c in range(2, 7)]
    return out


def _infer_account_id_from_source(wb) -> str | None:
    # Check PLSummary account row first (A4 in our builder)
    if "PLSummary" in wb.sheetnames:
        v = wb["PLSummary"].cell(4, 1).value
        if isinstance(v, str):
            m = _ACCOUNT_RE.search(v)
            if m:
                return m.group(1)
    # Search all sheets for "Portfolio No: ...366-3" patterns
    for name in wb.sheetnames:
        ws = wb[name]
        max_r = min(ws.max_row, 400)
        max_c = min(ws.max_column, 20)
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str):
                    m = _ACCOUNT_RE.search(v)
                    if m:
                        return m.group(1)
    return None


def _infer_period_end_from_source(wb) -> date | None:
    """
    Infer statement period-end date from source workbook text.
    """
    month_map = {
        "jan": 1, "january": 1,
        "feb": 2, "february": 2,
        "mar": 3, "march": 3,
        "apr": 4, "april": 4,
        "may": 5,
        "jun": 6, "june": 6,
        "jul": 7, "july": 7,
        "aug": 8, "august": 8,
        "sep": 9, "sept": 9, "september": 9,
        "oct": 10, "october": 10,
        "nov": 11, "november": 11,
        "dec": 12, "december": 12,
    }
    pat_long = re.compile(r"period ended\s+([A-Za-z]+)\s+(\d{1,2}),\s*(\d{4})", re.I)
    pat_short = re.compile(r"(?:as of|ended)\s+([A-Za-z]{3,9})\s+(\d{1,2})\s+(\d{2,4})", re.I)

    for name in wb.sheetnames:
        ws = wb[name]
        max_r = min(ws.max_row, 250)
        max_c = min(ws.max_column, 20)
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                v = ws.cell(r, c).value
                if not isinstance(v, str):
                    continue
                s = v.strip()
                m = pat_long.search(s)
                if m:
                    mon = month_map.get(m.group(1).lower())
                    if mon:
                        return date(int(m.group(3)), mon, int(m.group(2)))
                m2 = pat_short.search(s)
                if m2:
                    mon = month_map.get(m2.group(1).lower())
                    if mon:
                        y = int(m2.group(3))
                        if y < 100:
                            y += 2000
                        return date(y, mon, int(m2.group(2)))
    return None


def _find_account_block_start_col(ws, account_id: str) -> int | None:
    """In PLSummary Admin GS row 1, find account marker col and convert to block start col."""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, str) and v.strip() == account_id:
            # block starts 6 cols before account marker (template layout)
            return c - 6
    return None


def _extract_holdings_totals(wb) -> tuple[float | None, float | None, float | None, float | None]:
    """
    Parse Holdings TOTAL GS row if present.
    Returns (market_value, original_cost, unrealized_gain_loss, accrued_income).
    """
    if "Holdings" not in wb.sheetnames:
        return (None, None, None, None)
    ws = wb["Holdings"]
    max_r = min(ws.max_row, 800)
    max_c = min(ws.max_column, 20)
    for r in range(1, max_r + 1):
        label = ws.cell(r, 1).value
        if isinstance(label, str) and "total gs" in label.lower():
            nums = []
            for c in range(1, max_c + 1):
                n = _num(ws.cell(r, c).value)
                if n is not None:
                    nums.append(n)
            # often next row has accrued-income-only total
            accrued_income = None
            if r + 1 <= max_r:
                nums_next = []
                for c in range(1, max_c + 1):
                    n = _num(ws.cell(r + 1, c).value)
                    if n is not None:
                        nums_next.append(n)
                if len(nums_next) == 1:
                    accrued_income = nums_next[0]

            if len(nums) >= 3:
                market = nums[0]
                original_cost = nums[1]
                unreal = nums[2]
                return (market, original_cost, unreal, accrued_income)
    return (None, None, None, None)


def _to_date(v: Any) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        s = v.strip()
        # mm/dd/yyyy
        m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
        if m:
            return date(int(m.group(3)), int(m.group(1)), int(m.group(2)))
        # yyyy-mm-dd
        m2 = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
        if m2:
            return date(int(m2.group(1)), int(m2.group(2)), int(m2.group(3)))
    return None


def _find_latest_date_col_in_block(ws, account_row: int, lookahead_rows: int = 5) -> int | None:
    best: tuple[date, int] | None = None
    for rr in range(account_row, min(account_row + lookahead_rows, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            d = _to_date(ws.cell(rr, c).value)
            if d is None:
                continue
            if best is None or d > best[0]:
                best = (d, c)
    return best[1] if best else None


def _find_date_cols_in_block(ws, account_row: int, lookahead_rows: int = 5) -> list[tuple[date, int]]:
    seen_cols = set()
    out: list[tuple[date, int]] = []
    for rr in range(account_row, min(account_row + lookahead_rows, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            d = _to_date(ws.cell(rr, c).value)
            if d is None or c in seen_cols:
                continue
            seen_cols.add(c)
            out.append((d, c))
    out.sort(key=lambda x: x[0])
    return out


def _find_period_date_col_in_block(
    ws, account_row: int, period_end: date | None, lookahead_rows: int = 5
) -> tuple[int | None, int | None]:
    date_cols = _find_date_cols_in_block(ws, account_row, lookahead_rows=lookahead_rows)
    if not date_cols:
        return (None, None)

    target_idx = None
    if period_end is not None:
        for i, (d, _c) in enumerate(date_cols):
            if d == period_end:
                target_idx = i
                break
        if target_idx is None:
            le = [i for i, (d, _c) in enumerate(date_cols) if d <= period_end]
            if le:
                target_idx = le[-1]
    if target_idx is None:
        target_idx = len(date_cols) - 1

    target_col = date_cols[target_idx][1]
    prev_col = date_cols[target_idx - 1][1] if target_idx > 0 else None
    return (target_col, prev_col)


def _find_row_by_label_near(ws, start_row: int, label: str, window: int = 30) -> int | None:
    target = label.strip().lower()
    for r in range(start_row, min(start_row + window, ws.max_row) + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().lower() == target:
            return r
    return None


def _normalize_src_map_from_json(d: Any) -> dict[str, list[Any]]:
    """Ensure plsummary.source_map JSON becomes label -> [5 cols] like the live PLSummary sheet."""
    if not isinstance(d, dict):
        return {}
    out: dict[str, list[Any]] = {}
    for k, v in d.items():
        if not isinstance(k, str) or not k.strip():
            continue
        if isinstance(v, list):
            lst = list(v[:5])
            while len(lst) < 5:
                lst.append(None)
            out[k.strip()] = lst
        else:
            out[k.strip()] = [v, None, None, None, None]
    return out


def _first_src_val(row: list[Any] | None) -> float | None:
    if not row:
        return None
    return _num(row[0]) if len(row) > 0 else None


def _merge_optional_delta_fields_from_json(
    src_map: dict[str, list[Any]], by_key: dict[str, dict], get_num_fn
) -> None:
    """Fill change rows from delta.* fields when missing in plsummary.source_map."""
    pairs = (
        ("Change in unrealized gain/loss", "delta.change_in_unrealized_gain_loss"),
        ("Change in Accrued Interest", "delta.change_in_accrued_interest"),
        ("Change in Accrued Dividend", "delta.change_in_accrued_dividend"),
    )
    for label, fld in pairs:
        v = get_num_fn(fld)
        if v is None:
            continue
        if label not in src_map or _first_src_val(src_map.get(label)) is None:
            src_map[label] = [float(v), None, None, None, None]


def _apply_template_fills(
    wb_tpl,
    inferred_acct: str,
    period_end: date | None,
    src_map: dict[str, list[Any]],
    market_value: float | None,
    original_cost: float | None,
    unrealized_total: float | None,
    accrued_income: float | None,
) -> None:
    """Write template sheets from a PLSummary-style label map + holdings numbers."""
    if "PLSummary Admin GS " not in wb_tpl.sheetnames:
        raise ValueError("Template has no 'PLSummary Admin GS ' sheet")
    ws = wb_tpl["PLSummary Admin GS "]
    start_col = _find_account_block_start_col(ws, inferred_acct)
    if start_col is None:
        raise ValueError(f"Could not find account block '{inferred_acct}' in template PLSummary Admin GS")

    label_map = {
        "Accrued Interest and dividend": "Accrued Dividend and Interest",
        "Interest expense": "Interest Paid on Purchases",
        "Account Fees": "Account Fee",
        "Broker fee": "Broker fees",
    }
    full_width_labels = {"Investments", "Cash and cash equivalents", "Totals"}

    for r in range(1, ws.max_row + 1):
        lbl = ws.cell(r, start_col).value
        if not isinstance(lbl, str) or not lbl.strip():
            continue
        t_label = lbl.strip()
        s_label = label_map.get(t_label, t_label)
        if s_label not in src_map:
            continue
        vals = src_map[s_label]
        if t_label in full_width_labels:
            for i, v in enumerate(vals, start=1):
                ws.cell(r, start_col + i).value = v
        else:
            ws.cell(r, start_col + 1).value = vals[0]

    if market_value is not None:
        r_inv = _find_row_by_label_near(ws, 28, "Investment in Securities", 20)
        if r_inv:
            ws.cell(r_inv, start_col + 1).value = market_value

    def _set_admin_block_value(row_label: str, value: float | None) -> None:
        if value is None:
            return
        for rr in range(1, ws.max_row + 1):
            v = ws.cell(rr, start_col).value
            if isinstance(v, str) and v.strip().lower() == row_label.strip().lower():
                ws.cell(rr, start_col + 1).value = float(value)
                return

    unrealized_change_written: float | None = None
    if "Unrealized" in wb_tpl.sheetnames:
        wsu = wb_tpl["Unrealized"]
        acct_row = None
        for r in range(1, wsu.max_row + 1):
            v = wsu.cell(r, 1).value
            if isinstance(v, str) and v.strip() == inferred_acct:
                acct_row = r
                break
        if acct_row:
            date_col, prev_col = _find_period_date_col_in_block(wsu, acct_row, period_end, lookahead_rows=5)
            if date_col:
                r_cost = _find_row_by_label_near(wsu, acct_row, "Cost", 20)
                r_mv = _find_row_by_label_near(wsu, acct_row, "Market Value", 20)
                r_un = _find_row_by_label_near(wsu, acct_row, "Unrealized", 20)
                r_ch = _find_row_by_label_near(wsu, acct_row, "Change in Unrealized", 25)
                if r_cost and original_cost is not None:
                    wsu.cell(r_cost, date_col).value = original_cost
                if r_mv and market_value is not None:
                    wsu.cell(r_mv, date_col).value = market_value
                if r_un and unrealized_total is not None:
                    wsu.cell(r_un, date_col).value = unrealized_total
                if r_ch:
                    src_change_un = src_map.get("Change in unrealized gain/loss", [None])[0]
                    if src_change_un is not None and abs(float(src_change_un)) > 1e-9:
                        wsu.cell(r_ch, date_col).value = float(src_change_un)
                        unrealized_change_written = float(src_change_un)
                    elif r_un and unrealized_total is not None and prev_col is not None:
                        prev = _num(wsu.cell(r_un, prev_col).value)
                        if prev is not None:
                            unrealized_change_written = float(unrealized_total - prev)
                            wsu.cell(r_ch, date_col).value = unrealized_change_written
                    elif r_un and unrealized_total is not None and prev_col is None:
                        unrealized_change_written = float(unrealized_total)
                        wsu.cell(r_ch, date_col).value = unrealized_change_written

    accrued_interest_change_written: float | None = None
    if "Change in Interest" in wb_tpl.sheetnames:
        wsi = wb_tpl["Change in Interest"]
        acct_row = None
        for r in range(1, wsi.max_row + 1):
            v = wsi.cell(r, 1).value
            if isinstance(v, str) and v.strip() == inferred_acct:
                acct_row = r
                break
        if acct_row:
            date_col, prev_col = _find_period_date_col_in_block(wsi, acct_row, period_end, lookahead_rows=5)
            if date_col:
                r_ai = _find_row_by_label_near(wsi, acct_row, "Accrued Interest", 25)
                r_ci = _find_row_by_label_near(wsi, acct_row, "Change in Accrued Interest", 25)
                if r_ai and accrued_income is not None:
                    wsi.cell(r_ai, date_col).value = accrued_income
                if r_ci:
                    src_ci = src_map.get("Change in Accrued Interest", [None])[0]
                    if src_ci is not None and abs(float(src_ci)) > 1e-9:
                        wsi.cell(r_ci, date_col).value = float(src_ci)
                        accrued_interest_change_written = float(src_ci)
                    elif r_ai and accrued_income is not None and prev_col is not None:
                        prev = _num(wsi.cell(r_ai, prev_col).value)
                        if prev is not None:
                            accrued_interest_change_written = float(accrued_income - prev)
                            wsi.cell(r_ci, date_col).value = accrued_interest_change_written

    accrued_dividend_change_written: float | None = None
    if "Change in Dividend" in wb_tpl.sheetnames:
        wsd = wb_tpl["Change in Dividend"]
        acct_row = None
        for r in range(1, wsd.max_row + 1):
            v = wsd.cell(r, 1).value
            if isinstance(v, str) and v.strip() == inferred_acct:
                acct_row = r
                break
        if acct_row:
            date_col, _prev_col = _find_period_date_col_in_block(wsd, acct_row, period_end, lookahead_rows=5)
            if date_col:
                r_cd = _find_row_by_label_near(wsd, acct_row, "Change in Accrued Dividend", 25)
                if r_cd:
                    src_cd = src_map.get("Change in Accrued Dividend", [None])[0]
                    if src_cd is not None:
                        wsd.cell(r_cd, date_col).value = float(src_cd)
                        accrued_dividend_change_written = float(src_cd)

    if unrealized_change_written is not None:
        _set_admin_block_value("Change in unrealized gain/loss", unrealized_change_written)
    if accrued_interest_change_written is not None:
        _set_admin_block_value("Change in Accrued Interest", accrued_interest_change_written)
    if accrued_dividend_change_written is not None:
        _set_admin_block_value("Change in Accrued Dividend", accrued_dividend_change_written)


def populate_template_from_qb_output(
    template_path: str | Path,
    qb_output_xlsx: str | Path,
    output_path: str | Path,
    account_id: str | None = None,
) -> str:
    wb_src = load_workbook(qb_output_xlsx, data_only=True)
    # PLSummary is preferred (it standardizes many values), but not every PDF/QB build
    # yields a PLSummary sheet. For unattended service behavior we must not crash; we
    # populate what we can (holdings totals, dates, account id) and leave the rest blank.
    if "PLSummary" in wb_src.sheetnames:
        src_map = _extract_source_label_map(wb_src["PLSummary"])
    else:
        src_map = {}
    inferred_acct = account_id or _infer_account_id_from_source(wb_src) or "366-3"
    period_end = _infer_period_end_from_source(wb_src)
    market_value, original_cost, unrealized_total, accrued_income = _extract_holdings_totals(wb_src)

    wb_tpl = load_workbook(template_path)
    _apply_template_fills(
        wb_tpl,
        inferred_acct,
        period_end,
        src_map,
        market_value,
        original_cost,
        unrealized_total,
        accrued_income,
    )

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb_tpl.save(out)
    return str(out)


def populate_template_from_fields_json(
    template_path: str | Path,
    fields_json_path: str | Path,
    output_path: str | Path,
    account_id: str | None = None,
) -> str:
    """
    Populate the template using only fields.json (service output).

    Expects `plsummary.source_map` (from `fields` command) for parity with
    `populate-template`; otherwise falls back to cash + purchases/sales interest only.
    """
    payload = json.loads(Path(fields_json_path).read_text(encoding="utf-8"))
    raw_fields = payload.get("fields", [])
    by_key: dict[str, dict] = {}
    for f in raw_fields:
        k = f.get("key")
        if isinstance(k, str) and k:
            by_key[k] = f

    def _get_num(key: str) -> float | None:
        f = by_key.get(key)
        if not f:
            return None
        return _num(f.get("value"))

    def _parse_period_end() -> date | None:
        f = by_key.get("statement.period_end")
        if not f:
            return None
        val = f.get("value")
        if isinstance(val, str):
            d = _to_date(val)
            if d is not None:
                return d
            try:
                return date.fromisoformat(val.strip()[:10])
            except Exception:
                return None
        return None

    acc_f = by_key.get("account.id")
    acc_str = acc_f.get("value") if isinstance(acc_f, dict) else None
    if isinstance(acc_str, str) and acc_str.strip():
        inferred_acct = (account_id or acc_str.strip())
    else:
        inferred_acct = account_id or "366-3"

    period_end = _parse_period_end()

    src_map: dict[str, list[Any]] = {}
    sm_field = by_key.get("plsummary.source_map")
    if sm_field and isinstance(sm_field.get("value"), dict):
        src_map = _normalize_src_map_from_json(sm_field["value"])
        _merge_optional_delta_fields_from_json(src_map, by_key, _get_num)
    else:
        cash_eom = _get_num("cash.closing_balance")
        ps_total = _get_num("purchases_sales.total_accrued_interest")
        if cash_eom is not None:
            src_map["Cash and cash equivalents"] = [None, None, None, None, float(cash_eom)]
        if ps_total is not None:
            ip, ii = map_interest_from_purchases_sales(ps_total)
            src_map["Interest Paid on Purchases"] = [float(ip), None, None, None, None]
            src_map["Interest Income"] = [float(ii), None, None, None, None]
        _merge_optional_delta_fields_from_json(src_map, by_key, _get_num)

    market_value = _get_num("holdings.market_value")
    original_cost = _get_num("holdings.original_cost")
    unrealized_total = _get_num("holdings.unrealized_gain_loss")
    accrued_income = _get_num("holdings.accrued_income")

    wb_tpl = load_workbook(template_path)
    _apply_template_fills(
        wb_tpl,
        inferred_acct,
        period_end,
        src_map,
        market_value,
        original_cost,
        unrealized_total,
        accrued_income,
    )

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb_tpl.save(out)
    return str(out)

