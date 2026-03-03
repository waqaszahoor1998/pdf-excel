#!/usr/bin/env python3
"""
PDF → Excel via Anthropic (Claude) API.

Behaviour is config-driven: limits, model, prompts, single vs multi-table,
and optional structured output come from config file and environment (see config.py).
"""

import argparse
import base64
import csv
import io
import json
import logging
import os
import sys
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(message)s")
log = logging.getLogger(__name__)

import anthropic
from dotenv import load_dotenv
from openpyxl import Workbook

load_dotenv()

from config import CONFIG_DIR, get_system_prompt_path, load_config


def _load_prompt(path: Path) -> str:
    return path.read_text(encoding="utf-8").strip()


def load_pdf_base64(path: str, max_bytes: int) -> str:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"PDF not found: {path}")
    if path.suffix.lower() != ".pdf":
        raise ValueError("File must be a PDF")
    data = path.read_bytes()
    if len(data) > max_bytes:
        raise ValueError(f"PDF too large (max {max_bytes // (1024*1024)}MB)")
    return base64.standard_b64encode(data).decode("utf-8")


def extract_single_csv_from_response(text: str) -> str:
    """Return the first CSV block content."""
    s = text.strip()
    begin, end = "---BEGIN CSV---", "---END CSV---"
    i, j = s.find(begin), s.find(end)
    if i != -1 and j != -1 and j > i:
        return s[i + len(begin) : j].strip()
    for marker in ("```csv", "```CSV", "```"):
        if marker in s:
            start = s.find(marker) + len(marker)
            rest = s[start:].strip()
            end_m = rest.find("```")
            if end_m != -1:
                return rest[:end_m].strip()
            return rest
    lines = [ln.strip() for ln in s.splitlines() if ln.strip()]
    for k, line in enumerate(lines):
        if "," in line and k + 1 <= len(lines):
            return "\n".join(lines[k:])
    raise ValueError("No CSV block found in model response. Response was: " + text[:500])


# Backward compatibility for tests
extract_csv_from_response = extract_single_csv_from_response


def _parse_one_csv_block(block: str) -> tuple[str | None, list[list[str]]]:
    """Parse one CSV block; first line may be 'SheetName: X'. Returns (sheet_name, rows)."""
    block = block.strip()
    sheet_name = None
    if block.lower().startswith("sheetname:"):
        first_line, _, rest = block.partition("\n")
        sheet_name = first_line[10:].strip()
        block = rest.strip()
    reader = csv.reader(io.StringIO(block))
    rows = list(reader)
    if not rows:
        return sheet_name, []
    return sheet_name, rows


def extract_all_csv_blocks_from_response(text: str) -> list[tuple[str | None, list[list[str]]]]:
    """Return list of (sheet_name, rows) for each CSV block."""
    out = []
    begin, end = "---BEGIN CSV---", "---END CSV---"
    start = 0
    while True:
        i = text.find(begin, start)
        j = text.find(end, i) if i != -1 else -1
        if i == -1 or j == -1 or j <= i:
            break
        raw = text[i + len(begin) : j].strip()
        name, rows = _parse_one_csv_block(raw)
        out.append((name, rows))
        start = j + len(end)
    if not out:
        single = extract_single_csv_from_response(text)
        _, rows = _parse_one_csv_block(single)
        out = [(None, rows)]
    return out


def csv_to_excel(csv_content: str, out_path: str, sheet_name: str = "Extracted") -> None:
    reader = csv.reader(io.StringIO(csv_content))
    rows = list(reader)
    if not rows:
        raise ValueError("CSV has no rows")
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] if sheet_name else "Extracted"
    for r in rows:
        ws.append(r)
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _sanitize_sheet_name(name: str | None, fallback: str) -> str:
    if not name or not name.strip():
        return fallback
    s = (
        name.replace("\\", " ")
        .replace("/", " ")
        .replace("*", " ")
        .replace("?", " ")
        .replace("[", " ")
        .replace("]", " ")[:31]
        .strip()
    )
    return s or fallback


def tables_to_excel(tables: list[tuple[str | None, list[list[str]]]], out_path: str) -> None:
    """Write multiple tables to one Excel file, one sheet per table."""
    if not tables:
        raise ValueError("No tables to write")
    wb = Workbook()
    wb.remove(wb.active)
    for idx, (name, rows) in enumerate(tables):
        title = _sanitize_sheet_name(name, f"Sheet_{idx + 1}")
        ws = wb.create_sheet(title=title)
        for r in rows:
            ws.append(r if isinstance(r, list) else [r])
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def json_sections_to_excel(sections: list[dict], out_path: str) -> None:
    """Write structured sections (name, headers, rows) to Excel, one sheet per section."""
    if not sections:
        raise ValueError("No sections to write")
    tables = []
    for sec in sections:
        name = sec.get("name")
        headers = sec.get("headers") or []
        rows = sec.get("rows") or []
        rows = [headers] + list(rows) if headers else list(rows)
        tables.append((name, rows))
    tables_to_excel(tables, out_path)


def _call_api(
    client: anthropic.Anthropic,
    pdf_b64: str,
    user_text: str,
    system_prompt: str,
    model: str,
    config: dict,
) -> str:
    """Single API call; returns response text. Optional output_config from config."""
    user_content = [
        {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_b64}},
        {"type": "text", "text": user_text},
    ]
    kwargs = {
        "model": model,
        "max_tokens": 8192,
        "system": system_prompt,
        "messages": [{"role": "user", "content": user_content}],
    }
    if config.get("use_structured_output"):
        schema_path = config.get("structured_schema_path")
        if schema_path:
            path = Path(schema_path)
            if not path.is_absolute():
                path = CONFIG_DIR / path
            if path.exists():
                schema = json.loads(path.read_text(encoding="utf-8"))
                kwargs["output_config"] = {"format": {"type": "json_schema", "schema": schema}}
    message = client.messages.create(**kwargs)
    response_text = ""
    for block in message.content:
        if hasattr(block, "text"):
            response_text += block.text
    return response_text


def extract_pdf_to_excel(
    pdf_path: str,
    user_query: str,
    output_path: str,
    api_key: str | None = None,
    model: str | None = None,
    config: dict | None = None,
) -> str:
    """
    Extract data from PDF per user query using Anthropic API and save as Excel.
    All limits and behaviour come from config (and env). Returns the path to the saved Excel file.
    """
    cfg = config or load_config()
    api_key = api_key or os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise ValueError("Set ANTHROPIC_API_KEY in .env or pass api_key=...")
    model = model or cfg.get("default_model") or os.environ.get("ANTHROPIC_MODEL")
    if not model:
        raise ValueError("Set ANTHROPIC_MODEL in .env or default_model in config")

    max_bytes = int(cfg.get("max_pdf_bytes", 32 * 1024 * 1024))
    query_max = int(cfg.get("query_max_length", 8000))
    query_text = (user_query[:query_max] + "...") if len(user_query) > query_max else user_query

    prompt_path = get_system_prompt_path(cfg)
    if not prompt_path or not prompt_path.exists():
        raise ValueError(
            "No system prompt file found. Set SYSTEM_PROMPT_PATH or add prompts/extraction_single.txt (and extraction_all.txt) in prompts_dir."
        )
    system_prompt = _load_prompt(prompt_path)

    pdf_b64 = load_pdf_base64(pdf_path, max_bytes)
    log.info("Calling API…")
    client = anthropic.Anthropic(api_key=api_key)
    extraction_mode = (cfg.get("extraction_mode") or "single").strip().lower()

    # Optional: long-PDF structure step
    if cfg.get("long_pdf_enabled"):
        structure_path = Path(cfg.get("prompts_dir") or "").strip() or Path(__file__).resolve().parent / "prompts"
        if not structure_path.is_absolute():
            structure_path = Path(__file__).resolve().parent / structure_path
        structure_file = structure_path / "structure.txt"
        if structure_file.exists():
            structure_prompt = _load_prompt(structure_file)
            structure_text = _call_api(client, pdf_b64, structure_prompt, structure_prompt, model, cfg)
            user_text = f"Document structure (for context):\n{structure_text[:2000]}\n\nUser request: Extract the following from this PDF and return only the CSV block(s) as specified:\n\n{query_text}"
        else:
            user_text = f"Extract the following from this PDF and return only the CSV block(s) as specified:\n\n{query_text}"
    else:
        user_text = f"Extract the following from this PDF and return only the CSV block(s) as specified:\n\n{query_text}"

    response_text = _call_api(client, pdf_b64, user_text, system_prompt, model, cfg)

    if cfg.get("use_structured_output") and cfg.get("structured_schema_path"):
        try:
            data = json.loads(response_text)
            tables_data = data.get("tables") or []
            if tables_data:
                tables = []
                for t in tables_data:
                    name = t.get("name")
                    headers = t.get("headers") or []
                    rows = t.get("rows") or []
                    tables.append((name, [headers] + rows if headers else rows))
                tables_to_excel(tables, output_path)
                log.info("Done.")
                return output_path
        except (json.JSONDecodeError, KeyError):
            pass

    # CSV path
    if extraction_mode == "all":
        blocks = extract_all_csv_blocks_from_response(response_text)
        tables_to_excel(blocks, output_path)
    else:
        csv_content = extract_single_csv_from_response(response_text)
        csv_to_excel(csv_content, output_path)
    log.info("Done.")
    return output_path


def main() -> int:
    cfg = load_config()
    default_model = cfg.get("default_model") or os.environ.get("ANTHROPIC_MODEL") or ""
    parser = argparse.ArgumentParser(
        description="Extract data from a PDF using a natural-language query and save to Excel (via Anthropic API)."
    )
    parser.add_argument("pdf", help="Path to the PDF file")
    parser.add_argument("query", help="What to extract, e.g. 'company taxes for January 2026'")
    parser.add_argument("-o", "--output", default=None, help="Output Excel path")
    parser.add_argument(
        "--model",
        default=default_model,
        help="Anthropic model (or set ANTHROPIC_MODEL in .env / default_model in config)",
    )
    parser.add_argument("--config", default=None, help="Path to extract config JSON (optional)")
    args = parser.parse_args()

    config_path = args.config or os.environ.get("EXTRACT_CONFIG_PATH")
    config = load_config(config_path) if config_path else load_config()

    try:
        out = args.output or str(Path(args.pdf).with_suffix(".xlsx"))
        log.info("PDF: %s | Query: %s | Output: %s", args.pdf, args.query[:50] + "..." if len(args.query) > 50 else args.query, out)
        result = extract_pdf_to_excel(args.pdf, args.query, out, model=args.model or None, config=config)
        print(f"Saved: {result}")
        return 0
    except (FileNotFoundError, ValueError) as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1
    except anthropic.APIError as e:
        msg = str(e)
        if "authentication" in msg.lower() or "invalid_api_key" in msg.lower() or "401" in msg:
            print("Error: Invalid or missing API key. Set ANTHROPIC_API_KEY in .env.", file=sys.stderr)
        elif "rate" in msg.lower() or "429" in msg:
            print("Error: API rate limit exceeded. Try again later.", file=sys.stderr)
        elif "overloaded" in msg.lower() or "503" in msg:
            print("Error: API temporarily unavailable. Try again later.", file=sys.stderr)
        else:
            print(f"Error: API request failed. {msg}", file=sys.stderr)
        return 1
    except Exception as e:
        log.exception("Extract failed")
        print(f"Error: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
