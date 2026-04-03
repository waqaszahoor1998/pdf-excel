#!/usr/bin/env python3
"""
Transform extracted PDF tables into QB Automation Sheet format.

Pipeline: PDF → (tables_to_excel) → raw xlsx → (this module) → QB-format xlsx.

- Groups sheets by target name (e.g. all "Asset Allocation" into one sheet).
- Uses standard QB-style sheet names (see docs/EXPECTED_FORMAT.md).
- Preserves table structure; merges duplicate section types into one sheet per type.
"""

import logging
import re
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook

log = logging.getLogger(__name__)
from openpyxl.styles import Font, PatternFill

from config import load_qb_cleanup_config
from tables_to_excel import _cell_value as _normalize_cell_value
from tables_to_excel import _merge_fragmented_row

# QB-style colors (from sample workbook – see color scan below)
# GREEN 92D050: section separator row; first row of each block; cells that contain account IDs (902-7, 1004, E79271004)
# YELLOW FFFF00: formula cells (=SUM(...)) and "Checks" label
# ORANGE FFC000: emphasis (e.g. Net Assets subtotal area)
FILL_SECTION_HEADER = PatternFill(fill_type="solid", fgColor="92D050")   # Light green
FILL_TABLE_HEADER = PatternFill(fill_type="solid", fgColor="D9E1F2")    # Light blue – column header row
FILL_FORMULA = PatternFill(fill_type="solid", fgColor="FFFF00")         # Yellow – formulas / check cells
FILL_TOTALS = PatternFill(fill_type="solid", fgColor="FFC000")          # Orange – totals / emphasis

# Section titles that start a new block in per-account sheets (first cell match, case-insensitive)
_PER_ACCOUNT_SECTION_PATTERNS = re.compile(
    r"^(Account\s+Summary|Consolidated\s+Summary|Asset\s+Allocation|Portfolio\s+Activity|"
    r"Tax\s+Summary|Cost\s+Summary|Cash\s+[&]\s+Fixed\s+Income|Equity\s+Summary|Equity\s+Detail|"
    r"For\s+the\s+Period|Holdings|Interest\s+on\s+USD\s+Cash)$",
    re.I,
)


def _is_section_title_row(row: list) -> bool:
    """True if row's first non-empty cell looks like a section title."""
    first = next((str(c).strip() for c in (row or []) if c is not None and str(c).strip()), None)
    return bool(first and _PER_ACCOUNT_SECTION_PATTERNS.match(first))


def _is_continued_or_page_header_row(row: list) -> bool:
    """True if row is just 'CONTINUED' or a page-break marker like '— ABC TR ACCT. G41269004_11 —'."""
    first = next((str(c).strip() for c in (row or []) if c is not None and str(c).strip()), None)
    if not first:
        return True
    if first.upper() == "CONTINUED":
        return True
    if first.startswith("—") and first.endswith("—") and "_" in first:
        return True
    return False


def _merge_address_fragments_in_row(row: list) -> list:
    """
    Merge adjacent cells that look like one address or phrase split across cells
    (e.g. '383 Madison Avenue, New Yor' + 'k, NY 10179' -> one cell).
    """
    if not row or len(row) < 2:
        return list(row) if row else []
    out = []
    i = 0
    while i < len(row):
        cell = row[i]
        s = (str(cell).strip() if cell is not None else "").strip()
        if not s:
            out.append(cell)
            i += 1
            continue
        # Look for continuation in next cell only when first cell looks like address (long, or has comma)
        if i + 1 < len(row):
            next_cell = row[i + 1]
            next_s = (str(next_cell).strip() if next_cell is not None else "").strip()
            if (
                next_s
                and len(next_s) <= 20
                and len(s) >= 10
                and s[-1].isalpha()
                and next_s[0].isalpha()
            ):
                # No space for word completion ("Yor"+"k") or address continuation (next has comma); space for "Name"+"Role"
                sep = "" if len(next_s) <= 4 or "," in next_s else " "
                out.append(s + sep + next_s)
                i += 2
                continue
        out.append(cell)
        i += 1
    return out


def _split_rows_by_section(rows: list[list]) -> list[tuple[str, list[list]]]:
    """Split rows into [(section_name, sub_rows), ...] by section title. Drops CONTINUED/page-header rows from sub_rows."""
    if not rows:
        return []
    result = []
    current_section = "Section"
    current_rows = []
    for row in rows:
        if _is_continued_or_page_header_row(row):
            continue
        if _is_section_title_row(row):
            if current_rows:
                result.append((current_section, current_rows))
            first = next((str(c).strip() for c in (row or []) if c is not None and str(c).strip()), "")
            current_section = first or "Section"
            current_rows = [row]
        else:
            current_rows.append(row)
    if current_rows:
        result.append((current_section, current_rows))
    return result


def _target_sheet_name(source_name: str) -> str:
    """
    Normalize extracted sheet name for grouping. No fixed list: use document names.
    Only collapse to 'Other' when the name looks like a data row (e.g. long, starts with digit/coupon).
    """
    s = (source_name or "").strip()
    if not s:
        return "Other"
    # Strip trailing " 2", " 3", "_1", "_11" etc.
    base = re.sub(r"\s+\d+$", "", s)
    base = re.sub(r"_\d+$", "", base)
    # Merge "Page1", "Page2", ... into one sheet
    if re.match(r"^Page\d+$", base, re.I):
        return "By Page"
    # "PLSummary J.P. Morgan Chase" or similar
    if "plsummary" in base.lower() or "mtd pnl" in s.lower():
        return "PLSummary"
    # Period-covering headings often become the inferred sheet name; treat as Period Summary.
    if base.lower().startswith("for the period"):
        return "Period Summary"
    # Per-account pattern: "ABC TRUST ACCT. E79271004_1" -> "Account E79271004"
    m = re.search(r"ACCT\.\s*([A-Z0-9]+)", s, re.I)
    if m:
        return f"Account {m.group(1)}"
    # Broker header
    if "JPMorgan" in s or "J.P. Morgan" in s or "Chase Bank" in s:
        return "Broker Info"
    # Use document name unless it looks like a detail row (bond name, coupon, long string) -> group into Other
    if len(base) > 50:
        return "Other"
    if re.match(r"^\d", base) or re.match(r"^0\.\d", base):
        return "Other"
    return base


def _rows_from_sheet(ws) -> list[list]:
    """Read all rows from a worksheet as list of lists (cell values)."""
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _safe_sheet_name(name: str, max_len: int = 31) -> str:
    """Excel sheet name: max 31 chars (Excel limit), no \\ / * ? [ ]"""
    s = (name or "Sheet").replace("\\", "").replace("/", "").replace("*", "").replace("?", "").replace("[", "").replace("]", "")
    s = (s[:max_len] if s else "Sheet").strip() or "Sheet"
    return s[:31]  # enforce Excel limit even if max_len was overridden


def _fill_row(ws, row: int, num_cols: int, fill: PatternFill) -> None:
    """Apply a fill to columns 1 through num_cols in the given row."""
    if fill is None:
        return
    for c in range(1, num_cols + 1):
        ws.cell(row=row, column=c).fill = fill


def _looks_like_header_row(row: list) -> bool:
    """True if row looks like a table header (short text, no big numbers)."""
    if not row:
        return False
    cells = [str(c).strip() for c in (row if isinstance(row, (list, tuple)) else [row]) if c is not None]
    if len(cells) < 2:
        return False
    header_like = ("account", "market", "value", "bom", "eom", "mtd", "cash", "number", "date", "description", "debit", "credit")
    return any(h in " ".join(cells).lower() for h in header_like)


def _is_account_id(val) -> bool:
    """True if value looks like an account ID (e.g. 902-7, 1004, E79271004, 9511)."""
    if val is None:
        return False
    s = str(val).strip()
    if not s or len(s) > 20:
        return False
    # Digits with optional dash and digits (902-7, 1004, 9511)
    if re.match(r"^\d{3,}(-\d+)?$", s):
        return True
    # Alphanumeric like E79271004, G41269004
    if re.match(r"^[A-Z]?\d{6,}$", s, re.I):
        return True
    return False


def _is_formula_or_check(val) -> bool:
    """True if cell is a formula or 'Checks' label (yellow in sample)."""
    if val is None:
        return False
    s = str(val).strip()
    return s.startswith("=") or s.lower() == "checks"


def _is_totals_row(row: list) -> bool:
    """True if first cell looks like a Totals row."""
    if not row:
        return False
    first = (row[0] if isinstance(row, (list, tuple)) else row)
    s = (str(first).strip().lower() if first is not None else "")
    return s in ("total", "totals", "total value", "total assets", "total liabilities")


# Phrases that mark a row as disclaimer/prose (not table data). Keep lowercase for matching.
_PROSE_PHRASES = (
    "this page intentionally left blank",
    "form 1099",
    "1099 tax",
    "available online",
    "please see disclosure",
    "not fdic insured",
    "not insured by any government",
    "not a deposit",
    "subject to investment risks",
    "possible loss of the principal",
    "j.p. morgan securities",
    "jpmorgan chase bank",
    "jpms",
    "member financial regulatory",
    "finra",
    "sipc",
    "securities insurance protection",
    "margin account statement",
    "asset account statement",
    "additional information about your",
    "held at jpmorgan",
    "held in margin account",
    "exchange-listed options",
    "cleared through jpms",
    "investment and insurance products",
    "will be mailed shortly",
    "pending financial industry",
    "delay the delivery",
    "february 27, 2026",
    "contact your",
    "j.p. morgan team",
    "different entities",
    "investment accounts",
    "continued",
    # Pricing/valuation disclaimer block
    "important information about",
    "market value information",
    "believes to be reliable",
    "not guaranteed for accuracy",
    "current price is the value",
    "structured yield deposit",
    "gas and mineral interest",
    "non-producing interest",
    "drilling activity",
    "values in this statement",
    "valuations of over-the-counter",
    "not independently verified",
    "j.p. morgan makes no",
    "estimated annual income",
    "estimated yield",
    "bank deposit accounts",
    "certain assets, including",
    "assets not",
    "memo pos",
    "offshore dep",
    "unless we have otherwise agreed",
    "placement agent",
    "investment merits",
    "what to do if",
    "provide your",
    "oral com",
    "we must hear",
    "ncc1",
    "newark, de",
    "beginning on",
    "february 13, 202",
    "2026, your 2025",
    "consolidated summ",
    "tinued",
    "from the value shown",
    "would be overstated",
    "not independ",
    "govern your acc",
    "them at (202)",
    "with fdic",
    "receipts are no",
    "potential u.s. t",
    "publicly traded",
    "management or",
    "market fund at",
    "money market",
    "are held.",
    "services for th",
    "statement.",
    "500 stanton",
    "balances. for m",
    "equity protectio",
)


def _looks_like_account_summary_data_row(r: list) -> bool:
    """True if row has account ID in col 2 (0-based: 2) or numeric begin/end in cols 3,4 (typical Account Summary table)."""
    row_list = r if isinstance(r, (list, tuple)) else [r]
    if len(row_list) < 5:
        return False
    cells = list(row_list)[:8]
    # Account ID in column 2 (index 2); strip footnote superscripts ¹²
    cell2 = str(cells[2] or "").strip() if len(cells) > 2 else ""
    cell2_clean = cell2.replace("\u00b9", "").replace("\u00b2", "").strip()
    if _is_account_id(cell2_clean):
        return True
    # Or numeric beginning/ending in 3,4 (with optional superscript like ¹)
    for idx in (3, 4):
        if idx >= len(cells):
            continue
        v = cells[idx]
        if isinstance(v, (int, float, Decimal)):
            return True
        s = str(v or "").strip().replace("\u00b9", "").replace("\u00b2", "").replace(",", "")
        if s and s.replace(".", "").replace("-", "").isdigit():
            return True
    return False


def _extract_account_summary_table_from_block(rows: list[list]) -> tuple[list | None, list[list], list | None]:
    """
    From one block's rows, extract the account-summary table: header row, data rows, optional total row.
    Skips leading title rows. Collects every row that looks like an account line (account ID or begin/end numbers)
    until we hit a Total row. Returns (header, data_rows, total_row).
    """
    header_row = None
    data_rows = []
    total_row = None
    consecutive_skips = 0
    if not rows:
        return None, [], None
    i = 0
    # Skip section title rows
    while i < len(rows):
        r = rows[i] if isinstance(rows[i], (list, tuple)) else [rows[i]]
        first = (str(r[0] or "").strip().lower() if r else "")[:60]
        if first in ("for the period", "account summary", "period summary") or not first:
            i += 1
            continue
        break
    # Find header: first row that has "account" / "beginning" / "ending" / "number" and no account ID in col 2
    while i < len(rows):
        r = rows[i] if isinstance(rows[i], (list, tuple)) else [rows[i]]
        cells = [str(c).strip() if c is not None else "" for c in (r[:8] if len(r) >= 8 else r + [""] * (8 - len(r)))]
        first_lower = (cells[0] or "").lower()
        if not first_lower:
            i += 1
            continue
        if _is_totals_row(r):
            total_row = r
            i += 1
            break
        cell2 = cells[2] if len(cells) > 2 else ""
        has_account_id = _is_account_id(cell2)
        row_text = " ".join(cells).lower()
        looks_like_header = (
            "account" in first_lower or "account" in row_text[:80]
            or "beginning" in row_text or "ending" in row_text or "change" in row_text or "number" in row_text
            or (len(first_lower) < 30 and not has_account_id and not _looks_like_account_summary_data_row(r))
        )
        if not header_row and looks_like_header and not has_account_id:
            header_row = r
            i += 1
            continue
        if _is_totals_row(r):
            total_row = r
            i += 1
            break
        if _looks_like_account_summary_data_row(r):
            data_rows.append(r)
            consecutive_skips = 0
            i += 1
            continue
        if header_row and data_rows and first_lower in ("for the period", "account summary", "period summary"):
            break
        consecutive_skips += 1
        if header_row and data_rows and consecutive_skips >= 3:
            break
        i += 1
    return header_row, data_rows, total_row


def _merge_period_blocks_into_account_summary(
    blocks: list[tuple[str, list[list]]],
) -> tuple[str, list[list]]:
    """
    Merge all Period Summary blocks into one Account Summary table: one header, all account rows from
    every block, one total row. So we get both ABC Trust rows (and any other accounts) in the sheet.
    """
    merged = []
    header_emitted = False
    total_row = None
    for _sheet_name, rows in blocks:
        header, data_rows, total = _extract_account_summary_table_from_block(rows)
        if header and not header_emitted:
            merged.append(header)
            header_emitted = True
        merged.extend(data_rows)
        if total:
            total_row = total
    if total_row:
        merged.append(total_row)
    if not merged and blocks:
        # Fallback: use first block's rows as-is (original behaviour)
        _, first_rows = blocks[0]
        return "Account Summary", first_rows
    return "Account Summary", merged


def _row_has_table_data(row: list) -> bool:
    """True if row contains a clear table value: currency amount, account ID, or number (including 0)."""
    row_list = row if isinstance(row, (list, tuple)) else [row]
    for c in row_list:
        if c is None:
            continue
        s = str(c).strip().replace("\u00b9", "").replace("\u00b2", "").strip()
        if not s:
            continue
        if _is_account_id(s):
            return True
        if isinstance(c, (int, float, Decimal)):
            return True  # include 0 so account rows with zeros are kept
        if re.search(r"\$[\s\d,]+\.?\d*|[\d,]{3,}\.\d{2}", s):
            return True
    return False


def _get_footer_phrases() -> list:
    """Lazy-load footer phrases from config (so different PDFs can add phrases without code changes)."""
    if _get_footer_phrases._cache is None:
        _get_footer_phrases._cache = load_qb_cleanup_config().get("footer_phrases", [])
    return _get_footer_phrases._cache


_get_footer_phrases._cache = None  # type: ignore[attr-defined]


def _is_footer_row(row: list) -> bool:
    """
    True if row looks like PDF footer/disclaimer text.
    Uses config/qb_cleanup.json footer_phrases so you can add phrases for your PDFs without code changes.
    """
    if not row:
        return False
    cells = [c for c in (row if isinstance(row, (list, tuple)) else [row]) if c is not None and str(c).strip()]
    if not cells:
        return False
    text = " ".join(str(c).strip().lower() for c in cells)
    for phrase in _get_footer_phrases():
        if phrase.lower() in text:
            return True
    return False


def _is_prose_row(row: list) -> bool:
    """
    True if row looks like disclaimer, footnote, or long prose (not table data).
    Rows that contain table data (account ID, $ amounts) are kept even if they have some prose.
    """
    if not row:
        return True
    # Drop footer rows first (e.g. "Form will be available online")
    if _is_footer_row(row):
        return True
    cells = [c for c in (row if isinstance(row, (list, tuple)) else [row]) if c is not None and str(c).strip()]
    if not cells:
        return True
    # Keep rows that clearly have table data (so we don't drop "ABC TRUST, E79271004, $15,088,442..." with bleed)
    if _row_has_table_data(row):
        return False
    # Keep rows that look like table column headers (so we never drop the header row)
    text = " ".join(str(c).strip() for c in cells).lower()
    if any(
        phrase in text
        for phrase in (
            "beginning market value",
            "net deposits",
            "investment results",
            "ending market value",
            "current month",
            "year to date",
            "inception to date",
            "quarter to date",
            "account name",
            "market value",
            "cash in",
            "cash out",
            "quantity",
            "market price",
            "unit cost",
            "adjusted cost",
            "unrealized gain",
            "yield to maturity",
        )
    ):
        return False
    # Row with 2+ short cells and no long prose is likely a table header
    if len(cells) >= 2 and all(len(str(c).strip()) < 50 for c in cells) and len(text) < 120:
        return False
    # Explicit disclaimer / blank page
    if "this page intentionally left blank" in text:
        return True
    if text.strip() in ("continued", "investment accounts"):
        return True
    for phrase in _PROSE_PHRASES:
        if phrase in text:
            return True
    # Long all-caps line (disclaimer style)
    raw = " ".join(str(c).strip() for c in cells)
    if len(raw) > 25 and raw.isupper():
        return True
    # Long paragraph with no table-like number
    has_table_number = bool(re.search(r"\$[\s\d,]|[\d,]{2,}\.\d{2}|[\d,]+\s*%", text))
    if len(text) > 45 and not has_table_number:
        return True
    return False


def _infer_sheets_from_first_section_per_page(sections: list[tuple]) -> tuple[list[str], dict[str, list]]:
    """
    No TOC: infer sheet boundaries from the document. For each page, the first section
    (in document order) is the 'page heading'. Consecutive pages with the same
    first-section name form one sheet; all sections on those pages go to that sheet.
    Returns (ordered_sheet_names, by_sheet). Sections must have page_num (4-tuple).
    """
    from collections import defaultdict

    if not sections:
        return ([], {})
    # Require page numbers
    has_page = any(len(s) >= 4 and s[3] is not None for s in sections)
    if not has_page:
        return ([], {})

    # (page_num -> list of (index, section)) in document order
    by_page: dict[int, list[tuple[int, tuple]]] = defaultdict(list)
    for i, s in enumerate(sections):
        if len(s) >= 4 and s[3] is not None:
            by_page[s[3]].append((i, s))

    # First section on each page (earliest index on that page) = page heading
    page_heading: dict[int, str] = {}
    for p in sorted(by_page.keys()):
        # First section on this page = min index
        items = by_page[p]
        items.sort(key=lambda x: x[0])
        first_sec = items[0][1]
        name = (first_sec[0] or "").strip()[:80]
        page_heading[p] = name or f"Page {p}"

    # Runs: consecutive pages with same heading
    pages_sorted = sorted(page_heading.keys())
    runs: list[tuple[int, int, str]] = []  # (start_page, end_page, heading)
    if not pages_sorted:
        runs = []
    else:
        start = pages_sorted[0]
        cur_heading = page_heading[start]
        for p in pages_sorted[1:]:
            if page_heading[p] == cur_heading:
                continue
            runs.append((start, p - 1, cur_heading))
            start = p
            cur_heading = page_heading[p]
        runs.append((start, pages_sorted[-1], cur_heading))

    # For each page, sheet = heading of the run that contains it
    page_to_sheet: dict[int, str] = {}
    for start, end, heading in runs:
        for p in range(start, end + 1):
            page_to_sheet[p] = heading

    by_sheet = defaultdict(list)
    ordered = []  # sheet order = run order
    seen_sheets = set()
    for _start, _end, heading in runs:
        if heading not in seen_sheets:
            ordered.append(heading)
            seen_sheets.add(heading)

    for s in sections:
        if len(s) >= 4 and s[3] is not None:
            sheet = page_to_sheet.get(s[3], "Other")
            by_sheet[sheet].append(s)
        else:
            by_sheet["Other"].append(s)

    if by_sheet.get("Other"):
        ordered.append("Other")
    return (ordered, dict(by_sheet))


def _toc_sheet_for_page(page_num: int, toc: list[tuple[str, int]]) -> str:
    """Return the TOC heading that covers this page (last TOC entry with start_page <= page_num)."""
    best = None
    for heading, start in toc:
        if start <= page_num:
            best = heading
    return best if best else "Other"


def _section_to_block_rows(section: tuple) -> tuple[list[list], set[int]]:
    """Convert (sec_name, heading_rows, data_rows[, page[, bold_data_row_indices]]) to (rows, bold_row_indices) for writing. Drops PDF page-number footer rows."""
    from tables_to_excel import _drop_page_number_rows

    sec_name = section[0]
    heading_rows = section[1]
    data_rows = section[2]
    bold_data_indices = section[4] if len(section) >= 5 else set()
    rows = []
    for h in heading_rows or []:
        rows.append(h if isinstance(h, (list, tuple)) else [h])
    num_heading = len(rows)
    rows.extend(r if isinstance(r, (list, tuple)) else [r] for r in (data_rows or []))
    rows = _drop_page_number_rows(rows)
    # Indices into block rows that should be bold (data rows that were bold in PDF)
    bold_row_indices = {i for i in (num_heading + j for j in bold_data_indices) if i < len(rows)}
    return (rows, bold_row_indices)


def write_workbook_by_toc(
    sections: list[tuple],
    toc: list[tuple[str, int]],
    output_path: Path,
    validation_errors: list[str] | None = None,
) -> str:
    """
    Build Excel with one sheet per TOC heading. Sections must be 4-tuples (name, heading_rows, data_rows, page_num).
    Each sheet gets all sections whose page falls in that TOC range; sections without page go to "Other".
    """
    from collections import defaultdict

    by_sheet = defaultdict(list)
    for s in sections:
        if len(s) >= 4 and s[3] is not None:
            sheet_name = _toc_sheet_for_page(s[3], toc)
            by_sheet[sheet_name].append(s)
        else:
            by_sheet["Other"].append(s)

    toc_headings = [h for h, _ in toc]
    ordered = [h for h in toc_headings if by_sheet.get(h)]
    if by_sheet.get("Other"):
        ordered.append("Other")

    return _write_workbook_by_sheets(dict(by_sheet), ordered, output_path, validation_errors)


def _write_workbook_by_sheets(
    by_sheet: dict[str, list],
    ordered_sheet_names: list[str],
    output_path: Path,
    validation_errors: list[str] | None = None,
) -> str:
    """
    Build Excel with one sheet per key in ordered_sheet_names; each sheet gets sections from by_sheet[name].
    Shared by TOC-driven and heading-inferred (no-TOC) paths.
    """
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    for target_name in ordered_sheet_names:
        blocks = by_sheet.get(target_name)
        if not blocks:
            continue
        safe_name = _safe_sheet_name(target_name)[:31]
        ws = wb_out.create_sheet(title=safe_name)
        row_num = 1
        for i, section in enumerate(blocks):
            sec_name = section[0]
            block_rows, bold_row_indices = _section_to_block_rows(section)
            if i > 0:
                cell = ws.cell(row=row_num, column=1, value=f"— {sec_name} —")
                cell.font = Font(italic=True)
                _fill_row(ws, row_num, 20, FILL_SECTION_HEADER)
                row_num += 1
                row_num += 1
            for row_idx, r in enumerate(block_rows):
                r_list = r if r else []
                r_list = _merge_fragmented_row(r_list)
                # Skip duplicate section title row (e.g. "INVESTMENT RESULTS" again as first row)
                if row_idx == 0 and r_list and sec_name:
                    first_cell = str(r_list[0] or "").strip()
                    rest_empty = all(not str(c or "").strip() for c in r_list[1:])
                    if first_cell and rest_empty and first_cell.upper() == (sec_name or "").upper():
                        continue
                if _is_prose_row(r_list):
                    continue
                num_cols = max(len(r_list), 1)
                for col_idx, val in enumerate(r_list, start=1):
                    if val is not None and not isinstance(val, (int, float, Decimal)):
                        val = _normalize_cell_value(val)
                    cell = ws.cell(row=row_num, column=col_idx, value=val)
                    if row_idx in bold_row_indices:
                        cell.font = Font(bold=True)
                    if _is_formula_or_check(val):
                        cell.fill = FILL_FORMULA
                    elif _is_account_id(val):
                        cell.fill = FILL_SECTION_HEADER
                if row_idx == 0:
                    _fill_row(ws, row_num, num_cols, FILL_SECTION_HEADER)
                elif row_idx == 1 and _looks_like_header_row(r_list):
                    _fill_row(ws, row_num, num_cols, FILL_TABLE_HEADER)
                elif _is_totals_row(r_list):
                    _fill_row(ws, row_num, num_cols, FILL_TOTALS)
                row_num += 1
            row_num += 1

    if validation_errors:
        ws_val = wb_out.create_sheet(title="Validation Results")
        ws_val.cell(row=1, column=1, value="Requires Review")
        ws_val.cell(row=1, column=2, value="Yes")
        for i, msg in enumerate(validation_errors, start=2):
            ws_val.cell(row=i, column=1, value=msg)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(output_path)
    return str(output_path)


def transform_extracted_to_qb(
    extracted_xlsx_path: str,
    output_path: str,
    validation_errors: list[str] | None = None,
) -> str:
    """
    Read an extracted workbook (from pdf_tables_to_excel) and write a QB-format workbook.

    - Groups sheets by target name (e.g. all Asset Allocation → one sheet).
    - Writes sections one after another with a blank row and section title between.
    - If validation_errors is non-empty, adds a "Validation Results" sheet and marks Requires Review.
    - Returns the path to the written file.
    """
    from openpyxl import load_workbook

    in_path = Path(extracted_xlsx_path)
    out_path = Path(output_path)
    if not in_path.exists():
        raise FileNotFoundError(f"Extracted file not found: {in_path}")

    wb_in = load_workbook(in_path, read_only=True, data_only=True)
    # Collect: target_name -> [(source_sheet_name, rows), ...]
    by_target: dict[str, list[tuple[str, list[list]]]] = {}

    for sheet_name in wb_in.sheetnames:
        ws = wb_in[sheet_name]
        rows = _rows_from_sheet(ws)
        if not rows:
            continue
        target = _target_sheet_name(sheet_name)
        by_target.setdefault(target, []).append((sheet_name, rows))
    wb_in.close()

    # Populate Account Summary by merging all Period Summary blocks (so we get both ABC Trust rows / all accounts)
    if "Period Summary" in by_target and by_target["Period Summary"]:
        # Also carve other common sub-sections out of the Period Summary blocks so downstream
        # builders (PLSummary) and template population can work even when the PDF doesn't
        # have clean TOC headings.
        try:
            for _sheet_name, rows in list(by_target["Period Summary"]):
                for section_title, section_rows in _split_rows_by_section(rows or []):
                    key = (section_title or "").strip().lower()
                    if not key:
                        continue
                    if key.startswith("asset allocation"):
                        by_target.setdefault("Asset Allocation", []).append(("Asset Allocation", section_rows))
                    elif key.startswith("portfolio activity"):
                        by_target.setdefault("Portfolio Activity", []).append(("Portfolio Activity", section_rows))
                    elif key.startswith("tax summary"):
                        by_target.setdefault("Tax Summary", []).append(("Tax Summary", section_rows))
        except Exception:
            pass
        merged_account_summary = _merge_period_blocks_into_account_summary(by_target["Period Summary"])
        by_target["Account Summary"] = [merged_account_summary]
        by_target["Period Summary"] = []  # Skip noisy Period Summary sheet; data is in Account Summary

    # Build ordered list of (target_name, blocks) without any hard-coded ordering.
    # Keep "PLSummary" and "Account Summary" near the top when present, then the rest alphabetically.
    ordered_targets = []
    for preferred in ("PLSummary", "Account Summary"):
        if preferred in by_target:
            ordered_targets.append(preferred)
    for name in sorted(k for k in by_target.keys() if k not in set(ordered_targets)):
        ordered_targets.append(name)

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    # Build PLSummary only when we have QB-style data to fill it (otherwise we get an empty template first)
    plsummary_sources = (
        "Account Summary",
        "Asset Allocation",
        "Portfolio Activity",
        "Overview",
        "Tax Summary",
        "US Tax Summary",
        "Cash Activity",
        "Purchases and Sales",
    )
    has_plsummary_data = any(by_target.get(k) for k in plsummary_sources)
    if has_plsummary_data:
        try:
            from plsummary_builder import build_plsummary_jpm_sheet
            build_plsummary_jpm_sheet(wb_out, by_target)
        except Exception:
            pass  # if builder fails, continue with section-based sheets

    # Skip Period Summary (noisy repeated "Period Summary 2, 3, 4" and prose); keep the rest
    ordered_targets = [t for t in ordered_targets if t != "Period Summary"]

    for target_name in ordered_targets:
        blocks = by_target[target_name]
        safe_name = _safe_sheet_name(target_name)[:31]  # Excel sheet name limit
        ws = wb_out.create_sheet(title=safe_name)
        row_num = 1
        # Per-account sheets: split each block by section (Account Summary, Asset Allocation, etc.) and drop CONTINUED
        is_per_account = target_name.startswith("Account ") and target_name != "Account Summary"
        for i, (source_name, rows) in enumerate(blocks):
            if is_per_account:
                sub_blocks = _split_rows_by_section(rows)
            else:
                sub_blocks = [(source_name, rows)]
            for j, (block_title, block_rows) in enumerate(sub_blocks):
                if j > 0 or i > 0:
                    cell = ws.cell(row=row_num, column=1, value=f"— {block_title} —")
                    cell.font = Font(italic=True)
                    _fill_row(ws, row_num, 20, FILL_SECTION_HEADER)
                    row_num += 1
                    row_num += 1  # blank
                for row_idx, r in enumerate(block_rows):
                    r_list = r if r else []
                    r_list = _merge_fragmented_row(r_list)
                    if target_name == "Broker Info":
                        r_list = _merge_address_fragments_in_row(r_list)
                    if target_name == "Account Summary":
                        r_list = r_list[:6]
                    if _is_prose_row(r_list):
                        continue
                    num_cols = max(len(r_list), 1)
                    for col_idx, val in enumerate(r_list, start=1):
                        if val is not None and not isinstance(val, (int, float, Decimal)):
                            val = _normalize_cell_value(val)
                        cell = ws.cell(row=row_num, column=col_idx, value=val)
                        if _is_formula_or_check(val):
                            cell.fill = FILL_FORMULA
                        elif _is_account_id(val):
                            cell.fill = FILL_SECTION_HEADER
                    if row_idx == 0:
                        _fill_row(ws, row_num, num_cols, FILL_SECTION_HEADER)
                    elif row_idx == 1 and _looks_like_header_row(r_list):
                        _fill_row(ws, row_num, num_cols, FILL_TABLE_HEADER)
                    elif _is_totals_row(r_list):
                        _fill_row(ws, row_num, num_cols, FILL_TOTALS)
                    row_num += 1
                row_num += 1  # blank after each block

    # Validation Results sheet when there are validation errors
    if validation_errors:
        ws_val = wb_out.create_sheet(title="Validation Results")
        ws_val.cell(row=1, column=1, value="Requires Review")
        ws_val.cell(row=1, column=2, value="Yes")
        ws_val.cell(row=2, column=1, value="Message")
        for i, msg in enumerate(validation_errors, start=3):
            ws_val.cell(row=i, column=1, value=msg)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(out_path)
    return str(out_path)


def pdf_to_qb_excel(
    pdf_path: str,
    output_path: str,
    overwrite: bool = True,
    json_path_out: str | Path | None = None,
) -> str:
    """
    Full pipeline: PDF → JSON (canonical) → Excel. JSON is the intermediate; Excel is built from it.

    1. Extract PDF to sections; write JSON (to json_path_out if provided, else temp).
    2. Load sections from JSON; write raw Excel from it.
    3. Transform raw Excel to structured workbook (merge/rename sheets, colors).
    4. Save to output_path. Returns output_path.

    """
    import tempfile
    from datetime import datetime
    from tables_to_excel import (
        extract_sections_from_pdf,
        extract_toc_from_pdf,
        filter_sections_to_tables_only,
        load_sections_from_json,
        merge_section_header_rows,
        validate_sections,
        _write_json_from_sections,
        _write_sections_to_workbook,
    )
    from hybrid_extract import library_routing_meta

    out_path = Path(output_path)
    if out_path.exists() and not overwrite:
        raise FileExistsError(f"Output exists: {out_path}")

    # 1. PDF → JSON (canonical intermediate); keep only table-like sections (no long prose)
    sections_merged = extract_sections_from_pdf(pdf_path)
    sections_merged = [merge_section_header_rows(s) for s in sections_merged]
    # Same routing heuristics as hybrid (no VL): which pages look weak → candidate_vl_pages in meta
    routing_meta = library_routing_meta(sections_merged)
    sections = filter_sections_to_tables_only(sections_merged)
    json_path = Path(json_path_out) if json_path_out else Path(tempfile.NamedTemporaryFile(suffix=".json", delete=False).name)
    try:
        _write_json_from_sections(
            sections,
            json_path,
            overwrite=True,
            meta={
                "pdf_name": Path(pdf_path).name,
                "pdf_path": str(Path(pdf_path)),
                "generated_at_utc": datetime.utcnow().isoformat(timespec="seconds") + "Z",
                "generator": "pdf_to_qb.pdf_to_qb_excel",
                "note": "This JSON may include filtering to table-like sections only before QB transform.",
                "extraction_pipeline": "library_then_qb",
                **routing_meta,
            },
        )
        sections_from_json = load_sections_from_json(str(json_path))
        validation_errors, requires_review = validate_sections(sections_from_json)
        for err in validation_errors or []:
            log.warning("validation: %s", err)
        if requires_review:
            log.warning("Output marked as Requires Review (see Validation Results sheet)")

        # 2. Build an extracted workbook (organized by TOC/headings), then transform it to QB format.
        # The QB transform is where PLSummary builders run.
        toc = extract_toc_from_pdf(pdf_path)
        has_page = any(len(s) >= 4 and s[3] is not None for s in sections_from_json)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            temp_extracted = f.name
        try:
            if toc and has_page:
                log.info("Using table-of-contents from page 1: %d heading(s)", len(toc))
                write_workbook_by_toc(
                    sections_from_json, toc, Path(temp_extracted), validation_errors=validation_errors or None
                )
            else:
                # No TOC: infer sheet boundaries from document — first section on each page = page heading;
                # consecutive pages with same heading = one sheet (e.g. Holdings pages 6–16 → one sheet)
                ordered_names, by_sheet = _infer_sheets_from_first_section_per_page(sections_from_json)
                if ordered_names:
                    log.info("No TOC; inferred %d sheet(s) from first section per page", len(ordered_names))
                    _write_workbook_by_sheets(
                        by_sheet, ordered_names, Path(temp_extracted), validation_errors=validation_errors or None
                    )
                else:
                    _write_sections_to_workbook(sections_from_json, Path(temp_extracted))

            # Transform extracted workbook to QB-format workbook (adds PLSummary when possible).
            transform_extracted_to_qb(temp_extracted, output_path, validation_errors=validation_errors or None)
            return str(out_path)
        finally:
            Path(temp_extracted).unlink(missing_ok=True)
    finally:
        if not json_path_out:
            json_path.unlink(missing_ok=True)
