#!/usr/bin/env python3
"""
PDF → Excel (no AI). Extract all tables from a PDF and write them to an Excel file.

Part of the converter foundation. Use extract.py for the AI agent (natural-language extraction).
"""

import argparse
import json
import logging
import re
import sys
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from statistics import median

logging.basicConfig(level=logging.INFO, format="%(message)s")
log = logging.getLogger(__name__)

import pdfplumber

try:
    import jsonschema
except ImportError:
    jsonschema = None  # optional: validate extraction JSON when present

from config import load_qb_cleanup_config
from openpyxl import Workbook


def _get_header_fragment_merges() -> list:
    """Lazy-load header fragment merges from config (e.g. [['Year-to-', 'Date']])."""
    if _get_header_fragment_merges._cache is None:
        _get_header_fragment_merges._cache = load_qb_cleanup_config().get("header_fragment_merges", [])
    return _get_header_fragment_merges._cache


_get_header_fragment_merges._cache = None  # type: ignore[attr-defined]
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def _safe_sheet_name(name: str, fallback: int) -> str:
    """Excel sheet names: max 31 chars (Excel limit), no \\ / * ? [ ] :"""
    s = (name or f"Sheet{fallback}").replace("\\", "").replace("/", "").replace("*", "").replace("?", "").replace("[", "").replace("]", "").replace(":", " ")
    return (s[:31] if s else f"Sheet{fallback}")[:31]


# Max vertical gap (pt) to consider text as "heading above" a table
HEADING_GAP_PT = 35
# Min vertical gap (pt) between text blocks to start a new section when no grid tables
SECTION_GAP_PT = 18

# Table detection: "text" strategy finds tables without explicit grid lines (e.g. bank statements)
TABLE_SETTINGS = {"vertical_strategy": "text", "horizontal_strategy": "text"}

# Section title patterns to split one big table into sub-tables (e.g. Asset Allocation, Portfolio Activity, Tax Summary)
SECTION_TITLE_PATTERNS = re.compile(
    r"^(Asset\s+Allocation|Portfolio\s+Activity|Tax\s+Summary|Account\s+Summary)$",
    re.IGNORECASE,
)

# Map PDF report/section titles to target sheet names (see EXPECTED_FORMAT.md).
# Order matters: more specific patterns first. Includes broker combined-statement sections (e.g. 9004 PDF).
REPORT_TITLE_TO_SHEET = [
    # Fund-accounting / QB Automation target names
    (re.compile(r"statement\s+of\s+net\s+assets", re.I), "Net Assets"),
    (re.compile(r"statement\s+of\s+operations", re.I), "Operations"),
    (re.compile(r"change\s+in\s+partners['\u2019]?\s*capital", re.I), "Partner Capital"),
    (re.compile(r"mtd\s+pnl\s+per\s+trading\s+account\s+summary", re.I), "PLSummary"),
    (re.compile(r"journal\s+entry\s+import", re.I), "Journal Entry Import"),
    (re.compile(r"journal\s+entries?", re.I), "Journal Entries"),
    (re.compile(r"unrealized\s+gains?\s+and\s+losses?", re.I), "Unrealized"),
    (re.compile(r"changes?\s+in\s+accrued\s+dividend", re.I), "Change in Dividend"),
    (re.compile(r"changes?\s+in\s+accrued\s+interest", re.I), "Change in Interest"),
    (re.compile(r"alt\s+inv\s+transfer", re.I), "Alt Inv Transfer"),
    # Broker combined-statement sections (e.g. JPM 9004-20251231-Combined-Statement)
    (re.compile(r"consolidated\s+summary", re.I), "Consolidated Summary"),
    (re.compile(r"account\s+summary", re.I), "Account Summary"),
    (re.compile(r"for\s+the\s+period", re.I), "Period Summary"),
    (re.compile(r"asset\s+allocation", re.I), "Asset Allocation"),
    (re.compile(r"portfolio\s+activity", re.I), "Portfolio Activity"),
    (re.compile(r"tax\s+summary", re.I), "Tax Summary"),
    (re.compile(r"cash\s+[&]\s+fixed\s+income\s+(summary|detail)", re.I), "Cash & Fixed Income"),
    (re.compile(r"cash\s+[&]\s+fixed\s+income", re.I), "Cash & Fixed Income"),
    (re.compile(r"equity\s+summary", re.I), "Equity Summary"),
    (re.compile(r"equity\s+detail", re.I), "Equity Detail"),
    # Broker statement sections (e.g. GS Preferred and Hybrid, JPM statements)
    (re.compile(r"unrealized\s+gain\s*\(?\s*loss\s*\)?", re.I), "Unrealized"),
    (re.compile(r"holdings(?:\s+\(continued\))?", re.I), "Holdings"),
    (re.compile(r"cash\s+activity(?:\s+\(continued\))?", re.I), "Cash Activity"),
    (re.compile(r"dividends?\s+and\s+distributions?", re.I), "Dividends and Distributions"),
    (re.compile(r"fixed\s+income", re.I), "Fixed Income"),
    (re.compile(r"reportable\s+income", re.I), "Reportable Income"),
    (re.compile(r"reportable\s+interest", re.I), "Reportable Interest"),
    (re.compile(r"purchases?\s*[&]\s*sales?", re.I), "Purchases and Sales"),
    (re.compile(r"realized\s+capital\s+gains?", re.I), "Realized Capital Gains"),
    (re.compile(r"portfolio\s+information", re.I), "Portfolio Information"),
]


def _get_title_to_sheet_config() -> list[tuple[re.Pattern, str]]:
    """Lazy-load title→sheet mappings from config (so you can add patterns for new PDFs without code changes)."""
    if _get_title_to_sheet_config._cache is None:
        raw = load_qb_cleanup_config().get("title_to_sheet") or []
        out = []
        for item in raw:
            if isinstance(item, (list, tuple)) and len(item) >= 2:
                try:
                    # Pattern: plain text matches anywhere; use regex if you need (e.g. "statement.*holdings")
                    out.append((re.compile(str(item[0]), re.I), str(item[1]).strip()))
                except re.error:
                    out.append((re.compile(re.escape(str(item[0])), re.I), str(item[1]).strip()))
        _get_title_to_sheet_config._cache = out
    return _get_title_to_sheet_config._cache


_get_title_to_sheet_config._cache = None  # type: ignore[attr-defined]


def _normalize_section_name_for_lookup(text: str) -> str:
    """Strip ' (Continued)', ' (Continued)1', etc., so continuations map to the same sheet as the base section."""
    if not (text and isinstance(text, str)):
        return ""
    s = text.strip()
    # Remove " (Continued)" or " (Continued)1", " (Continued)2", etc.
    s = re.sub(r"\s*\(Continued\)\s*\d*\s*$", "", s, flags=re.I).strip()
    return s


def _preferred_sheet_name_from_title(text: str) -> str | None:
    """If text matches a known report type (built-in or config), return the target sheet name; else None."""
    if not (text and isinstance(text, str)):
        return None
    s = _normalize_section_name_for_lookup(text)
    if not s:
        return None
    # Config first (so your PDF's titles override built-in)
    for pattern, sheet_name in _get_title_to_sheet_config():
        if pattern.search(s):
            return sheet_name
    for pattern, sheet_name in REPORT_TITLE_TO_SHEET:
        if pattern.search(s):
            return sheet_name
    return None


def _cell_looks_numeric(c) -> bool:
    """True if cell parses as a number (for table-boundary detection: title-only rows have no number)."""
    if c is None or (isinstance(c, str) and not c.strip()):
        return False
    if isinstance(c, (int, float, Decimal)) and not isinstance(c, bool):
        return True
    s = str(c).strip().lstrip("$€£\u00a0").replace(",", "")
    if s.startswith("(") and s.endswith(")"):
        s = s[1:-1].strip()
    return bool(re.match(r"^-?\d+\.?\d*%?$|^-?\d+\.?\d*e[-+]?\d+$", s, re.I))


def _cell_value(c) -> str | int | Decimal:
    """
    Prefer numeric type for Excel (so calculations work); otherwise return cleaned string.
    Strips $ € £ and commas; handles trailing %; parenthetical negatives (123.45); rounds to 2 decimals.
    """
    if c is None:
        return ""
    if isinstance(c, (int, float, Decimal)) and not isinstance(c, bool):
        if isinstance(c, Decimal):
            return c
        if isinstance(c, float) and c != int(c):
            try:
                return Decimal(str(c)).quantize(Decimal("0.01"))
            except Exception:
                return c
        return c if isinstance(c, int) else int(c)
    s = str(c).strip()
    if not s:
        return ""
    # Parenthetical negative: (308.60) or ($37,303.03)
    if s.startswith("(") and s.endswith(")"):
        inner = s[1:-1].strip().lstrip("$€£\u00a0").replace(",", "")
        try:
            val = Decimal(inner)
            return -val.quantize(Decimal("0.01")) if abs(val) >= Decimal("0.01") or val == 0 else -val
        except (ValueError, Exception):
            pass
    # "09 $24,157,595.24" or "24 $24,284,278.98" -> take the dollar amount part only
    m = re.match(r"^\d+\s+[\$€£]?\s*([\d,]+\.?\d*)\s*$", s)
    if m:
        try:
            val = Decimal(m.group(1).replace(",", ""))
            return val.quantize(Decimal("0.01")) if abs(val) >= Decimal("0.01") or val == 0 else val
        except (ValueError, Exception):
            pass
    # Strip currency symbols and commas for parsing
    is_pct = s.endswith("%")
    if is_pct:
        s = s[:-1].strip()
    clean = s.lstrip("$€£\u00a0").replace(",", "")
    try:
        if "." in clean or "e" in clean.lower():
            val = Decimal(clean)
            if is_pct:
                val = val / 100
            if abs(val) >= Decimal("0.01") or val == 0:
                val = val.quantize(Decimal("0.01"))
            return val
        if clean.isdigit() or (clean.startswith("-") and clean[1:].isdigit()):
            return int(clean)
    except (ValueError, Exception):
        pass
    # Strip footnote/superscript markers from strings (e.g. E79271004¹ -> E79271004) for clean account IDs
    s = re.sub(r"[\u00B9\u00B2\u00B3\u2070-\u2079]+$", "", s)
    return s


def _merge_fragmented_row(cells: list) -> list:
    """
    Merge cells that are fragments of one value so numbers and headers stay in one cell.
    - "15,088,442.", "61" -> "15,088,442.61"
    - "22,913,59", "5.63" -> "22,913,595.63"
    - "Beginni", "n", "g" -> "Beginning"
    - "1,421,910.", "03 1,494,773.17" -> "1,421,910.03", "1,494,773.17"
    """
    if not cells:
        return []
    out = []
    for c in cells:
        s = (str(c) if c is not None else "").strip()
        if s and isinstance(c, (int, float, Decimal)) and not isinstance(c, bool):
            s = str(c)
        if not s:
            out.append("")
            continue
        if out and out[-1] != "":
            last = str(out[-1])
            # Merge single digit (e.g. "2" after "15,041,566.")
            if re.match(r"^\d$", s):
                out[-1] = last + s
                continue
            # Merge decimal part: last ends with "." and current is digits (e.g. "61" after "15,088,442.")
            if last.endswith(".") and re.match(r"^\d+$", s):
                out[-1] = last + s
                continue
            # Last ends with "." and current is "DD next_number" (e.g. "03 1,494,773.17"): merge DD, then append next_number
            if last.endswith(".") and re.match(r"^\d+\s+[\d,]+", s):
                m = re.match(r"^(\d+)\s+(.+)$", s)
                if m:
                    dec_part, rest = m.group(1), m.group(2).strip()
                    out[-1] = last + dec_part
                    # Push rest as new value (may be number like "1,494,773.17")
                    try:
                        rest_clean = rest.replace(",", "").lstrip("$€£")
                        if "." in rest_clean or re.search(r"\d", rest_clean):
                            out.append(Decimal(rest_clean).quantize(Decimal("0.01")) if "." in rest_clean else int(rest_clean))
                        else:
                            out.append(rest)
                    except ValueError:
                        out.append(rest)
                    continue
            # Merge decimal part in next cell: last ends with digits and current is short .dd or d.dd (not a full amount)
            # e.g. "22,913,59" + "5.63" -> 22913595.63; avoid "24044839" + "924157595.24" -> one huge number
            if re.search(r"[\d,]+$", last) and re.match(r"^\d*\.?\d+$", s) and len(s) <= 8:
                try:
                    clean_last = last.replace(",", "").lstrip("$€£")
                    combined = clean_last + s
                    combined_clean = combined.replace(",", "")
                    if "." in combined_clean:
                        out[-1] = Decimal(combined_clean).quantize(Decimal("0.01"))
                    else:
                        out[-1] = int(combined_clean)
                except (ValueError, Exception):
                    out.append(s)
                continue
            # Merge word fragments: last ends with letter (4+ chars), current is 1–2 letters only (e.g. "Beginni" + "n" + "g")
            # Use 1–2 so we get "Beginning" from "Beginni"+"n"+"g" but not "Beginning"+"Endi" -> "BeginningEnding"
            if len(last) >= 4 and last[-1].isalpha() and re.match(r"^[a-zA-Z]{1,2}$", s):
                out[-1] = last + s
                continue
            # Merge word continuations when current starts with lowercase and is short
            # e.g. "Accou" + "nt(s)" -> "Account(s)", "Year-En" + "d" -> "Year-End"
            if len(last) >= 4 and last[-1].isalpha() and re.match(r"^[a-z][a-zA-Z()']{0,8}$", s):
                out[-1] = last + s
                continue
            # Merge first token of a split cell, keep the rest as separate cell
            # e.g. "Year-En" + "d Tax Form 1099" -> "Year-End", "Tax Form 1099"
            if len(last) >= 4 and last[-1].isalpha() and re.match(r"^[a-z][a-zA-Z()']{0,8}\s+.+$", s):
                m = re.match(r"^([a-z][a-zA-Z()']{0,8})\s+(.+)$", s)
                if m:
                    token, rest = m.group(1), m.group(2).strip()
                    out[-1] = last + token
                    out.append(rest)
                    continue
            # Merge split parenthetical negative: "(37,30" + "3.03)" -> one number -37303.03
            if re.match(r"^\(\d{1,3}(,\d{2,3})?$", last) and re.match(r"^\d+\.\d{2}\)$", s):
                try:
                    combined = last + s  # "(37,30" + "3.03)" -> "(37,303.03)"
                    inner = combined[1:-1].strip().replace(",", "")
                    val = Decimal(inner)
                    out[-1] = -val.quantize(Decimal("0.01")) if abs(val) >= Decimal("0.01") or val == 0 else -val
                except ValueError:
                    out.append(s)
                continue
            # Merge header fragments from config (e.g. "Year-to-" + "Date" -> "Year-to-Date")
            for prefix, suffix in _get_header_fragment_merges():
                if last.rstrip().endswith(prefix) and s.strip() == suffix:
                    out[-1] = last.rstrip() + suffix
                    break
            else:
                out.append(s)
            continue
        out.append(s)
    return out


def _drop_empty_rows(rows: list[list]) -> list[list]:
    """Remove rows that are entirely empty or only whitespace."""
    return [
        row for row in rows
        if any(c is not None and str(c).strip() != "" for c in (row if isinstance(row, (list, tuple)) else [row]))
    ]


# Match "Page 3 of 54" or "Page  3 of  54" (PDF footer)
_PAGE_OF_PAGE_RE = re.compile(r"^\s*Page\s*\d+\s*of\s*\d+\s*$", re.I)


def _is_page_number_row(row: list) -> bool:
    """True if this row is a PDF page-number footer (e.g. single cell 'Page 3 of 54' or row containing it)."""
    if not row:
        return False
    cells = row if isinstance(row, (list, tuple)) else [row]
    for c in cells:
        if c is None:
            continue
        s = str(c).strip()
        if _PAGE_OF_PAGE_RE.match(s):
            return True
    return False


def _drop_page_number_rows(rows: list[list]) -> list[list]:
    """Remove rows that are PDF page-number footers so they don't appear in Excel/JSON."""
    return [r for r in rows if not _is_page_number_row(r)]


def _clean_table_rows(rows: list) -> list[list]:
    """Apply merge of fragmented cells and drop empty rows. Use for all table row sources."""
    merged = [_merge_fragmented_row(list(r) if isinstance(r, (list, tuple)) else [r]) for r in rows]
    return _drop_empty_rows(merged)


def _split_table_by_section_titles(rows: list[list]) -> list[tuple[str, list[list]]]:
    """
    If the table has rows that look like section titles (e.g. Asset Allocation, Portfolio Activity, Tax Summary),
    split into sub-tables so each gets its own section in Excel. Returns [(title, rows), ...].
    """
    if not rows:
        return []
    result = []
    current_title = None
    current_rows = []
    leading_rows = []
    for row in rows:
        cells = list(row) if isinstance(row, (list, tuple)) else [row]
        first = (cells[0] if cells else None) and str(cells[0]).strip()
        if first and SECTION_TITLE_PATTERNS.match(first):
            if current_title is not None and current_rows:
                result.append((current_title, current_rows))
            elif current_title is None and leading_rows:
                result.append(("Section", leading_rows))
                leading_rows = []
            current_title = first
            current_rows = []
            continue
        if current_title is not None:
            current_rows.append(row)
        else:
            leading_rows.append(row)
    if current_title is not None and current_rows:
        result.append((current_title, current_rows))
    elif leading_rows:
        result.append(("Section", leading_rows))
    return result if result else [("Section", rows)]


def _page_sections_with_headings(page, page_num: int):
    """
    Yield (section_name, heading_rows, data_rows) for the page.
    Uses positioned text and tables so we get: heading, then table, then next heading, then table.
    """
    def _extract_text_lines_with_style(pg):
        # Prefer char-level details so we can use font-size/bold cues.
        try:
            return pg.extract_text_lines(return_chars=True) or []
        except TypeError:
            return pg.extract_text_lines() or []

    def _line_style(line: dict) -> tuple[float, float]:
        """
        Returns (avg_font_size, bold_ratio) for a text line.
        If char-level info is unavailable, returns (0.0, 0.0).
        """
        chars = line.get("chars") if isinstance(line, dict) else None
        if not chars:
            return 0.0, 0.0
        sizes = []
        bold_hits = 0
        for ch in chars:
            sz = ch.get("size")
            if isinstance(sz, (int, float)):
                sizes.append(float(sz))
            fn = str(ch.get("fontname") or "").lower()
            if "bold" in fn or "black" in fn or "demi" in fn:
                bold_hits += 1
        if not sizes:
            return 0.0, 0.0
        avg_size = sum(sizes) / len(sizes)
        bold_ratio = (bold_hits / len(chars)) if chars else 0.0
        return avg_size, bold_ratio

    def _is_heading_candidate(line: dict, body_font_size: float, table_top: float, prev_bottom: float | None) -> bool:
        txt = (line.get("text") or "").strip()
        if not txt:
            return False
        if len(txt) > 140:
            return False  # likely prose
        top_ln = float(line.get("top", 0))
        bottom_ln = float(line.get("bottom", top_ln))
        if prev_bottom is not None and top_ln < prev_bottom + 5:
            return False
        if not (bottom_ln <= table_top + HEADING_GAP_PT and top_ln >= (prev_bottom or 0) - 5):
            return False
        sz, bold_ratio = _line_style(line)
        # Heading cues: visually stronger text near table.
        size_signal = body_font_size > 0 and sz >= (body_font_size + 0.6)
        bold_signal = bold_ratio >= 0.35
        # Also keep obvious title-like short lines even without char info.
        title_shape = len(txt.split()) <= 8 and txt == txt.strip()
        return size_signal or bold_signal or title_shape

    # Get text lines with position (top, bottom)
    if not hasattr(page, "extract_text_lines"):
        return
    text_lines = _extract_text_lines_with_style(page)
    font_sizes = []
    for ln in text_lines:
        sz, _ = _line_style(ln)
        if sz > 0:
            font_sizes.append(sz)
    body_font_size = median(font_sizes) if font_sizes else 0.0
    find_tables_kw = {"table_settings": TABLE_SETTINGS} if hasattr(page, "find_tables") else {}
    tables = page.find_tables(**find_tables_kw) if hasattr(page, "find_tables") else []

    # Build list of (top, kind, payload): "text" -> line dict, "table" -> (bbox, rows)
    elements = []
    for line in text_lines:
        text = (line.get("text") or "").strip()
        if not text:
            continue
        top = float(line.get("top", 0))
        bottom = float(line.get("bottom", top))
        elements.append((top, "text", {"text": text, "top": top, "bottom": bottom}))
    for tbl in tables:
        bbox = tbl.bbox if hasattr(tbl, "bbox") else (0, 0, 0, 0)
        try:
            rows = tbl.extract() or []
        except Exception:
            rows = []
        if not rows:
            continue
        rows = _clean_table_rows(rows)
        if not rows:
            continue
        top = float(bbox[1])
        elements.append((top, "table", {"bbox": bbox, "rows": rows}))

    if not elements:
        return
    elements.sort(key=lambda x: (x[0], 0 if x[1] == "text" else 1))

    prev_table_bottom = None
    section_idx = 0
    for _, kind, payload in elements:
        if kind == "text":
            continue  # we'll attach text to the next table or flush as section
        if kind == "table":
            table_top = payload["bbox"][1]
            table_bottom = payload["bbox"][3]
            # Heading = style-aware lines just above this table.
            heading_rows = []
            for line in text_lines:
                if _is_heading_candidate(line, body_font_size, table_top, prev_table_bottom):
                    t = (line.get("text") or "").strip()
                    if t:
                        heading_rows.append(t)
            # Fallback to positional-only heading capture if style filter found nothing.
            if not heading_rows:
                for line in text_lines:
                    t = (line.get("text") or "").strip()
                    if not t:
                        continue
                    bottom = float(line.get("bottom", 0))
                    top_ln = float(line.get("top", 0))
                    if prev_table_bottom is not None and top_ln < prev_table_bottom + 5:
                        continue
                    if bottom <= table_top + HEADING_GAP_PT and top_ln >= (prev_table_bottom or 0) - 5:
                        heading_rows.append(t)
            rows = payload["rows"]
            sub_tables = _split_table_by_section_titles(rows)
            for sub_title, sub_rows in sub_tables:
                if not sub_rows:
                    continue
                section_idx += 1
                # Use sub-table title (e.g. Asset Allocation) as section name; keep all headings (e.g. NON-REPORTABLE ITEMS) so parent headings appear in Excel
                head = [sub_title] if sub_title != "Section" else (heading_rows or [f"Page{page_num}_T{section_idx}"])
                if heading_rows and sub_title != "Section":
                    head = heading_rows + [sub_title]  # all lines above table (period, parent section) then table title
                name = (sub_title[:28] + f"_{section_idx}") if sub_title != "Section" else ((heading_rows[0][:28] + f"_{section_idx}") if heading_rows else f"Page{page_num}_T{section_idx}")
                yield (_safe_sheet_name(name, section_idx), head, sub_rows)
            prev_table_bottom = table_bottom

    # If no grid tables, group text lines by vertical gap: first line of group = heading, rest = table
    if not tables and text_lines:
        text_lines_sorted = sorted(
            [{"text": (l.get("text") or "").strip(), "top": float(l.get("top", 0)), "bottom": float(l.get("bottom", 0))}
             for l in text_lines if (l.get("text") or "").strip()],
            key=lambda x: x["top"],
        )
        if not text_lines_sorted:
            return
        group = [text_lines_sorted[0]["text"]]
        group_start = text_lines_sorted[0]["top"]
        for i in range(1, len(text_lines_sorted)):
            line = text_lines_sorted[i]
            gap = line["top"] - (text_lines_sorted[i - 1]["bottom"])
            if gap > SECTION_GAP_PT and group:
                # Flush group: first line = heading, rest = rows (split on 2+ spaces); then merge fragments
                heading = group[0]
                section_idx += 1
                rows = []
                for g in group[1:]:
                    parts = re.split(r"[\t ]{2,}", g)
                    rows.append(parts if len(parts) > 1 else [g])
                rows = _clean_table_rows(rows)
                yield (_safe_sheet_name(heading[:28] + f"_{section_idx}", section_idx), [heading], rows)
                group = []
            group.append(line["text"])
        if group:
            heading = group[0]
            section_idx += 1
            rows = []
            for g in group[1:]:
                parts = re.split(r"[\t ]{2,}", g)
                rows.append(parts if len(parts) > 1 else [g])
            rows = _clean_table_rows(rows)
            yield (_safe_sheet_name(heading[:28] + f"_{section_idx}", section_idx), [heading], rows)


def _normalize_row(row) -> list:
    """Turn a table row into a list of cell values (numbers preserved where possible)."""
    if isinstance(row, (list, tuple)):
        return [_cell_value(c) for c in row]
    return [_cell_value(row)]


def _parse_summary_lines(heading_rows: list[str]) -> tuple[str, list[list], list[str]]:
    """
    From heading lines above a table, get section title, optional key-value summary, and any
    lines that look like table headers (to be prepended to the data table).
    Returns (title, summary_rows, table_header_lines). summary_rows is list of [label, value].
    table_header_lines are raw lines to use as the first row(s) of the data table.
    """
    if not heading_rows:
        return "Section", [], []
    title = heading_rows[0].strip()
    if len(heading_rows) == 1:
        return title, [], []

    summary_rows = []
    table_header_lines = []
    for line in heading_rows[1:]:
        line = (line or "").strip()
        if not line:
            continue
        # Key-value summary: "Label: value" or "Label - value"
        if ":" in line:
            idx = line.index(":")
            summary_rows.append([line[:idx].strip(), line[idx + 1 :].strip()])
        elif " - " in line and not re.search(r"\d{2,}", line):
            a, _, b = line.partition(" - ")
            summary_rows.append([a.strip(), b.strip()])
        else:
            # Could be a table header row (multiple columns when split on 2+ spaces) or body text
            parts = re.split(r"  +", line)
            if len(parts) >= 2 and len(line) < 120:
                table_header_lines.append(line)
            else:
                # Single block of text or long line: treat as summary row (label, value) or skip
                if len(parts) == 2:
                    summary_rows.append([parts[0].strip(), parts[1].strip()])
                elif not table_header_lines:
                    summary_rows.append([line, ""])
                else:
                    table_header_lines.append(line)
    return title, summary_rows, table_header_lines


def _split_table_cell(cell: str) -> list:
    """
    Try to split a single cell that may contain tabular data into columns.
    Uses 2+ spaces as column separator; if none, tries splitting before numbers (e.g. "Name 1,234.56 2,345.67").
    """
    if not (cell and isinstance(cell, str)):
        return [cell] if cell is not None else [""]
    cell = cell.strip()
    if not cell:
        return [""]
    parts = re.split(r"  +", cell)
    if len(parts) > 1:
        return [_cell_value(p) for p in parts]
    # Two or more spaces before a digit (amounts)
    parts = re.split(r"  +(?=[\d$€£])", cell)
    if len(parts) > 1:
        return [_cell_value(p) for p in parts]
    # Single space before an amount (digits, optional commas, decimal): "Name 15,088.61 16,000.00" -> [Name, 15,088.61, 16,000.00]
    parts = re.split(r" ([\d][\d,]*\.[\d]+)", cell)
    if len(parts) > 1:
        out = []
        for i, p in enumerate(parts):
            p = p.strip()
            if not p:
                continue
            if i % 2 == 1:
                out.append(_cell_value(p))
            else:
                out.append(_cell_value(p))
        if len(out) > 1:
            return out
    return [_cell_value(cell)]


def _row_has_data_like_cells(row: list) -> bool:
    """True if row looks like a data row: contains CURRENT MONTH/YEAR or has 2+ numeric cells."""
    if not row:
        return False
    cells = row if isinstance(row, (list, tuple)) else [row]
    text = " ".join(str(c) for c in cells).upper()
    if "CURRENT MONTH" in text or "CURRENT YEAR" in text or "YEAR TO DATE" in text or "QUARTER TO DATE" in text:
        return True
    numeric_count = sum(1 for c in cells if _cell_looks_numeric(c))
    return numeric_count >= 2


def merge_section_header_rows(section: tuple) -> tuple:
    """
    If section has data_rows that start with multi-line header rows (then a data row like CURRENT MONTH),
    merge those header rows into one. Returns section with same structure and merged data_rows.
    """
    if not section or len(section) < 3:
        return section
    sec_name, heading_rows, data_rows = section[0], section[1], section[2]
    rest = section[3:] if len(section) > 3 else ()
    if not data_rows:
        return section
    merged_data = _merge_multi_line_header_rows(data_rows)
    if len(rest) > 0:
        return (sec_name, heading_rows, merged_data, *rest)
    return (sec_name, heading_rows, merged_data)


def _is_inception_or_performance_title_row(row: list) -> bool:
    """True if row looks like a standalone title (e.g. 'Inception Date for Performance May 9 25') - keep above merged header."""
    if not row:
        return False
    cells = row if isinstance(row, (list, tuple)) else [row]
    first = str(cells[0] or "").strip()
    if not first or len(first) < 15:
        return False
    lower = first.lower()
    return "inception" in lower and ("date" in lower or "performance" in lower)


def _merge_multi_line_header_rows(rows: list[list]) -> list[list]:
    """
    Merge consecutive rows at the top that look like multi-line column headers into one header row.
    Uses the first row that looks like data (CURRENT MONTH/YEAR or 2+ numeric cells) to get column count N.
    Header rows are right-aligned to N columns (so last column aligns with data), then merged column-wise.
    First column (row label column) is set to '-' when empty.
    If the very first row looks like "Inception Date for Performance..." (PERFORMANCE table), keep it separate and merge only the rows below it.
    """
    if not rows or len(rows) < 2:
        return rows
    rows_as_lists = []
    for r in rows:
        rows_as_lists.append(list(r) if isinstance(r, (list, tuple)) else [r])
    data_idx = None
    for i, r in enumerate(rows_as_lists):
        if _row_has_data_like_cells(r):
            data_idx = i
            break
    if data_idx is None or data_idx <= 1:
        return rows
    N = len(rows_as_lists[data_idx])
    if N <= 0:
        return rows
    # PERFORMANCE-style: keep "Inception Date for Performance May 9 25" as its own row, merge only rows below it
    prefix_row = None
    if data_idx >= 2 and _is_inception_or_performance_title_row(rows_as_lists[0]):
        prefix_row = rows_as_lists[0]
        header_rows = rows_as_lists[1:data_idx]
    else:
        header_rows = rows_as_lists[:data_idx]
    # Right-align each header row to N columns: row with K cells → columns [N-K .. N-1]
    grid = []
    for r in header_rows:
        K = len(r)
        padded = [""] * N
        for j, val in enumerate(r):
            cell = str(val).strip() if val is not None else ""
            if N - K + j >= 0 and N - K + j < N:
                padded[N - K + j] = cell
        grid.append(padded)
    merged = []
    for j in range(N):
        parts = [grid[i][j].strip() for i in range(len(grid)) if grid[i][j] and str(grid[i][j]).strip()]
        merged.append(" ".join(parts).strip() if parts else "")
    if not merged[0].strip():
        merged[0] = "-"
    if prefix_row is not None:
        return [prefix_row, merged] + rows_as_lists[data_idx:]
    return [merged] + rows_as_lists[data_idx:]


def _normalize_table_rows(data_rows: list) -> list[list]:
    """
    Ensure all rows have the same length (pad with empty strings).
    If every row is single-cell, try splitting into columns (2+ spaces, or space-before-number).
    """
    if not data_rows:
        return []
    rows_as_lists = []
    for row in data_rows:
        if isinstance(row, (list, tuple)):
            r = [str(c).strip() if c is not None else "" for c in row]
        else:
            r = [str(row).strip()]
        rows_as_lists.append(r)

    # If all rows are single-cell, try splitting into columns
    if all(len(r) == 1 for r in rows_as_lists):
        split_rows = []
        for r in rows_as_lists:
            split_rows.append(_split_table_cell(r[0]))
        if split_rows and max(len(s) for s in split_rows) > 1:
            rows_as_lists = split_rows

    max_cols = max(len(r) for r in rows_as_lists)
    out = []
    for r in rows_as_lists:
        vals = [_cell_value(c) if isinstance(c, str) else c for c in r]
        out.append(vals + [""] * (max_cols - len(r)))
    return out


def _is_single_header_row(heading_rows: list) -> bool:
    """True if heading_rows is a single row of column headers (e.g. from JSON/VL: ["Account Name", "Market Value", ...])."""
    if not heading_rows or len(heading_rows) < 2:
        return False
    for h in heading_rows:
        if not isinstance(h, str):
            return False
        s = (h or "").strip()
        if ":" in s or "  " in s or len(s) > 80:
            return False
    return True


def _write_section_to_sheet(
    ws, sec_name: str, heading_rows: list, data_rows: list, start_row: int = 1, bold_data_row_indices: set | None = None
) -> int:
    """
    Write one section to a worksheet like the PDF: section title, optional summary table,
    then the main data table with header row and columns. Uses bold for title and table header.
    If bold_data_row_indices is set (0-based indices into data_rows), those data rows are written bold to match PDF.
    Handles both PDF-style heading lines and a single row of column headers (from JSON/VL) for QB-style output.
    Returns the next row number after this section (so callers can append more sections).
    """
    bold_data_row_indices = bold_data_row_indices or set()
    # Single row of column headers (e.g. from VL/JSON): use as table header row directly for clean QB-style sheets
    if _is_single_header_row(heading_rows):
        title = (sec_name or "Section").strip()
        summary_rows = []
        combined_data = [[str(h).strip() for h in heading_rows]] + _drop_page_number_rows(data_rows)
        skip_merge = True  # keep header row unchanged
    else:
        title, summary_rows, table_header_lines = _parse_summary_lines(heading_rows)
        combined_data = []
        for line in table_header_lines:
            parts = re.split(r"  +", line)
            combined_data.append(parts if len(parts) > 1 else [line])
        combined_data.extend(_drop_page_number_rows(data_rows))
        skip_merge = False

    row_num = start_row

    # Section title (bold)
    ws.cell(row=row_num, column=1, value=title).font = Font(bold=True)
    row_num += 1

    # Optional summary as 2-column table (label | value)
    if summary_rows:
        for label, value in summary_rows:
            ws.cell(row=row_num, column=1, value=label)
            ws.cell(row=row_num, column=2, value=value)
            row_num += 1
        row_num += 1  # blank row after summary

    data_rows = combined_data

    # Merge multi-line column headers only when not already a clean single header row (avoids dropping/changing headers)
    if not skip_merge:
        data_rows = _merge_multi_line_header_rows(data_rows)

    # Main data table: header row + data rows
    table_rows = _normalize_table_rows(data_rows)
    # Merge fragmented text/number cells so Excel doesn't look "cut" or shifted.
    table_rows = [_merge_fragmented_row(list(r)) for r in (table_rows or [])]
    # Drop all-blank rows so we don't write empty lines
    table_rows = [r for r in table_rows if r and any(str(c or "").strip() for c in r)]
    if not table_rows:
        if row_num == start_row + 1 and not summary_rows:
            ws.cell(row=row_num, column=1, value="(No table data)")
            row_num += 1
        return row_num

    # Avoid duplicate section title: only drop first row if it's clearly a title repeat (single non-empty cell matching title), not a real header row
    if title and len(table_rows) >= 1:
        first_row = table_rows[0]
        first_cell = str(first_row[0] or "").strip() if first_row else ""
        non_empty_count = sum(1 for c in (first_row or []) if str(c or "").strip())
        rest_empty = all(not str(c or "").strip() for c in (first_row[1:] if len(first_row) > 1 else []))
        if non_empty_count >= 2:
            pass  # real header row (multiple columns) — never drop
        elif first_cell and rest_empty and first_cell.upper() == title.upper():
            table_rows = table_rows[1:]
        elif first_cell and first_cell.upper() == title.upper():
            table_rows = [["-"] + list(first_row[1:])] + list(table_rows[1:])
    if not table_rows:
        return row_num

    # Table header (first row, bold)
    for col_idx, val in enumerate(table_rows[0], start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=_cell_value(val))
        cell.font = Font(bold=True)
    row_num += 1

    # Data rows (bold where PDF had bold, e.g. TOTAL lines)
    for data_idx, r in enumerate(table_rows[1:]):
        for col_idx, val in enumerate(r, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=_cell_value(val) if not isinstance(val, (int, float, Decimal)) else val)
            if data_idx in bold_data_row_indices:
                cell.font = Font(bold=True)
        row_num += 1

    # Auto-fit column widths for this section's columns (only if we're the first section or single-sheet)
    num_cols = len(table_rows[0]) if table_rows else 2
    for col_idx in range(1, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for r in range(start_row, row_num):
            cell = ws.cell(row=r, column=col_idx)
            if cell.value is not None:
                max_len = max(max_len, min(50, len(str(cell.value))))
        if max_len > 0:
            current_width = ws.column_dimensions[col_letter].width or 0
            ws.column_dimensions[col_letter].width = max(current_width, max_len + 1)
    return row_num


def extract_toc_from_pdf(pdf_path: str | Path) -> list[tuple[str, int]]:
    """
    Extract table-of-contents from the first page: (heading, page_number) for each TOC entry.
    Looks for rows with a 'Page N' style second column (e.g. "General Information" / "Page 2").
    Returns [] if no TOC found or PyMuPDF unavailable. Used to drive one sheet per TOC heading.
    """
    pdf_path = Path(pdf_path)
    if not pdf_path.exists() or pdf_path.suffix.lower() != ".pdf":
        return []
    try:
        import fitz
    except ImportError:
        return []
    ROW_TOLERANCE = 5
    COL_GAP = 15

    def page_to_rows(page) -> list[list[str]]:
        """Build list of rows; each row is list of cell strings."""
        blocks = page.get_text("dict", flags=fitz.TEXT_PRESERVE_WHITESPACE).get("blocks") or []
        items = []
        for block in blocks:
            for line in block.get("lines") or []:
                for span in line.get("spans") or []:
                    text = (span.get("text") or "").strip()
                    if not text:
                        continue
                    bbox = span.get("bbox", (0, 0, 0, 0))
                    x0, y0 = bbox[0], bbox[1]
                    items.append((y0, x0, text))
        if not items:
            return []
        items.sort(key=lambda t: (round(t[0] / ROW_TOLERANCE), t[1]))
        rows = []
        current_y = None
        current_cells = []
        for y0, x0, text in items:
            y_key = round(y0 / ROW_TOLERANCE)
            if current_y is not None and y_key != current_y:
                if current_cells:
                    rows.append([c for _, c in sorted(current_cells, key=lambda t: t[0])])
                current_cells = []
            current_y = y_key
            if current_cells and abs(x0 - current_cells[-1][0]) < COL_GAP:
                current_cells[-1] = (current_cells[-1][0], (current_cells[-1][1] + " " + text).strip())
            else:
                current_cells.append((x0, text))
        if current_cells:
            rows.append([c for _, c in sorted(current_cells, key=lambda t: t[0])])
        return rows

    toc = []
    total_pages = None
    try:
        doc = fitz.open(str(pdf_path))
        if len(doc) == 0:
            doc.close()
            return []
        total_pages = len(doc)
        page = doc[0]
        rows = page_to_rows(page)
        doc.close()
    except Exception:
        return []
    # Find rows where last cell is "Page N" or just a digit (TOC page number)
    page_re = re.compile(r"Page\s*(\d+)", re.I)
    digits_only = re.compile(r"^\d+$")
    numericish = re.compile(r"^[\d,]+(\.\d+)?$")
    for row in rows:
        if not row or not isinstance(row, (list, tuple)) or len(row) < 2:
            continue
        page_num = None
        col_idx = -1
        last_cell = str(row[-1]).strip() if row else ""
        if digits_only.match(last_cell):
            page_num = int(last_cell)
            col_idx = len(row) - 1
        else:
            for col_idx in range(len(row) - 1, -1, -1):
                cell = (row[col_idx] if col_idx < len(row) else "") and str(row[col_idx]).strip()
                if not cell:
                    continue
                m = page_re.search(cell)
                if m:
                    page_num = int(m.group(1))
                    break
        if page_num is None:
            continue
        # Heading = cell immediately before the page number (TOC section name)
        if col_idx > 0:
            heading = str(row[col_idx - 1]).strip()
        else:
            heading = " ".join(str(row[i]).strip() for i in range(col_idx) if i < len(row) and row[i]).strip()
        if not heading or len(heading) > 80:
            continue
        # Reject "TOC headings" that are actually numbers/amounts.
        if numericish.match(heading.replace("$", "").replace("(", "").replace(")", "").strip()):
            continue
        # Skip document title, period, and account/portfolio ID row
        if heading.upper().startswith("GS:") or heading.upper().startswith("PERIOD COVERING"):
            continue
        if "PORTFOLIO NO" in heading.upper() or "XXX-XX" in heading:
            continue
        # Skip if heading looks like a name/contact (e.g. "Brandon Geer, 310-...")
        if re.search(r"\d{3}-\d{3}-\d{4}", heading) or heading.startswith("Suite ") or "Boulevard" in heading:
            continue
        # Sanity-check page number: must be within the PDF page count.
        if total_pages is not None and (page_num < 1 or page_num > total_pages):
            continue
        toc.append((heading, page_num))
    # Sort by page number; dedupe by heading (keep first)
    seen_headings = set()
    ordered = []
    for h, p in sorted(toc, key=lambda x: x[1]):
        key = h.strip().lower()
        if key in seen_headings:
            continue
        seen_headings.add(key)
        ordered.append((h, p))
    return ordered


def _extract_sections_from_pdf_pymupdf(pdf_path: Path, max_pages: int | None = None) -> list[tuple[str, list, list]]:
    """
    Fallback when pdfplumber gets no text: use PyMuPDF (fitz) with layout.
    Uses get_text("dict") so we keep table structure: group spans by row (y), sort by x for columns.
    Returns list of (sec_name, heading_rows, data_rows, page_num) when used for TOC-driven sheets.
    """
    try:
        import fitz
    except ImportError:
        log.warning("PyMuPDF not installed; pip install pymupdf for fallback extraction")
        return []

    ROW_TOLERANCE = 5  # pts: spans within this y-distance are same row
    COL_GAP = 15  # pts: gap in x to treat as new column

    def page_to_rows(page) -> tuple[list[list[str]], list[bool]]:
        """Build list of rows; each row is list of cell strings. Also return row_bold[i]=True if row i has bold text in PDF."""
        blocks = page.get_text("dict", flags=fitz.TEXT_PRESERVE_WHITESPACE).get("blocks") or []
        # PDF font flags: bit 1 (2)=italic, bit 4 (16)=bold; also check font name
        def _span_bold(span) -> bool:
            flags = span.get("flags", 0) or 0
            if (flags & 16) != 0:
                return True
            font = (span.get("font") or "").lower()
            return "bold" in font
        # Collect (y0, x0, text, is_bold) for each span
        items = []
        for block in blocks:
            for line in block.get("lines") or []:
                for span in line.get("spans") or []:
                    text = (span.get("text") or "").strip()
                    if not text:
                        continue
                    bbox = span.get("bbox", (0, 0, 0, 0))
                    x0, y0 = bbox[0], bbox[1]
                    items.append((y0, x0, text, _span_bold(span)))
        if not items:
            return [], []
        # Group by row (similar y); track if any span in row was bold
        items.sort(key=lambda t: (round(t[0] / ROW_TOLERANCE), t[1]))
        rows = []
        row_bold = []
        current_y = None
        current_cells = []  # list of (x0, text, is_bold)
        for y0, x0, text, is_bold in items:
            y_key = round(y0 / ROW_TOLERANCE)
            if current_y is not None and y_key != current_y:
                if current_cells:
                    row_vals = [c[1] for c in sorted(current_cells, key=lambda t: t[0])]
                    rows.append(row_vals)
                    row_bold.append(any(c[2] for c in current_cells))
                current_cells = []
            current_y = y_key
            if current_cells and abs(x0 - current_cells[-1][0]) < COL_GAP:
                prev = current_cells[-1]
                current_cells[-1] = (prev[0], (prev[1] + " " + text).strip(), prev[2] or is_bold)
            else:
                current_cells.append((x0, text, is_bold))
        if current_cells:
            row_vals = [c[1] for c in sorted(current_cells, key=lambda t: t[0])]
            rows.append(row_vals)
            row_bold.append(any(c[2] for c in current_cells))
        return rows, row_bold

    def _is_table_start_row(row) -> bool:
        """True if row looks like a table section title only (no numeric in same row). Used for boundaries."""
        if not row or not isinstance(row, (list, tuple)):
            return False
        first_cell = (row[0] if row else "").strip()
        if not first_cell or len(first_cell) < 10 or len(first_cell) > 45:
            return False
        if not first_cell.isupper() or re.match(r"^[\d,.\s\$\(\)\-%]+$", first_cell):
            return False
        if " AS OF " in first_cell or "INCLUDING" in first_cell:
            return False
        # Title-only: no numeric value in the rest of the row (so "INTEREST RECEIVED", 221145.71 is not a new table)
        rest = row[1:] if len(row) > 1 else []
        if any(_cell_looks_numeric(c) for c in rest):
            return False
        return True

    sections = []
    doc = fitz.open(str(pdf_path))
    try:
        n_pages = len(doc)
        if max_pages is not None:
            n_pages = min(n_pages, max_pages)
        for page_num in range(n_pages):
            page = doc[page_num]
            rows, row_bold = page_to_rows(page)
            if not rows:
                continue
            # Find table boundaries: rows that are section titles only (no data in same row)
            starts = [i for i, row in enumerate(rows) if _is_table_start_row(row)]
            for k, i in enumerate(starts):
                sec_name = (rows[i][0] if rows[i] else "").strip()[:50]
                end_i = starts[k + 1] if k + 1 < len(starts) else len(rows)
                data_rows = rows[i + 1 : end_i]
                bold_slice = row_bold[i + 1 : end_i]
                # Prepend any rows before the first table on this page (e.g. "TOTAL PORTFOLIO: ..." with value)
                if k == 0 and i > 0:
                    data_rows = rows[0:i] + data_rows
                    bold_slice = row_bold[0:i] + bold_slice
                cleaned = _clean_table_rows(data_rows)
                data_rows = [_normalize_row(r) for r in cleaned]
                # Bold indices into data_rows (by position; cleaning may drop rows so we cap by len(data_rows))
                bold_data_row_indices = {j for j in range(min(len(bold_slice), len(data_rows))) if bold_slice[j]}
                if data_rows:  # only emit if there is content
                    # If previous table-start had no data (e.g. "NON-REPORTABLE ITEMS"), add it as parent heading
                    heading_rows = [sec_name]
                    if k > 0:
                        prev_i = starts[k - 1]
                        prev_sec = (rows[prev_i][0] if rows[prev_i] else "").strip()[:50]
                        prev_data = _clean_table_rows(rows[prev_i + 1 : i])
                        if not prev_data:
                            heading_rows = [prev_sec, sec_name]
                    if bold_data_row_indices:
                        sections.append((sec_name, heading_rows, data_rows, page_num + 1, bold_data_row_indices))
                    else:
                        sections.append((sec_name, heading_rows, data_rows, page_num + 1))
            if not starts and rows:
                # No table-start row on page: one block with all rows (e.g. continuation)
                raw_title = (rows[0][0] if rows and rows[0] else "")
                title = (str(raw_title).strip()[:50] if raw_title else f"Page{page_num + 1}")
                if len(title) < 5:
                    title = f"Page{page_num + 1}"
                data_rows = [_normalize_row(r) for r in _clean_table_rows(rows[1:])]
                bold_slice = row_bold[1:] if len(row_bold) > 1 else []
                bold_data_row_indices = {j for j, b in enumerate(bold_slice[: len(data_rows)]) if b} if bold_slice else set()
                if data_rows:
                    if bold_data_row_indices:
                        sections.append((f"Page{page_num + 1}", [title], data_rows, page_num + 1, bold_data_row_indices))
                    else:
                        sections.append((f"Page{page_num + 1}", [title], data_rows, page_num + 1))
    finally:
        doc.close()
    return sections


def _score_sections(sections: list[tuple[str, list, list]]) -> float:
    """
    Score extraction result for comparison. Higher = better.
    total_data_rows + 2 * num_table_like_sections - penalty for inconsistent column counts.
    Table-like = any section with at least one row (single-row summaries and key-value rows count).
    """
    if not sections:
        return 0.0
    total_rows = 0
    table_like = 0
    inconsistent = 0
    for s in sections:
        _name, _headings, data_rows = s[0], s[1], s[2]
        if not data_rows:
            continue
        total_rows += len(data_rows)
        # Any row/column structure with data counts as table (including 1 row × N columns)
        if len(data_rows) >= 1:
            table_like += 1
        header_len = len(data_rows[0]) if data_rows[0] else 0
        for r in data_rows[1:]:
            row_len = len(r) if isinstance(r, (list, tuple)) else 0
            if header_len and row_len != header_len:
                inconsistent += 1
    penalty = min(inconsistent * 2, total_rows)  # cap penalty
    return total_rows + 2.0 * table_like - penalty


def _extract_sections_from_pdf_camelot(pdf_path: Path, max_pages: int | None = None) -> list[tuple[str, list, list]]:
    """Extract tables using Camelot. Returns list of (section_name, heading_rows, data_rows). On failure returns []."""
    try:
        import camelot
    except ImportError:
        log.debug("Camelot not installed; skip Camelot extraction")
        return []
    sections = []
    pages_arg = "all" if max_pages is None else f"1-{max_pages}"
    try:
        # Try lattice first (bordered tables), then stream (whitespace-separated)
        for flavor in ("lattice", "stream"):
            try:
                tables = camelot.read_pdf(str(pdf_path), pages=pages_arg, flavor=flavor)
                break
            except Exception:
                continue
        else:
            return []
        for i, t in enumerate(tables):
            try:
                df = t.df
                if df is None or df.empty:
                    continue
                rows = df.astype(str).values.tolist()
                page = getattr(t, "page", i + 1)
                name = f"Page{page}_Camelot_{i + 1}"
                cleaned = _clean_table_rows(rows)
                data_rows = [_normalize_row(r) for r in cleaned]
                if data_rows:
                    sections.append((name, [name], data_rows))
            except Exception as e:
                log.debug("Camelot table %s skip: %s", i, e)
    except Exception as e:
        log.warning("Camelot extraction failed: %s", e)
    return sections


def _extract_sections_from_pdf_tabula(pdf_path: Path, max_pages: int | None = None) -> list[tuple[str, list, list]]:
    """Extract tables using Tabula (requires Java). Returns list of (section_name, heading_rows, data_rows). On failure returns []."""
    try:
        import tabula
    except ImportError:
        log.debug("Tabula not installed; skip Tabula extraction")
        return []
    sections = []
    pages_arg = "all" if max_pages is None else f"1-{max_pages}"
    try:
        dfs = tabula.read_pdf(str(pdf_path), pages=pages_arg, multiple_tables=True)
        if not dfs:
            return []
        for i, df in enumerate(dfs):
            try:
                if df is None or df.empty:
                    continue
                rows = df.astype(str).values.tolist()
                name = f"Tabula_{i + 1}"
                cleaned = _clean_table_rows(rows)
                data_rows = [_normalize_row(r) for r in cleaned]
                if data_rows:
                    sections.append((name, [name], data_rows))
            except Exception as e:
                log.debug("Tabula table %s skip: %s", i, e)
    except Exception as e:
        log.warning("Tabula extraction failed: %s", e)
    return sections


def extract_sections_from_pdf(pdf_path: str, max_pages: int | None = None) -> list[tuple[str, list, list]]:
    """
    Extract all tables/sections from the PDF in document order.
    Returns list of (section_name, heading_rows, data_rows).
    Same structure used for Excel and JSON; no file is written.

    If max_pages is set, only pages 1..max_pages are read (faster; hybrid --max-pages).

    A "table" is any row/column layout with values: multi-row grids, single-row summaries,
    key-value lines, and calculation breakdowns. There is no minimum row count; one row
    with multiple columns (e.g. a summary line) is kept.
    """
    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        raise FileNotFoundError(f"Not found: {pdf_path}")
    if pdf_path.suffix.lower() != ".pdf":
        raise ValueError("File must be a .pdf")

    sections = []
    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        for page_num, page in enumerate(pdf.pages, start=1):
            if max_pages is not None and page_num > max_pages:
                break
            if total_pages > 1:
                log.info("Page %d/%d", page_num, total_pages)
            used_sections = False
            if hasattr(page, "extract_text_lines"):
                for sec_name, heading_rows, data_rows in _page_sections_with_headings(page, page_num):
                    sections.append((sec_name, heading_rows, [_normalize_row(r) for r in data_rows], page_num))
                    used_sections = True
            if not used_sections:
                tables = page.extract_tables()
                if tables:
                    for i, table in enumerate(tables):
                        if not table:
                            continue
                        name = f"Page{page_num}" if len(tables) == 1 else f"Page{page_num}_T{i+1}"
                        heading_rows = [name]
                        cleaned = _clean_table_rows(table)
                        data_rows = [_normalize_row(r) for r in cleaned]
                        sections.append((name, heading_rows, data_rows, page_num))
                else:
                    text = page.extract_text()
                    if text and text.strip():
                        lines = [line.strip() for line in text.strip().splitlines() if line.strip()]
                        if lines:
                            heading = lines[0]
                            rows = []
                            for line in lines[1:]:
                                parts = re.split(r"[\t ]{2,}", line)
                                rows.append(parts if len(parts) > 1 else [line])
                            rows = _clean_table_rows(rows)
                            data_rows = [_normalize_row(r) for r in rows]
                            sections.append((f"Page{page_num}", [heading], data_rows, page_num))
                    else:
                        sections.append((f"Page{page_num}", ["(No text extracted from this page)"], [], page_num))

    def _section_has_data(data_rows):
        """True if section has any non-empty cell. Keeps single-row tables (e.g. summary lines, key-value rows)."""
        if not data_rows:
            return False
        for r in data_rows:
            if isinstance(r, (list, tuple)):
                if r and any(c is not None and str(c).strip() != "" for c in r):
                    return True
            elif r is not None and str(r).strip() != "":
                return True
        return False

    merged = []
    pending_title = None
    for s in sections:
        sec_name = s[0]
        heading_rows = s[1]
        data_rows = s[2]
        page_num = s[3] if len(s) >= 4 else None
        if not _section_has_data(data_rows):
            log.debug("section=%s rows=0 skipped (no data)", sec_name[:50] if sec_name else "(unnamed)")
            if heading_rows and heading_rows[0].strip():
                pending_title = heading_rows[0].strip() if not pending_title else f"{pending_title} — {heading_rows[0].strip()}"
            continue
        if pending_title:
            heading_rows = [pending_title, *heading_rows]
            pending_title = None
        log.debug("section=%s rows=%d", (sec_name or "")[:50], len(data_rows))
        if page_num is not None:
            merged.append((sec_name, heading_rows, data_rows, page_num))
        else:
            merged.append((sec_name, heading_rows, data_rows))

    # Fallback: if pdfplumber got no real content, try PyMuPDF (many PDFs expose text to fitz only)
    has_real_data = any(_section_has_data(s[2]) for s in merged)
    if not has_real_data:
        pymupdf_sections = _extract_sections_from_pdf_pymupdf(pdf_path, max_pages=max_pages)
        if pymupdf_sections:
            log.info("Using PyMuPDF fallback (pdfplumber had no extractable text)")
            merged = [s for s in pymupdf_sections if _section_has_data(s[2])]

    # Multi-extractor comparison: run Camelot and Tabula, score all candidates, pick best
    candidates = [("pdfplumber/PyMuPDF", merged)]
    camelot_sections = _extract_sections_from_pdf_camelot(pdf_path, max_pages=max_pages)
    if camelot_sections:
        candidates.append(("Camelot", camelot_sections))
    tabula_sections = _extract_sections_from_pdf_tabula(pdf_path, max_pages=max_pages)
    if tabula_sections:
        candidates.append(("Tabula", tabula_sections))

    best_name = candidates[0][0]
    best_sections = candidates[0][1]
    best_score = _score_sections(best_sections)
    for name, secs in candidates[1:]:
        score = _score_sections(secs)
        if score > best_score:
            best_score = score
            best_name = name
            best_sections = secs
    log.info("Extractor comparison: %s selected (score=%.0f)", best_name, best_score)
    if max_pages is not None:

        def _page_within_cap(s: tuple) -> bool:
            if len(s) < 4:
                return True
            p = s[3]
            if p is None:
                return True
            try:
                return int(p) <= max_pages
            except (TypeError, ValueError):
                return True

        best_sections = [s for s in best_sections if _page_within_cap(s)]
    return best_sections


def _cell_to_json(c):
    """One cell to a JSON-safe value. Decimal -> float for JSON."""
    if c is None:
        return None
    if isinstance(c, Decimal):
        return float(c)
    if isinstance(c, (int, float)) and not isinstance(c, bool):
        return c
    s = str(c).strip()
    return s if s else None


def _build_header_grid(rows: list[list]) -> dict | None:
    """
    Build row/column header view from a table: first row = column_headers, first column = row_headers.
    Returns { "column_headers": [...], "row_headers": [...], "data": [[...], ...] } so that
    value at (row_headers[i], column_headers[j]) = data[i][j]. Enables (x, y) lookup by heading names.
    """
    if not rows or len(rows) < 2:
        return None
    first_row = rows[0]
    col_headers = [_cell_to_json(c) for c in (first_row if isinstance(first_row, (list, tuple)) else [first_row])]
    if not col_headers or all(c is None or (isinstance(c, str) and not c.strip()) for c in col_headers):
        return None
    row_headers = []
    data = []
    for r in rows[1:]:
        row = r if isinstance(r, (list, tuple)) else [r]
        row_headers.append(_cell_to_json(row[0]) if row else None)
        data.append([_cell_to_json(row[j]) if j < len(row) else None for j in range(1, len(col_headers) + 1)])
    return {"column_headers": col_headers, "row_headers": row_headers, "data": data}


def _sections_to_json_serializable(sections: list[tuple]) -> list[dict]:
    """Turn (name, heading_rows, data_rows[, page_num]) into list of dicts safe for JSON (no tuples, consistent types)."""
    out = []
    for s in sections:
        sec_name, heading_rows, data_rows = s[0], s[1], s[2]
        page_num = s[3] if len(s) >= 4 else None
        data_rows = _drop_page_number_rows(data_rows or [])
        # Headings: list of strings (first line can be section title)
        headings = []
        for h in (heading_rows or []):
            if isinstance(h, (list, tuple)):
                headings.append(" ".join(str(x) for x in h if x is not None))
            else:
                headings.append(str(h) if h is not None else "")
        # Rows: list of lists; cells as numbers or strings
        rows = []
        for r in (data_rows or []):
            row = r if isinstance(r, (list, tuple)) else [r]
            # Merge common fragmentation patterns early (reduces cut words/decimals).
            row = _merge_fragmented_row(list(row))
            cells = [_cell_to_json(c) for c in row]
            rows.append(cells)
        section_dict = {"name": str(sec_name), "headings": headings, "rows": rows}
        if page_num is not None:
            section_dict["page"] = page_num
        # Grid size for verification: how many rows/columns were extracted
        section_dict["row_count"] = len(rows)
        section_dict["column_count"] = len(rows[0]) if rows else 0
        # Add header-grid view for (row_heading, column_heading) → value lookup
        # Canonical key is (row_index, column_index); headers are labels and may repeat
        grid = _build_header_grid(rows)
        if grid:
            section_dict["column_headers"] = grid["column_headers"]
            section_dict["row_headers"] = grid["row_headers"]
            section_dict["data"] = grid["data"]
        out.append(section_dict)
    return out


def _json_row_is_narrative_noise(row: list) -> bool:
    """True for long text rows with no numeric cells (aligns with Excel prose pruning)."""
    if not row:
        return False
    cells = [c for c in row if c is not None and str(c).strip() != ""]
    if not cells:
        return False
    if any(_cell_looks_numeric(c) for c in cells):
        return False
    text_cells = [str(c).strip() for c in cells]
    longest = max(len(t) for t in text_cells)
    return longest > 45 and len(text_cells) <= 2


def refine_json_sections(sections: list[dict]) -> list[dict]:
    """Drop obvious narrative rows from each section; refresh row/column counts and header grid."""
    out: list[dict] = []
    for sec in sections:
        if not isinstance(sec, dict):
            out.append(sec)
            continue
        rows = sec.get("rows") or []
        if not isinstance(rows, list):
            out.append(sec)
            continue
        kept = [r for r in rows if isinstance(r, list) and not _json_row_is_narrative_noise(r)]
        new_sec = dict(sec)
        new_sec["rows"] = kept
        new_sec["row_count"] = len(kept)
        new_sec["column_count"] = len(kept[0]) if kept else 0
        grid = _build_header_grid(kept)
        if grid:
            new_sec["column_headers"] = grid["column_headers"]
            new_sec["row_headers"] = grid["row_headers"]
            new_sec["data"] = grid["data"]
        else:
            new_sec.pop("column_headers", None)
            new_sec.pop("row_headers", None)
            new_sec.pop("data", None)
        out.append(new_sec)
    return out


def _looks_like_table(sec_name: str, heading_rows: list, data_rows: list) -> bool:
    """
    True if section looks like a table (rows/columns with values). Drops long prose and junk.
    Keeps: multiple columns, or multiple rows with values; drops sections that are mostly paragraphs.
    """
    if not data_rows:
        return False
    total_cells = 0
    max_cells_in_row = 0
    has_numeric = False
    long_single_cell = 0
    rows_with_long_text = 0  # row with any cell > 100 chars (prose)
    for r in data_rows:
        row = r if isinstance(r, (list, tuple)) else [r]
        nc = len([c for c in row if c is not None and str(c).strip()])
        total_cells += nc
        max_cells_in_row = max(max_cells_in_row, nc)
        for c in row:
            if c is not None and _cell_looks_numeric(c):
                has_numeric = True
            if c is not None:
                s = str(c).strip()
                if len(s) > 200:
                    long_single_cell += 1
                if len(s) > 100:
                    rows_with_long_text += 1
                    break
    # Drop if it's just one long paragraph (single-cell section with 200+ char)
    if max_cells_in_row <= 1 and long_single_cell > 0:
        return False
    # Drop if mostly long paragraphs (e.g. disclaimer blocks)
    if rows_with_long_text > len(data_rows) / 2:
        return False
    # Keep if structured: at least 2 cells in some row, or 2+ rows with at least one value
    if max_cells_in_row >= 2:
        return True
    if len(data_rows) >= 2 and total_cells >= 2:
        return True
    if len(data_rows) == 1 and total_cells >= 1 and has_numeric:
        return True  # single row with label + value(s)
    return False


def filter_sections_to_tables_only(sections: list[tuple]) -> list[tuple]:
    """Keep only sections that look like tables (drop long prose and non-table content). Preserves page num if present (4-tuple)."""
    out = []
    for s in sections:
        sec_name, heading_rows, data_rows = s[0], s[1], s[2]
        if _looks_like_table(sec_name, heading_rows, data_rows):
            if len(s) >= 4:
                out.append((sec_name, heading_rows, data_rows, s[3]))
            else:
                out.append((sec_name, heading_rows, data_rows))
        else:
            log.debug("skip non-table section=%s rows=%d", (sec_name or "")[:40], len(data_rows or []))
    return out


def _write_json_from_sections(
    sections: list[tuple],
    out: Path,
    overwrite: bool = True,
    meta: dict | None = None,
) -> None:
    """Write sections to a JSON file (used for one-stream Excel + JSON output)."""
    if out.exists() and not overwrite:
        return
    out.parent.mkdir(parents=True, exist_ok=True)
    # Serialize and refine to reduce narrative noise and improve downstream usability.
    payload_sections = _sections_to_json_serializable(sections)
    payload_sections = refine_json_sections(payload_sections)
    payload = {"sections": payload_sections}
    payload["meta"] = meta if isinstance(meta, dict) else {}
    try:
        evaluation = evaluate_extraction_json_correctness(payload)
        payload["meta"].update(
            {
                "status": evaluation.get("status"),
                "requires_review": evaluation.get("requires_review"),
                "quality_score": evaluation.get("quality_score"),
                "validation_errors": evaluation.get("errors"),
                "validation_warnings": evaluation.get("warnings"),
            }
        )
    except Exception:
        # Validation meta must never break JSON writing.
        pass
    with open(out, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
    log.info("Wrote %d section(s) to %s", len(sections), out)
    for s in sections:
        log.debug("json section=%s rows=%d", (s[0] or "")[:50], len((s[2] or [])))


# Schema for extraction JSON (sections with name, headings, rows). Used to validate before Excel.
_EXTRACTION_JSON_SCHEMA = {
    "type": "object",
    "required": ["sections"],
    "properties": {
        "sections": {
            "type": "array",
            "items": {
                "type": "object",
                "required": ["name", "headings", "rows"],
                "properties": {
                    "name": {"type": "string"},
                    "headings": {"type": "array", "items": {"type": "string"}},
                    "rows": {
                        "type": "array",
                        "items": {
                            "type": "array",
                            "items": {"type": ["string", "number", "null"]},
                        },
                    },
                    "row_count": {"type": "integer"},
                    "column_count": {"type": "integer"},
                },
                "additionalProperties": True,
            },
        },
    },
    "additionalProperties": True,
}


def validate_extraction_json(payload: dict) -> None:
    """
    Validate extraction JSON against schema. Raises ValueError if invalid.
    Ensures sections have name, headings, rows so we don't write malformed data to Excel.
    """
    if jsonschema is None:
        return
    try:
        jsonschema.validate(instance=payload, schema=_EXTRACTION_JSON_SCHEMA)
    except jsonschema.ValidationError as e:
        raise ValueError(f"Extraction JSON validation failed: {e}") from e


def evaluate_extraction_json_correctness(
    payload: dict,
    *,
    numeric_tol: float = 0.01,
) -> dict:
    """
    Deep correctness validator for extraction JSON.

    This is stricter than schema validation:
    - checks row/column "tally" consistency (all rows match header-row width)
    - checks `row_count`/`column_count` if present
    - if the optional header-grid is present (`column_headers`, `row_headers`, `data`),
      verifies its dimensions and that it matches the `rows` matrix by (row_index, col_index)

    Returns dict with:
      - status: "ok" | "requires_review" | "failed"
      - requires_review: bool
      - quality_score: float in [0, 1]
      - errors: list[str]
      - warnings: list[str]
    """

    def _cell_normalize_for_compare(c):
        # Normalize using existing project logic so numeric strings and numeric types
        # compare consistently across different pipelines.
        v = _cell_value(c)
        if v == "":
            return None
        if isinstance(v, (int, float, Decimal)) and not isinstance(v, bool):
            return ("num", Decimal(str(v)))
        return ("str", str(v).strip())

    def _cells_equal(a, b) -> bool:
        na = _cell_normalize_for_compare(a)
        nb = _cell_normalize_for_compare(b)
        if na is None and nb is None:
            return True
        if na is None or nb is None:
            return False
        if na[0] == "num" and nb[0] == "num":
            try:
                diff = abs(na[1] - nb[1])
                return diff <= Decimal(str(numeric_tol))
            except Exception:
                return False
        if na[0] != nb[0]:
            return False
        return na[1] == nb[1]

    errors: list[str] = []
    warnings: list[str] = []

    if not isinstance(payload, dict):
        return {
            "status": "failed",
            "requires_review": True,
            "quality_score": 0.0,
            "errors": ["payload is not a dict"],
            "warnings": [],
        }

    sections = payload.get("sections")
    if not isinstance(sections, list):
        return {
            "status": "failed",
            "requires_review": True,
            "quality_score": 0.0,
            "errors": ["payload.sections missing or not a list"],
            "warnings": [],
        }

    if not sections:
        warnings.append("no sections found in payload")

    for sec in sections:
        if not isinstance(sec, dict):
            errors.append("section is not an object/dict")
            continue

        sec_name = str(sec.get("name") or "")
        rows = sec.get("rows")
        if not isinstance(rows, list):
            errors.append(f"section={sec_name[:60]!r}: rows missing or not a list")
            continue

        row_count_actual = len(rows)
        if row_count_actual == 0:
            # Not necessarily wrong (some documents have empty sections),
            # but flag for review.
            warnings.append(f"section={sec_name[:60]!r}: empty rows")
            # Still check row_count/column_count if present.
            rc = sec.get("row_count")
            cc = sec.get("column_count")
            if rc is not None and isinstance(rc, int) and rc != 0:
                errors.append(f"section={sec_name[:60]!r}: row_count={rc} != 0")
            if cc is not None and isinstance(cc, int) and cc != 0:
                errors.append(f"section={sec_name[:60]!r}: column_count={cc} != 0")
            continue

        # Header row width is the canonical column width.
        header_row = rows[0]
        header_len = len(header_row) if isinstance(header_row, list) else 0
        if header_len <= 0:
            errors.append(f"section={sec_name[:60]!r}: header row invalid/empty")
            continue

        # Verify `rows` are lists and have consistent width.
        for i, r in enumerate(rows):
            if not isinstance(r, list):
                errors.append(f"section={sec_name[:60]!r}: row_index={i} is not a list")
                continue
            if len(r) != header_len:
                errors.append(
                    f"section={sec_name[:60]!r} row_index={i}: row_len={len(r)} != header_len={header_len}"
                )

        # Verify row_count/column_count fields when present.
        if "row_count" in sec and isinstance(sec.get("row_count"), int):
            if sec["row_count"] != row_count_actual:
                errors.append(f"section={sec_name[:60]!r}: row_count={sec['row_count']} != len(rows)={row_count_actual}")
        if "column_count" in sec and isinstance(sec.get("column_count"), int):
            if sec["column_count"] != header_len:
                errors.append(
                    f"section={sec_name[:60]!r}: column_count={sec['column_count']} != header_len={header_len}"
                )

        # Optional deep check: verify the precomputed header-grid matches `rows`.
        grid_col_headers = sec.get("column_headers")
        grid_row_headers = sec.get("row_headers")
        grid_data = sec.get("data")
        if grid_col_headers is not None or grid_row_headers is not None or grid_data is not None:
            if not isinstance(grid_col_headers, list) or not isinstance(grid_row_headers, list) or not isinstance(grid_data, list):
                errors.append(
                    f"section={sec_name[:60]!r}: grid keys present but types invalid "
                    f"(column_headers/row_headers/data must be lists)"
                )
            else:
                expected_col_count = header_len
                expected_row_header_count = max(0, row_count_actual - 1)
                if len(grid_col_headers) != expected_col_count:
                    errors.append(
                        f"section={sec_name[:60]!r}: len(column_headers)={len(grid_col_headers)} != column_count={expected_col_count}"
                    )
                if len(grid_row_headers) != expected_row_header_count:
                    errors.append(
                        f"section={sec_name[:60]!r}: len(row_headers)={len(grid_row_headers)} != row_count-1={expected_row_header_count}"
                    )
                if len(grid_data) != expected_row_header_count:
                    errors.append(
                        f"section={sec_name[:60]!r}: len(data)={len(grid_data)} != row_count-1={expected_row_header_count}"
                    )

                # data is expected to be [row_count-1][column_count] where data[ri][ci] == rows[ri+1][ci+1]
                # (this follows the project’s own `_build_header_grid` indexing).
                for ri in range(expected_row_header_count):
                    if ri >= len(grid_data):
                        continue
                    row_data = grid_data[ri]
                    if not isinstance(row_data, list):
                        errors.append(f"section={sec_name[:60]!r}: data row_index={ri} is not a list")
                        continue
                    if len(row_data) != expected_col_count:
                        errors.append(
                            f"section={sec_name[:60]!r}: data[{ri}] len={len(row_data)} != column_count={expected_col_count}"
                        )
                        continue
                    for ci in range(expected_col_count):
                        expected_cell = None
                        # rows[ri+1] is the original table row
                        if ri + 1 < len(rows) and (ci + 1) < len(rows[ri + 1]):
                            expected_cell = rows[ri + 1][ci + 1]
                        actual_cell = row_data[ci] if ci < len(row_data) else None
                        if not _cells_equal(expected_cell, actual_cell):
                            errors.append(
                                f"section={sec_name[:60]!r}: grid/data mismatch at (row={ri}, col={ci})"
                            )
                            break

                # Check row_headers/column_headers labels match the corresponding `rows` cells.
                for ci in range(expected_col_count):
                    if ci < len(grid_col_headers) and ci < len(header_row):
                        if not _cells_equal(header_row[ci], grid_col_headers[ci]):
                            errors.append(f"section={sec_name[:60]!r}: column_headers mismatch at index={ci}")
                            break
                for ri in range(expected_row_header_count):
                    if ri < len(grid_row_headers) and (ri + 1) < len(rows):
                        if not _cells_equal(rows[ri + 1][0] if rows[ri + 1] else None, grid_row_headers[ri]):
                            errors.append(f"section={sec_name[:60]!r}: row_headers mismatch at row_index={ri}")
                            break

    # Simple warnings (quality heuristics). Keep these low-noise for unattended runs.
    if not errors:
        # If everything has very few numeric cells, it’s suspicious.
        total_cells = 0
        numeric_cells = 0
        for sec in sections:
            if not isinstance(sec, dict):
                continue
            rows = sec.get("rows")
            if not isinstance(rows, list) or not rows:
                continue
            for r in rows:
                if not isinstance(r, list):
                    continue
                for c in r:
                    total_cells += 1
                    if c is not None and _cell_looks_numeric(c):
                        numeric_cells += 1
        if total_cells > 0:
            numeric_ratio = numeric_cells / total_cells
            if numeric_ratio < 0.005:
                warnings.append(f"low numeric density (numeric_ratio={numeric_ratio:.4f})")

    # Compute status/score.
    if errors:
        requires_review = True
        status = "failed"
        quality_score = 0.0
    elif warnings:
        requires_review = True
        status = "requires_review"
        quality_score = max(0.0, 1.0 - 0.05 * len(warnings))
    else:
        requires_review = False
        status = "ok"
        quality_score = 1.0

    return {
        "status": status,
        "requires_review": requires_review,
        "quality_score": float(quality_score),
        "errors": errors,
        "warnings": warnings,
    }


def _clean_loaded_cell(c):
    """Data cleaning: trim strings, leave numbers/null as-is. Used after load from JSON."""
    if c is None:
        return None
    if isinstance(c, (int, float, Decimal)) and not isinstance(c, bool):
        return c
    s = str(c).strip()
    return s if s else None


def _is_numeric_cell(c) -> bool:
    """True if cell value is numeric (int, float, Decimal)."""
    return isinstance(c, (int, float, Decimal)) and not isinstance(c, bool)


def validate_sections(
    sections: list[tuple[str, list, list]],
) -> tuple[list[str], bool]:
    """
    Run validation on extracted sections. Returns (list of error messages, requires_review).
    - Inconsistent column count per section -> error + needs_review.
    - Optional: Total row sum vs sum of detail rows (when detectable).
    """
    errors = []
    requires_review = False
    for s in sections:
        sec_name, _headings, data_rows = s[0], s[1], s[2]
        if not data_rows:
            continue
        header_len = len(data_rows[0]) if data_rows[0] else 0
        for i, row in enumerate(data_rows[1:], start=1):
            row_len = len(row) if isinstance(row, (list, tuple)) else 0
            if header_len and row_len != header_len:
                err = f"section={sec_name[:40]!r} row={i+1} column_count={row_len} != header_count={header_len}"
                errors.append(err)
                requires_review = True
    return (errors, requires_review)


def _clean_loaded_sections(sections: list[tuple]) -> list[tuple]:
    """
    Data cleaning layer: trim string cells in loaded sections before Excel.
    Reduces noise from extra spaces; keeps numbers and structure unchanged. Preserves page if 4-tuple.
    """
    out = []
    for s in sections:
        name, heading_rows, data_rows = s[0], s[1], s[2]
        clean_headings = [_clean_loaded_cell(h) or "" for h in (heading_rows or [])]
        clean_rows = []
        for r in data_rows or []:
            row = r if isinstance(r, (list, tuple)) else [r]
            clean_rows.append([_clean_loaded_cell(c) for c in row])
        if len(s) >= 4:
            out.append((name, clean_headings, clean_rows, s[3]))
        else:
            out.append((name, clean_headings, clean_rows))
    return out


def load_sections_from_json(json_path: str | Path) -> list[tuple[str, list, list]]:
    """
    Load sections from a JSON file (same format we write).
    Validates payload against schema, then applies data cleaning (trim strings). Returns (section_name, heading_rows, data_rows).
    """
    path = Path(json_path)
    if not path.exists():
        raise FileNotFoundError(f"JSON not found: {path}")
    with open(path, encoding="utf-8") as f:
        payload = json.load(f)
    validate_extraction_json(payload)
    sections_data = payload.get("sections") or []
    out = []
    for sec in sections_data:
        name = (sec.get("name") or "").strip()
        headings = sec.get("headings") or []
        heading_rows = [str(h).strip() for h in headings] if isinstance(headings, list) else []
        rows = sec.get("rows") or []
        data_rows = [r if isinstance(r, (list, tuple)) else [r] for r in rows]
        page = sec.get("page")
        if page is not None:
            out.append((name, heading_rows, data_rows, int(page)))
        else:
            out.append((name, heading_rows, data_rows))
    return _clean_loaded_sections(out)


def _normalize_page_to_sheet(mapping: dict) -> dict[int, str]:
    """Convert page_to_sheet from config/JSON (keys may be int or str) to dict[int, str]."""
    if not mapping:
        return {}
    out = {}
    for k, v in mapping.items():
        try:
            p = int(k)
        except (TypeError, ValueError):
            continue
        if v and isinstance(v, str):
            out[p] = v.strip()
    return out


def write_sections_to_workbook_by_page(
    sections: list[tuple],
    page_to_sheet: dict[int, str],
    out_path: Path,
) -> int:
    """
    Build Excel with one sheet per document section, using page numbers to group.
    page_to_sheet maps page number -> sheet name (e.g. {2: "General Information", 3: "Overview", 4: "US Tax Summary", 5: "US Tax Summary"}).
    Sections with the same target sheet are written in document order; sections without page use their section name as sheet.
    Returns number of sheets written.
    """
    wb = Workbook()
    wb.remove(wb.active)
    if not sections:
        ws = wb.create_sheet(title="Info")
        ws.append(["No sections in this data."])
        wb.save(out_path)
        return 1

    def _is_prose_noise_row(row: list) -> bool:
        """
        True for narrative/junk lines that pollute numeric sheets.
        Keep short label rows (e.g. 'Cash', 'Total') but drop long prose rows with no numbers.
        """
        if not row:
            return False
        cells = [c for c in row if c is not None and str(c).strip() != ""]
        if not cells:
            return False
        text_cells = [str(c).strip() for c in cells]
        if any(_cell_looks_numeric(c) for c in cells):
            return False
        longest = max(len(t) for t in text_cells)
        # Narrative rows are usually long and sparse.
        return longest > 45 and len(text_cells) <= 2

    def _prune_prose_rows(data_rows: list[list]) -> list[list]:
        return [r for r in data_rows if not _is_prose_noise_row(r)]

    # Group sections by target sheet name (by page or fallback to section name)
    groups: dict[str, list[tuple]] = {}  # sheet_name -> [(sec_name, heading_rows, data_rows, bold_indices?), ...]
    sheet_order: list[str] = []  # preserve order by first occurrence

    for idx, s in enumerate(sections):
        sec_name, heading_rows, data_rows = s[0], s[1], s[2]
        page_num = s[3] if len(s) >= 4 and s[3] is not None else None
        bold_indices = s[4] if len(s) >= 5 else None
        if page_num is not None and page_num in page_to_sheet:
            sheet_name = page_to_sheet[page_num]
        elif page_num is not None:
            preferred = _preferred_sheet_name_from_title(sec_name)
            if not preferred and heading_rows:
                preferred = _preferred_sheet_name_from_title(
                    heading_rows[0] if isinstance(heading_rows[0], str) else ""
                )
            # Avoid generic "Page N" unless we truly cannot infer a meaningful sheet.
            if preferred:
                sheet_name = _safe_sheet_name(preferred, idx + 1)
            else:
                sheet_name = _safe_sheet_name(f"Page {page_num}", idx + 1)
        else:
            sheet_name = _safe_sheet_name(sec_name or f"Section_{idx + 1}", idx + 1)
        if sheet_name not in groups:
            groups[sheet_name] = []
            sheet_order.append(sheet_name)
        cleaned_rows = _prune_prose_rows(data_rows)
        groups[sheet_name].append((sec_name, heading_rows, cleaned_rows, bold_indices))

    for sheet_name in sheet_order:
        group_sections = groups[sheet_name]
        ws = wb.create_sheet(title=sheet_name[:31])
        row_num = 1
        for i, (sec_name, heading_rows, data_rows, bold_indices) in enumerate(group_sections):
            row_num = _write_section_to_sheet(
                ws, sec_name, heading_rows, data_rows, start_row=row_num, bold_data_row_indices=bold_indices
            )
            if i < len(group_sections) - 1:
                row_num += 1
        log.debug("excel sheet=%s sections=%d", sheet_name, len(group_sections))

    wb.save(out_path)
    return len(groups)


def _write_sections_to_workbook(
    sections: list[tuple],
    out_path: Path,
    single_sheet: bool = False,
) -> int:
    """Build an Excel workbook from sections (same logic as pdf_tables_to_excel). Sections may be 3- or 4-tuples (page ignored). Returns number of sheets."""
    wb = Workbook()
    wb.remove(wb.active)
    if not sections:
        ws = wb.create_sheet(title="Info")
        ws.append(["No sections in this data."])
        wb.save(out_path)
        return 1
    if single_sheet:
        ws = wb.create_sheet(title="Extracted")
        next_row = 1
        for s in sections:
            sec_name, heading_rows, data_rows = s[0], s[1], s[2]
            page = s[3] if len(s) >= 4 else None
            bold_indices = s[4] if len(s) >= 5 else None
            if page is not None:
                heading_rows = (["Page"] + list(heading_rows)) if heading_rows else ["Page"]
                data_rows = [[page] + list(row) for row in data_rows]
            next_row = _write_section_to_sheet(ws, sec_name, heading_rows, data_rows, start_row=next_row, bold_data_row_indices=bold_indices)
            next_row += 2
        sheet_count = 1
    else:
        # Group sections by sheet name so "Holdings (Continued)", "Holdings (Continued)1" → one "Holdings" sheet
        groups = {}  # sheet_key -> list of (sec_name, heading_rows, data_rows, bold_indices, page)
        for idx, s in enumerate(sections):
            sec_name, heading_rows, data_rows = s[0], s[1], s[2]
            page = s[3] if len(s) >= 4 else None
            preferred = _preferred_sheet_name_from_title(sec_name)
            if not preferred and heading_rows:
                preferred = _preferred_sheet_name_from_title(
                    heading_rows[0] if isinstance(heading_rows[0], str) else ""
                )
            if preferred:
                sheet_key = _safe_sheet_name(preferred, idx + 1)
            else:
                # Use normalized name so continuations merge even without a pattern match
                norm = _normalize_section_name_for_lookup(sec_name)
                sheet_key = _safe_sheet_name(norm or sec_name, idx + 1)
            if sheet_key not in groups:
                groups[sheet_key] = []
            bold_indices = s[4] if len(s) >= 5 else None
            groups[sheet_key].append((sec_name, heading_rows, data_rows, bold_indices, page))

        used_sheet_names = {}
        for sheet_key, group_sections in groups.items():
            base = sheet_key
            count = used_sheet_names.get(base, 0) + 1
            used_sheet_names[base] = count
            if count > 1:
                suffix = f" {count}"
                base = (base.strip() or base)[: 31 - len(suffix)]
                name = _safe_sheet_name(f"{base}{suffix}", 0)
            else:
                name = base
            ws = wb.create_sheet(title=name[:31])
            row_num = 1
            for i, group_item in enumerate(group_sections):
                sec_name, heading_rows, data_rows = group_item[0], group_item[1], group_item[2]
                bold_indices = group_item[3] if len(group_item) >= 4 else None
                page = group_item[4] if len(group_item) >= 5 else None
                if page is not None:
                    heading_rows = (["Page"] + list(heading_rows)) if heading_rows else ["Page"]
                    data_rows = [[page] + list(row) for row in data_rows]
                row_num = _write_section_to_sheet(
                    ws, sec_name, heading_rows, data_rows, start_row=row_num, bold_data_row_indices=bold_indices
                )
                # Blank row between merged sections (except after the last)
                if i < len(group_sections) - 1:
                    row_num += 1
            log.debug("excel sheet=%s sections=%d", name, len(group_sections))
        sheet_count = len(groups)
    wb.save(out_path)
    return sheet_count


def json_to_excel(
    json_path: str | Path,
    output_path: str | Path | None = None,
    overwrite: bool = True,
    single_sheet: bool = False,
) -> str:
    """
    Produce an Excel file from a JSON file (sections + rows).
    Pipeline: JSON is the canonical intermediate; Excel is built from it.
    """
    json_path = Path(json_path)
    if not json_path.exists():
        raise FileNotFoundError(f"JSON not found: {json_path}")
    out = Path(output_path or json_path.with_suffix(".xlsx"))
    if out.exists() and not overwrite:
        raise FileExistsError(f"Output exists: {out}")
    out.parent.mkdir(parents=True, exist_ok=True)
    sections = load_sections_from_json(json_path)
    _write_sections_to_workbook(sections, out, single_sheet=single_sheet)
    log.info("Wrote Excel from JSON: %s", out)
    return str(out)


def pdf_to_json(
    pdf_path: str,
    output_path: str | None = None,
    overwrite: bool = True,
) -> str:
    """
    Extract all tables/sections from the PDF and write them to a JSON file.
    Structure: { "sections": [ { "name": "...", "headings": [...], "rows": [[...], ...] }, ... ] }
    Same data as Excel, so you can edit/filter the JSON then convert to Excel or QB format later.
    """
    pdf_path = Path(pdf_path)
    out = Path(output_path or pdf_path.with_suffix(".json"))
    if out.exists() and not overwrite:
        raise FileExistsError(f"Output exists (use --overwrite to replace): {out}")

    sections = extract_sections_from_pdf(pdf_path)
    page_count = None
    try:
        import fitz  # PyMuPDF

        with fitz.open(str(pdf_path)) as doc:
            page_count = len(doc)
    except Exception:
        page_count = None
    out.parent.mkdir(parents=True, exist_ok=True)
    _write_json_from_sections(
        sections,
        out,
        overwrite,
        meta={
            "pdf_name": pdf_path.name,
            "pdf_path": str(pdf_path),
            "page_count": page_count,
            "generated_at_utc": datetime.utcnow().isoformat(timespec="seconds") + "Z",
            "generator": "tables_to_excel.pdf_to_json",
        },
    )
    return str(out)


def pdf_tables_to_excel(
    pdf_path: str,
    output_path: str | None = None,
    overwrite: bool = True,
    single_sheet: bool = False,
    write_json: bool = False,
    json_path: str | Path | None = None,
) -> str:
    """
    Extract every table from the PDF and write to one Excel file.

    By default (single_sheet=False), creates one sheet per section so each sheet mirrors
    the PDF: a section title, optional summary (key-value lines as a 2-column table),
    then the main data table with header row and columns. Organized and readable.

    If single_sheet=True, writes one sheet "Extracted" with all sections in that same
    format (title, summary, table, blank rows, next section...).

    If write_json=True and json_path is set, writes the same extraction to a JSON file
    (one stream: one extraction → Excel + JSON).
    """
    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        raise FileNotFoundError(f"Not found: {pdf_path}")
    if pdf_path.suffix.lower() != ".pdf":
        raise ValueError("File must be a .pdf")

    out = Path(output_path or pdf_path.with_suffix(".xlsx"))
    if out.exists() and not overwrite:
        raise FileExistsError(f"Output exists (use --overwrite to replace): {out}")

    out.parent.mkdir(parents=True, exist_ok=True)

    try:
        sections = extract_sections_from_pdf(pdf_path)
    except Exception as e:
        msg = str(e).lower()
        if "password" in msg or "encrypted" in msg:
            raise ValueError("PDF appears password-protected or encrypted; not supported.") from e
        if "invalid" in msg or "cannot read" in msg or "failed" in msg:
            raise ValueError("PDF could not be read (corrupt or invalid file).") from e
        raise

    if not sections:
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title="Info")
        ws.append(["No tables or structured data could be extracted from this PDF."])
        wb.save(out)
        if write_json and json_path:
            _write_json_from_sections(
                [],
                Path(json_path),
                overwrite,
                meta={
                    "pdf_name": pdf_path.name,
                    "pdf_path": str(pdf_path),
                    "page_count": None,
                    "generated_at_utc": datetime.utcnow().isoformat(timespec="seconds") + "Z",
                    "generator": "tables_to_excel.pdf_tables_to_excel",
                    "note": "No tables/structured data extracted; wrote empty sections.",
                },
            )
        log.info("Wrote 1 sheet to %s", out)
        return str(out)

    if write_json and json_path:
        _write_json_from_sections(
            sections,
            Path(json_path),
            overwrite,
            meta={
                "pdf_name": pdf_path.name,
                "pdf_path": str(pdf_path),
                "page_count": None,
                "generated_at_utc": datetime.utcnow().isoformat(timespec="seconds") + "Z",
                "generator": "tables_to_excel.pdf_tables_to_excel",
            },
        )
    sheet_count = _write_sections_to_workbook(sections, out, single_sheet=single_sheet)
    log.info("Wrote %d sheet(s) to %s", sheet_count, out)
    return str(out)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Extract all tables from a PDF into an Excel file (no API key required)."
    )
    parser.add_argument("pdf", help="Path to the PDF file")
    parser.add_argument("-o", "--output", default=None, help="Output .xlsx path")
    parser.add_argument("--no-overwrite", action="store_false", dest="overwrite", default=True, help="Do not overwrite; fail if output file already exists")
    parser.add_argument("--separate-sheets", action="store_true", dest="separate_sheets", help="One sheet per section (default: one sheet with all sections in order)")
    args = parser.parse_args()

    try:
        log.info("Input: %s", args.pdf)
        result = pdf_tables_to_excel(args.pdf, args.output, overwrite=args.overwrite, single_sheet=not args.separate_sheets)
        print(f"Saved: {result}")
        return 0
    except (FileNotFoundError, ValueError, FileExistsError) as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
