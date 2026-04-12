#!/usr/bin/env python3
"""
Single entry point for PDF → Excel.

  python run.py tables <pdf> [pdf2 ...]   Extract all tables to QB-format Excel. Batch: multiple PDFs → multiple .xlsx.
  python run.py json <pdf> [pdf2 ...]     Extract PDF to JSON first (sections + rows); then you can convert to Excel.
  python run.py ask <pdf> <query>         AI agent: extract what you ask for. Optional: multiple PDFs with same query.
  python run.py populate-family-template --template family_template.xlsx --jpm-pdf jpm.pdf --gs-pdf gs.pdf
                                            One workbook: JPM + Goldman fills in the same template output.

  Web UI: flask --app app run  →  http://127.0.0.1:8003  (port from .flaskenv)
"""

import argparse
import sys
from pathlib import Path

def _get_version():
    p = Path(__file__).resolve().parent / "VERSION"
    return p.read_text().strip() if p.exists() else "0.0.0"

# Set CUDA before any VL/llama_cpp use so GPU is used
try:
    from extract_vl import _ensure_cuda_path
    _ensure_cuda_path()
except ImportError:
    pass
# Project modules
from tables_to_excel import pdf_tables_to_excel, pdf_to_json, json_to_excel
from extract import extract_pdf_to_excel
from pdf_to_qb import pdf_to_qb_excel, transform_extracted_to_qb
from template_populator import populate_template_from_fields_json, populate_template_from_qb_output
from fields_from_qb_output import extract_fields
from pdf_json_audit import audit_pdf_vs_extraction_json, apply_audit_to_extraction_file
from scripts.populate_jpm_template_from_pdf import populate_jpm_template_from_pdf
from scripts.family_template_merge import populate_family_template


def _expand_pdfs(paths):
    """Expand paths to a list of .pdf files. If a path is a dir, add its .pdf children."""
    out = []
    for p in paths:
        path = Path(p)
        if not path.exists():
            raise FileNotFoundError(f"Not found: {path}")
        if path.is_file():
            if path.suffix.lower() == ".pdf":
                out.append(path)
            else:
                raise ValueError(f"Not a PDF: {path}")
        else:
            for f in sorted(path.glob("*.pdf")):
                out.append(f)
    return out


def _default_output_dir():
    """Default output directory (e.g. output/). Can be overridden via OUTPUT_DIR in .env."""
    try:
        from dotenv import load_dotenv
        import os
        load_dotenv()
        return os.environ.get("OUTPUT_DIR", "output")
    except Exception:
        return "output"


def cmd_tables(args) -> int:
    overwrite = not args.no_overwrite
    pdfs = _expand_pdfs(args.pdfs)
    if not pdfs:
        print("Error: No PDF files found.", file=sys.stderr)
        return 1
    if args.output and len(pdfs) > 1:
        print("Error: -o/--output only allowed for a single PDF.", file=sys.stderr)
        return 1
    default_dir = _default_output_dir()
    for i, pdf in enumerate(pdfs):
        if args.output and len(pdfs) == 1:
            out = args.output
        else:
            out = str(Path(default_dir) / Path(pdf).with_suffix(".xlsx").name)
        if len(pdfs) > 1:
            print(f"[{i+1}/{len(pdfs)}] {pdf}")
        try:
            out_path = Path(out)
            result = pdf_to_qb_excel(
                str(pdf),
                out,
                overwrite=overwrite,
                json_path_out=str(out_path.with_suffix(".json")),
            )
            print(f"Saved: {result}")
            if out_path.with_suffix(".json").exists():
                print(f"Saved: {out_path.with_suffix('.json')}")
                if not getattr(args, "no_audit", False):
                    _, ok = apply_audit_to_extraction_file(
                        Path(pdf),
                        out_path.with_suffix(".json"),
                        audit_pages=getattr(args, "audit_pages", None),
                        strict=getattr(args, "audit_strict", False),
                        report_path=Path(getattr(args, "audit_report", "")).expanduser()
                        if getattr(args, "audit_report", None)
                        else None,
                    )
                    if getattr(args, "audit_strict", False) and not ok:
                        return 2
        except Exception as e:
            print(f"Error: {e}", file=sys.stderr)
            return 1
    return 0


def cmd_clean_json(args) -> int:
    """Remove repetitive sections from extraction JSON. When --pdf is given, only collapse/drop when the repeated phrase appears few times on that PDF page (compare with source). Overwrites the file."""
    import json as _json
    from extract_vl import _drop_repetitive_sections, pdf_phrase_count_for_file
    json_path = Path(args.json_file)
    if not json_path.exists():
        print(f"Error: File not found: {json_path}", file=sys.stderr)
        return 1
    with open(json_path, encoding="utf-8") as f:
        payload = _json.load(f)
    sections = payload.get("sections") or []
    if not sections:
        print("No sections to clean.", file=sys.stderr)
        return 0
    pdf_path = getattr(args, "pdf", None)
    if not pdf_path:
        meta = payload.get("meta") or {}
        if meta.get("pdf_path") and Path(meta["pdf_path"]).exists():
            pdf_path = meta["pdf_path"]
        elif meta.get("pdf_name"):
            candidate = json_path.parent / meta["pdf_name"]
            if candidate.exists():
                pdf_path = candidate
    pdf_phrase_count = None
    if pdf_path and Path(pdf_path).exists():
        pdf_phrase_count = pdf_phrase_count_for_file(pdf_path)
        if pdf_phrase_count:
            print(f"Using PDF for comparison: {pdf_path}")
    before = len(sections)
    sections = _drop_repetitive_sections(sections, pdf_phrase_count=pdf_phrase_count)
    after = len(sections)
    payload["sections"] = sections
    with open(json_path, "w", encoding="utf-8") as f:
        _json.dump(payload, f, indent=2, ensure_ascii=False)
    print(f"Cleaned: {before} -> {after} sections (removed {before - after} repetitive). Saved: {json_path}")
    return 0


def cmd_from_json(args) -> int:
    """Convert a JSON file (from pdf→json) to Excel. Use after editing JSON to map tables correctly.
    If the JSON has meta.page_to_sheet or config has page_to_sheet, sections are grouped by page
    into sheets (e.g. page 2 → General Information, page 3 → Overview, pages 4–5 → US Tax Summary).
    """
    import json as _json
    json_path = Path(args.json_file)
    if not json_path.exists():
        print(f"Error: File not found: {json_path}", file=sys.stderr)
        return 1
    if json_path.suffix.lower() != ".json":
        print("Error: File must be a .json file.", file=sys.stderr)
        return 1
    out = args.output or str(json_path.with_suffix(".xlsx"))
    out_path = Path(out)
    overwrite = not getattr(args, "no_overwrite", False)
    if out_path.exists() and not overwrite:
        print(f"Error: Output exists: {out_path} (use --overwrite to replace)", file=sys.stderr)
        return 1
    try:
        import tempfile
        from tables_to_excel import (
            load_sections_from_json,
            _write_sections_to_workbook,
            write_sections_to_workbook_by_page,
            _normalize_page_to_sheet,
        )
        sections = load_sections_from_json(json_path)
        # Use page_to_sheet only when explicitly present in this JSON's meta.
        page_to_sheet = {}
        with open(json_path, encoding="utf-8") as f:
            payload = _json.load(f)
        meta = payload.get("meta") or {}
        page_to_sheet = _normalize_page_to_sheet(meta.get("page_to_sheet") or {})
        has_page = any(len(s) >= 4 and s[3] is not None for s in sections)
        if page_to_sheet and has_page:
            write_sections_to_workbook_by_page(sections, page_to_sheet, out_path)
        else:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
                temp_xlsx = f.name
            try:
                _write_sections_to_workbook(sections, Path(temp_xlsx))
                transform_extracted_to_qb(temp_xlsx, str(out_path))
            finally:
                Path(temp_xlsx).unlink(missing_ok=True)
        print(f"Saved: {out_path}")
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1
    return 0


def cmd_hybrid(args) -> int:
    """Extract PDF with hybrid (library first, VL only on bad pages). Writes JSON with per-page VL timing in meta."""
    pdfs = _expand_pdfs(args.pdfs)
    if not pdfs:
        print("Error: No PDF files found.", file=sys.stderr)
        return 1
    if args.output and len(pdfs) > 1:
        print("Error: -o/--output only allowed for a single PDF.", file=sys.stderr)
        return 1
    default_dir = _default_output_dir()
    try:
        from hybrid_extract import hybrid_pdf_to_json
    except ImportError as e:
        print(f"Error: hybrid extraction requires hybrid_extract: {e}", file=sys.stderr)
        return 1
    for i, pdf in enumerate(pdfs):
        out = args.output if (args.output and len(pdfs) == 1) else str(Path(default_dir) / Path(pdf).with_suffix(".json").stem + "_hybrid.json")
        if len(pdfs) > 1:
            print(f"[{i+1}/{len(pdfs)}] {pdf}")
        try:
            result = hybrid_pdf_to_json(
                str(pdf),
                out,
                schema_type=getattr(args, "schema_type", None),
                max_pages=getattr(args, "max_pages", None),
                overwrite=not args.no_overwrite,
            )
            print(f"Saved: {result}")
            if not getattr(args, "no_audit", False):
                _, ok = apply_audit_to_extraction_file(
                    Path(pdf),
                    Path(result),
                    audit_pages=getattr(args, "audit_pages", None),
                    strict=getattr(args, "audit_strict", False),
                    report_path=Path(getattr(args, "audit_report", "")).expanduser()
                    if getattr(args, "audit_report", None)
                    else None,
                )
                if getattr(args, "audit_strict", False) and not ok:
                    return 2
            if getattr(args, "excel", False):
                import tempfile
                from tables_to_excel import (
                    load_sections_from_json,
                    _write_sections_to_workbook,
                    write_sections_to_workbook_by_page,
                    _normalize_page_to_sheet,
                )
                import json as _json
                xlsx_path = Path(result).with_suffix(".xlsx")
                if xlsx_path.exists() and args.no_overwrite:
                    print(f"Excel exists (skip): {xlsx_path}")
                else:
                    sections = load_sections_from_json(result)
                    with open(result, encoding="utf-8") as f:
                        payload = _json.load(f)
                    meta = payload.get("meta") or {}
                    page_to_sheet = _normalize_page_to_sheet(meta.get("page_to_sheet") or {})
                    has_page = any(len(s) >= 4 and s[3] is not None for s in sections)
                    if page_to_sheet and has_page:
                        write_sections_to_workbook_by_page(sections, page_to_sheet, xlsx_path)
                    else:
                        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
                            tmp = f.name
                        try:
                            _write_sections_to_workbook(sections, Path(tmp))
                            transform_extracted_to_qb(tmp, str(xlsx_path))
                        finally:
                            Path(tmp).unlink(missing_ok=True)
                    print(f"Saved: {xlsx_path}")
        except Exception as e:
            print(f"Error: {e}", file=sys.stderr)
            return 1
    return 0


def cmd_json(args) -> int:
    pdfs = _expand_pdfs(args.pdfs)
    if not pdfs:
        print("Error: No PDF files found.", file=sys.stderr)
        return 1
    if args.output and len(pdfs) > 1:
        print("Error: -o/--output only allowed for a single PDF.", file=sys.stderr)
        return 1
    default_dir = _default_output_dir()
    for i, pdf in enumerate(pdfs):
        out = args.output if (args.output and len(pdfs) == 1) else str(Path(default_dir) / Path(pdf).with_suffix(".json").name)
        if len(pdfs) > 1:
            print(f"[{i+1}/{len(pdfs)}] {pdf}")
        try:
            result = pdf_to_json(
                str(pdf),
                out,
                overwrite=not args.no_overwrite,
                refine_sections=not getattr(args, "json_raw", False),
                page_text_preview=getattr(args, "json_page_text_preview", False),
                page_text_preview_chars=int(getattr(args, "json_page_text_chars", 800) or 800),
                group_by_page=getattr(args, "json_by_page", False),
            )
            print(f"Saved: {result}")
            if not getattr(args, "no_audit", False):
                _, ok = apply_audit_to_extraction_file(
                    Path(pdf),
                    Path(result),
                    audit_pages=getattr(args, "audit_pages", None),
                    strict=getattr(args, "audit_strict", False),
                    report_path=Path(getattr(args, "audit_report", "")).expanduser()
                    if getattr(args, "audit_report", None)
                    else None,
                )
                if getattr(args, "audit_strict", False) and not ok:
                    return 2
        except Exception as e:
            print(f"Error: {e}", file=sys.stderr)
            return 1
    return 0


def cmd_audit_json(args) -> int:
    """
    Audit a canonical extraction JSON against its source PDF.
    Writes a report JSON (or prints to stdout) with numeric/text gaps, structural issues, invented values.
    """
    import json

    pdf_path = Path(args.pdf)
    json_path = Path(args.json_file)
    if not pdf_path.exists():
        print(f"Error: PDF not found: {pdf_path}", file=sys.stderr)
        return 1
    if not json_path.exists():
        print(f"Error: JSON not found: {json_path}", file=sys.stderr)
        return 1

    report = audit_pdf_vs_extraction_json(
        pdf_path,
        json_path,
        max_pages=getattr(args, "max_pages", None),
    )

    s = report.get("summary") or {}
    print(f"\nAudit: {pdf_path}", file=sys.stderr)
    print(f"  Pages audited            : {s.get('pages_audited')}", file=sys.stderr)
    print(f"  Numeric gaps             : {s.get('pages_with_numeric_gaps')}", file=sys.stderr)
    print(f"  Text gaps                : {s.get('pages_with_text_gaps')}", file=sys.stderr)
    print(f"  Structural issues        : {s.get('pages_with_structural_issues')}", file=sys.stderr)
    print(f"  Invented values          : {s.get('pages_with_invented_values')}", file=sys.stderr)
    print(f"  No sections (has content): {s.get('pages_no_sections_but_content')}", file=sys.stderr)

    out = getattr(args, "output", None)
    if out:
        out_path = Path(out)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"\nFull report written to {out}", file=sys.stderr)
        return 0

    print(json.dumps(report, indent=2, ensure_ascii=False))
    return 0


def cmd_ask(args) -> int:
    pdfs = _expand_pdfs(args.pdf)
    if not pdfs:
        print("Error: No PDF files found.", file=sys.stderr)
        return 1
    if args.output and len(pdfs) > 1:
        print("Error: -o/--output only allowed for a single PDF.", file=sys.stderr)
        return 1
    default_dir = _default_output_dir()
    for i, pdf in enumerate(pdfs):
        if args.output and len(pdfs) == 1:
            out = args.output
        else:
            out = str(Path(default_dir) / Path(pdf).with_suffix(".xlsx").name)
        if len(pdfs) > 1:
            print(f"[{i+1}/{len(pdfs)}] {pdf}")
        try:
            backend = getattr(args, "backend", "anthropic")
            if backend == "smollm":
                from extract_smollm import extract_pdf_to_excel as smollm_extract
                result = smollm_extract(str(pdf), args.query, out, model_name=getattr(args, "smollm_model", "HuggingFaceTB/SmolLM2-360M-Instruct"))
            else:
                from config import load_config
                cfg = load_config()
                result = extract_pdf_to_excel(str(pdf), args.query, out, model=args.model, config=cfg)
            print(f"Saved: {result}")
        except Exception as e:
            msg = str(e).lower()
            if "401" in msg or "auth" in msg or "invalid_api_key" in msg:
                print("Error: Invalid or missing API key. Set ANTHROPIC_API_KEY in .env.", file=sys.stderr)
            elif "429" in msg or "rate" in msg:
                print("Error: API rate limit exceeded. Try again later.", file=sys.stderr)
            else:
                print(f"Error: {e}", file=sys.stderr)
            return 1
    return 0


def cmd_populate_template(args) -> int:
    """
    Template-first flow:
      PDF -> QB extracted workbook -> populate user template -> output populated template.
    """
    import tempfile

    pdfs = _expand_pdfs(args.pdfs)
    if not pdfs:
        print("Error: No PDF files found.", file=sys.stderr)
        return 1
    if len(pdfs) > 1:
        print("Error: populate-template currently supports one PDF at a time.", file=sys.stderr)
        return 1

    pdf = str(pdfs[0])
    template = Path(args.template)
    if not template.exists():
        print(f"Error: Template not found: {template}", file=sys.stderr)
        return 1

    default_dir = _default_output_dir()
    out = args.output or str(Path(default_dir) / f"{Path(template).stem}_populated.xlsx")
    out_path = Path(out)
    if out_path.exists() and args.no_overwrite:
        print(f"Error: Output exists: {out_path}", file=sys.stderr)
        return 1

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        qb_tmp = f.name
    try:
        # 1) Build QB extracted workbook from PDF (existing pipeline).
        pdf_to_qb_excel(pdf, qb_tmp, overwrite=True)
        # 2) Populate the provided template from that workbook.
        result = populate_template_from_qb_output(
            template_path=str(template),
            qb_output_xlsx=qb_tmp,
            output_path=str(out_path),
            account_id=getattr(args, "account_id", None),
        )
        print(f"Saved: {result}")
        return 0
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1
    finally:
        Path(qb_tmp).unlink(missing_ok=True)


def cmd_fields(args) -> int:
    """Extract generic fields as JSON (service-style)."""
    import json
    import tempfile
    from openpyxl import load_workbook

    pdfs = _expand_pdfs(args.pdfs)
    if not pdfs:
        print("Error: No PDF files found.", file=sys.stderr)
        return 1
    if len(pdfs) > 1:
        print("Error: fields currently supports one PDF at a time.", file=sys.stderr)
        return 1

    pdf = str(pdfs[0])
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        qb_tmp = f.name
    try:
        # Build QB workbook first (same path as tables command)
        pdf_to_qb_excel(pdf, qb_tmp, overwrite=True)
        wb = load_workbook(qb_tmp, data_only=True, read_only=True)
        # Build by_target-compatible structure from workbook sheets (simple adapter)
        by_target = {
            name: [(name, [list(r) for r in wb[name].iter_rows(values_only=True)])]
            for name in wb.sheetnames
        }
        wb.close()
        fields = extract_fields(by_target)
        payload = {"fields": [f.to_json() for f in fields]}
        if args.output:
            Path(args.output).write_text(json.dumps(payload, indent=2), encoding="utf-8")
            print(f"Saved: {args.output}")
        else:
            print(json.dumps(payload, indent=2))
        return 0
    finally:
        Path(qb_tmp).unlink(missing_ok=True)


def cmd_populate_template_from_fields(args) -> int:
    """
    Template fill from service output:
      fields.json -> populate user template -> output populated template.
    """
    template = Path(args.template)
    if not template.exists():
        print(f"Error: Template not found: {template}", file=sys.stderr)
        return 1
    fields_path = Path(args.fields_json)
    if not fields_path.exists():
        print(f"Error: fields.json not found: {fields_path}", file=sys.stderr)
        return 1

    default_dir = _default_output_dir()
    out = args.output or str(Path(default_dir) / f"{Path(template).stem}_populated_from_fields.xlsx")
    out_path = Path(out)
    if out_path.exists() and args.no_overwrite:
        print(f"Error: Output exists: {out_path}", file=sys.stderr)
        return 1

    try:
        result = populate_template_from_fields_json(
            template_path=str(template),
            fields_json_path=str(fields_path),
            output_path=str(out_path),
            account_id=getattr(args, "account_id", None),
        )
        print(f"Saved: {result}")
        return 0
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1


def cmd_populate_template_jpm(args) -> int:
    """
    Populate the JPM blocks in a template directly from a JPM PDF.

    This bypasses the QB extraction step and reads the PDF text layer directly,
    writing values into the template's `PLSummary JP Morgan Chase` account blocks.
    """
    pdfs = _expand_pdfs(args.pdfs)
    if not pdfs:
        print("Error: No PDF files found.", file=sys.stderr)
        return 1
    if len(pdfs) > 1:
        print("Error: populate-template-jpm currently supports one PDF at a time.", file=sys.stderr)
        return 1

    pdf = str(pdfs[0])
    template = Path(args.template)
    if not template.exists():
        print(f"Error: Template not found: {template}", file=sys.stderr)
        return 1

    default_dir = _default_output_dir()
    out = args.output or str(Path(default_dir) / f"{Path(template).stem}_jpm_populated.xlsx")
    out_path = Path(out)
    if out_path.exists() and args.no_overwrite:
        print(f"Error: Output exists: {out_path}", file=sys.stderr)
        return 1

    try:
        accounts = None
        if getattr(args, "accounts", None):
            accounts = [a.strip() for a in str(args.accounts).split(",") if a.strip()]
        result = populate_jpm_template_from_pdf(
            template_path=str(template),
            pdf_path=pdf,
            output_path=str(out_path),
            accounts=accounts,
        )
        print(f"Saved: {result}")
        return 0
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1


def cmd_populate_family_template(args) -> int:
    """
    Single output workbook: fill JPM blocks from a JPM PDF and Goldman blocks from
    one or more Goldman PDFs (via QB extraction), without separate per-bank files.
    """
    template = Path(args.template)
    if not template.exists():
        print(f"Error: Template not found: {template}", file=sys.stderr)
        return 1

    jpm_pdf = getattr(args, "jpm_pdf", None)
    gs_pdfs = list(getattr(args, "gs_pdfs", None) or [])
    if not jpm_pdf and not gs_pdfs:
        print("Error: provide --jpm-pdf and/or --gs-pdf (repeat --gs-pdf for multiple).", file=sys.stderr)
        return 1

    default_dir = _default_output_dir()
    out = args.output or str(Path(default_dir) / "family_template_filled.xlsx")
    out_path = Path(out)
    if out_path.exists() and args.no_overwrite:
        print(f"Error: Output exists: {out_path}", file=sys.stderr)
        return 1

    gs_account_arg = getattr(args, "gs_account", None)
    pairs: list[tuple[str, str | None]] = []
    if gs_pdfs:
        if gs_account_arg:
            parts = [x.strip() for x in str(gs_account_arg).split(",") if x.strip()]
            if len(parts) == len(gs_pdfs):
                accts: list[str | None] = parts
            elif len(parts) == 1:
                accts = [parts[0]] * len(gs_pdfs)
            else:
                print(
                    "Error: --gs-account must be one account for all GS PDFs, "
                    "or a comma-separated list with the same length as --gs-pdf.",
                    file=sys.stderr,
                )
                return 1
        else:
            accts = [None] * len(gs_pdfs)
        pairs = [(str(p), a) for p, a in zip(gs_pdfs, accts)]

    jpm_accounts = None
    if getattr(args, "jpm_accounts", None):
        jpm_accounts = [a.strip() for a in str(args.jpm_accounts).split(",") if a.strip()]

    try:
        result = populate_family_template(
            template,
            out_path,
            jpm_pdf=jpm_pdf,
            jpm_accounts=jpm_accounts,
            gs_pdf_account_pairs=pairs or None,
        )
        print(f"Saved: {result}")
        return 0
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1


def main() -> int:
    parser = argparse.ArgumentParser(
        description="PDF → Excel: extract tables (offline) or ask the AI agent for specific data.",
        epilog="Examples:\n  %(prog)s tables report.pdf\n  %(prog)s tables a.pdf b.pdf\n  %(prog)s ask report.pdf \"taxes for January 2026\"",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--version", action="version", version=_get_version())
    sub = parser.add_subparsers(dest="cmd", required=True, help="Command")

    # tables: one or more PDFs
    p_tables = sub.add_parser("tables", help="Extract all tables from PDF(s) to Excel (no AI)")
    p_tables.add_argument("pdfs", nargs="+", help="PDF file(s) or directory containing PDFs")
    p_tables.add_argument("-o", "--output", default=None, help="Output .xlsx path (single PDF only)")
    p_tables.add_argument("--no-overwrite", action="store_true", help="Do not overwrite existing output")
    p_tables.add_argument("--no-audit", action="store_true", help="Skip PDF-vs-JSON audit step (not recommended)")
    p_tables.add_argument(
        "--audit-pages",
        type=int,
        default=None,
        help="Audit page cap: default=all pages; N=first N only (smoke); 0=disable audit",
    )
    p_tables.add_argument("--audit-strict", action="store_true", help="Fail command (exit 2) when audit requires review")
    p_tables.add_argument("--audit-report", default=None, help="Optional path to write full audit report JSON")
    p_tables.set_defaults(func=cmd_tables)

    p_hybrid = sub.add_parser("hybrid", help="Extract PDF with hybrid (library + VL on bad pages); writes JSON with per-page VL timing in meta")
    p_hybrid.add_argument("pdfs", nargs="+", help="PDF file(s) or directory containing PDFs")
    p_hybrid.add_argument("-o", "--output", default=None, help="Output .json path (single PDF only)")
    p_hybrid.add_argument("--schema-type", default=None, help="VL prompt profile: universal, generic, broker_statement, tax_statement")
    p_hybrid.add_argument("--max-pages", type=int, default=None, help="Max pages to process (default: all)")
    p_hybrid.add_argument("--no-overwrite", action="store_true", help="Do not overwrite existing output")
    p_hybrid.add_argument("--excel", action="store_true", help="Also convert the JSON to Excel (same path with .xlsx)")
    p_hybrid.add_argument("--no-audit", action="store_true", help="Skip PDF-vs-JSON audit step (not recommended)")
    p_hybrid.add_argument(
        "--audit-pages",
        type=int,
        default=None,
        help="Audit page cap: default=all pages; N=first N only; 0=disable audit",
    )
    p_hybrid.add_argument("--audit-strict", action="store_true", help="Fail command (exit 2) when audit requires review")
    p_hybrid.add_argument("--audit-report", default=None, help="Optional path to write full audit report JSON")
    p_hybrid.set_defaults(func=cmd_hybrid)

    p_json = sub.add_parser("json", help="Extract PDF to JSON (sections + rows); easy to edit, then convert to Excel")
    p_json.add_argument("pdfs", nargs="+", help="PDF file(s) or directory containing PDFs")
    p_json.add_argument("-o", "--output", default=None, help="Output .json path (single PDF only)")
    p_json.add_argument("--no-overwrite", action="store_true", help="Do not overwrite existing output")
    p_json.add_argument("--no-audit", action="store_true", help="Skip PDF-vs-JSON audit step (not recommended)")
    p_json.add_argument(
        "--audit-pages",
        type=int,
        default=None,
        help="Audit page cap: default=all pages; N=first N only; 0=disable audit",
    )
    p_json.add_argument("--audit-strict", action="store_true", help="Fail command (exit 2) when audit requires review")
    p_json.add_argument("--audit-report", default=None, help="Optional path to write full audit report JSON")
    p_json.add_argument(
        "--json-raw",
        action="store_true",
        help="Keep long narrative-only rows in JSON (skip refine_json_sections; closer to raw extraction)",
    )
    p_json.add_argument(
        "--json-page-text-preview",
        action="store_true",
        help="Add meta.page_text_preview: first N chars of each page text (PyMuPDF) for auditing gaps",
    )
    p_json.add_argument(
        "--json-page-text-chars",
        type=int,
        default=800,
        help="Max characters per page for --json-page-text-preview (default: 800)",
    )
    p_json.add_argument(
        "--json-by-page",
        action="store_true",
        help="Also write top-level 'pages': { \"1\": [sections...], \"2\": [...] } (flat 'sections' kept for compatibility)",
    )
    p_json.set_defaults(func=cmd_json)

    p_audit = sub.add_parser("audit-json", help="Audit extraction JSON against its source PDF (numeric/text gaps, invented values)")
    p_audit.add_argument("pdf", help="Path to the source PDF")
    p_audit.add_argument("json_file", help="Path to extraction JSON (our canonical sections/rows format)")
    p_audit.add_argument("-o", "--output", default=None, help="Write report JSON to this path (optional)")
    p_audit.add_argument("--max-pages", type=int, default=None, help="Limit audit to first N pages (optional)")
    p_audit.set_defaults(func=cmd_audit_json)

    p_from_json = sub.add_parser("from-json", help="Convert JSON (from pdf→json) to Excel; use after editing JSON to map tables")
    p_from_json.add_argument("json_file", help="Path to .json file (extraction output)")
    p_from_json.add_argument("-o", "--output", default=None, help="Output .xlsx path")
    p_from_json.add_argument("--no-overwrite", action="store_true", help="Do not overwrite existing output")
    p_from_json.set_defaults(func=cmd_from_json)

    p_clean_json = sub.add_parser("clean-json", help="Remove repetitive sections from extraction JSON; with --pdf, only clean when PDF confirms phrase is rare on that page")
    p_clean_json.add_argument("json_file", help="Path to .json file (extraction output)")
    p_clean_json.add_argument("--pdf", default=None, help="Path to source PDF; if set, we only collapse/drop when the repeated phrase appears ≤10 times on that page")
    p_clean_json.set_defaults(func=cmd_clean_json)

    p_pop_tpl = sub.add_parser(
        "populate-template",
        help="Populate a provided Excel template from PDF using QB extraction + mapping rules",
    )
    p_pop_tpl.add_argument("pdfs", nargs="+", help="One PDF file path")
    p_pop_tpl.add_argument("--template", required=True, help="Path to template .xlsx")
    p_pop_tpl.add_argument("-o", "--output", default=None, help="Output populated .xlsx path")
    p_pop_tpl.add_argument("--account-id", default=None, help="Optional account id override (e.g. 366-3)")
    p_pop_tpl.add_argument("--no-overwrite", action="store_true", help="Do not overwrite existing output")
    p_pop_tpl.set_defaults(func=cmd_populate_template)

    p_pop_tpl_fields = sub.add_parser(
        "populate-template-from-fields",
        help="Populate a provided Excel template from fields.json (service output)",
    )
    p_pop_tpl_fields.add_argument("--fields-json", required=True, help="Path to fields.json")
    p_pop_tpl_fields.add_argument("--template", required=True, help="Path to template .xlsx")
    p_pop_tpl_fields.add_argument("-o", "--output", default=None, help="Output populated .xlsx path")
    p_pop_tpl_fields.add_argument("--account-id", default=None, help="Optional account id override (e.g. 366-3)")
    p_pop_tpl_fields.add_argument("--no-overwrite", action="store_true", help="Do not overwrite existing output")
    p_pop_tpl_fields.set_defaults(func=cmd_populate_template_from_fields)

    p_pop_tpl_jpm = sub.add_parser(
        "populate-template-jpm",
        help="Populate JPM blocks in a template directly from a JPM PDF (text-layer extraction)",
    )
    p_pop_tpl_jpm.add_argument("pdfs", nargs="+", help="One JPM PDF file path")
    p_pop_tpl_jpm.add_argument("--template", required=True, help="Path to template .xlsx")
    p_pop_tpl_jpm.add_argument("-o", "--output", default=None, help="Output populated .xlsx path")
    p_pop_tpl_jpm.add_argument("--accounts", default=None, help="Optional comma-separated account list (e.g. 1004,9004)")
    p_pop_tpl_jpm.add_argument("--no-overwrite", action="store_true", help="Do not overwrite existing output")
    p_pop_tpl_jpm.set_defaults(func=cmd_populate_template_jpm)

    p_pop_fam = sub.add_parser(
        "populate-family-template",
        help="Fill one template workbook from JPM PDF + Goldman PDF(s) (single .xlsx output)",
    )
    p_pop_fam.add_argument("--template", required=True, help="Path to family_template.xlsx (empty template)")
    p_pop_fam.add_argument("-o", "--output", default=None, help="Output .xlsx path (default: output/family_template_filled.xlsx)")
    p_pop_fam.add_argument("--jpm-pdf", default=None, help="JP Morgan statement PDF (optional)")
    p_pop_fam.add_argument(
        "--jpm-accounts",
        default=None,
        help="Comma-separated JPM account blocks to fill, e.g. 1004,9004 (default: all in template)",
    )
    p_pop_fam.add_argument(
        "--gs-pdf",
        action="append",
        default=[],
        dest="gs_pdfs",
        metavar="PATH",
        help="Goldman PDF (repeat for multiple statements)",
    )
    p_pop_fam.add_argument(
        "--gs-account",
        default=None,
        help="Goldman template account id(s): one value for all --gs-pdf, or comma list matching each PDF",
    )
    p_pop_fam.add_argument("--no-overwrite", action="store_true", help="Do not overwrite existing output")
    p_pop_fam.set_defaults(func=cmd_populate_family_template)

    p_fields = sub.add_parser(
        "fields",
        help="Extract generic fields (JSON) from a PDF using the hybrid/QB pipeline",
    )
    p_fields.add_argument("pdfs", nargs="+", help="One PDF file path")
    p_fields.add_argument("-o", "--output", default=None, help="Write output JSON to file (optional)")
    p_fields.set_defaults(func=cmd_fields)

    # ask: PDF(s) + query
    p_ask = sub.add_parser("ask", help="AI agent: extract what you ask for from PDF(s)")
    p_ask.add_argument("pdf", nargs="+", help="PDF file(s) or directory containing PDFs")
    p_ask.add_argument("query", help="What to extract, e.g. 'company taxes for January 2026'")
    p_ask.add_argument("-o", "--output", default=None, help="Output .xlsx path (single PDF only)")
    p_ask.add_argument("--backend", choices=("anthropic", "smollm"), default="anthropic", help="AI backend: anthropic (API) or smollm (offline local)")
    p_ask.add_argument("--model", default="claude-sonnet-4-20250514", help="Anthropic model (when backend=anthropic)")
    p_ask.add_argument("--smollm-model", default="HuggingFaceTB/SmolLM2-360M-Instruct", help="HuggingFace model name (when backend=smollm)")
    p_ask.set_defaults(func=cmd_ask)

    args = parser.parse_args()
    try:
        return args.func(args)
    except FileNotFoundError as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
