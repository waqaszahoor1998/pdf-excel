#!/usr/bin/env python3
"""
Build a PLSummary-style sheet (like QB Automation Sheet sample) from extracted PDF data.

Creates "PLSummary J.P. Morgan Chase" with:
- Header: Entity, Report title, Broker, Account ID
- Columns: BOM, MTD, MTD, MTD, EOM | Account Name, Market Value, Cash In, Cash Out, PNL, Market Value
- Rows: Investments, Cash and cash equivalents, Totals, then line items (Accrued Dividend, Interest Income, etc.)
- Balance Sheet block: Accrued Dividend and Interest, Cash - Due from Broker, Investment in Securities, Total Asset, etc.

Maps from our extracted Account Summary, Asset Allocation, Portfolio Activity, Tax Summary.
"""

import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

FILL_HEADER = PatternFill(fill_type="solid", fgColor="92D050")
FILL_TABLE_HEADER = PatternFill(fill_type="solid", fgColor="D9E1F2")
FILL_TOTALS = PatternFill(fill_type="solid", fgColor="FFC000")


def _fill_row(ws, row: int, num_cols: int, fill: PatternFill) -> None:
    for c in range(1, num_cols + 1):
        try:
            ws.cell(row=row, column=c).fill = fill
        except Exception:
            pass


def _num(cell) -> float | int | None:
    """Extract number from cell; return None if not a number."""
    if cell is None:
        return None
    if isinstance(cell, (int, float)):
        return cell
    s = str(cell).strip().replace(",", "").lstrip("$€£")
    if not s or s in ("", "-", "—"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _find_first_data_row(rows: list[list], min_cols: int = 3) -> list[list]:
    """Return rows that look like data (have at least one number in first min_cols columns)."""
    out = []
    for row in rows or []:
        r = row if isinstance(row, (list, tuple)) else [row]
        vals = [r[i] if i < len(r) else None for i in range(max(min_cols, 6))]
        if any(_num(v) is not None for v in vals):
            out.append(r)
    return out


def _parse_account_summary_block(rows: list[list]) -> dict:
    """From first Account Summary block get: total_beginning, total_ending, total_change, accounts (list of {name, id, begin, end, change})."""
    data = {"total_beginning": None, "total_ending": None, "total_change": None, "accounts": []}
    data_rows = _find_first_data_row(rows, 4)
    for r in data_rows:
        cells = list(r) if len(r) >= 6 else (list(r) + [None] * (6 - len(r)))
        name = (cells[0] or "").strip() if cells else ""
        acct_id = (cells[2] or "").strip() if len(cells) > 2 else ""
        begin = _num(cells[3]) if len(cells) > 3 else None
        end = _num(cells[4]) if len(cells) > 4 else None
        chg = _num(cells[5]) if len(cells) > 5 else None
        if name and (begin is not None or end is not None):
            if "total" in name.lower():
                data["total_beginning"] = begin
                data["total_ending"] = end
                data["total_change"] = chg
            else:
                data["accounts"].append({"name": name, "id": acct_id, "begin": begin, "end": end, "change": chg})
    return data


def _parse_asset_allocation_block(rows: list[list]) -> dict:
    """Get equity, cash_fixed, market_value, accruals, market_value_with_accruals from first block."""
    data = {"equity": None, "cash_fixed": None, "market_value": None, "accruals": None, "market_value_with_accruals": None}
    for r in (rows or []):
        r = r if isinstance(r, (list, tuple)) else [r]
        label = (str(r[0] or "").strip().lower() if r else "")[:50]
        if "equity" in label and len(r) >= 2:
            data["equity"] = _num(r[2]) or _num(r[1])  # ending or first number
        elif "cash" in label and "fixed" in label and len(r) >= 2:
            data["cash_fixed"] = _num(r[2]) or _num(r[1])
        elif label == "market value" and len(r) >= 3:
            data["market_value"] = _num(r[2]) or _num(r[1])
        elif label == "accruals" and len(r) >= 2:
            data["accruals"] = _num(r[2]) or _num(r[1])
        elif "market value with accruals" in label and len(r) >= 3:
            data["market_value_with_accruals"] = _num(r[2]) or _num(r[1])
    return data


def _parse_portfolio_activity_block(rows: list[list]) -> dict:
    """Get beginning_mv, ending_mv, income_dist, change_inv, accruals, mv_with_accruals."""
    data = {"beginning_mv": None, "ending_mv": None, "income_dist": None, "change_inv": None, "accruals": None, "mv_with_accruals": None}
    for r in (rows or []):
        r = r if isinstance(r, (list, tuple)) else [r]
        label = (str(r[0] or "").strip().lower() if r else "")[:50]
        v = _num(r[2]) if len(r) > 2 else _num(r[1]) if len(r) > 1 else None
        if "beginning market" in label:
            data["beginning_mv"] = v
        elif "ending market" in label and "accrual" not in label:
            data["ending_mv"] = v
        elif "income" in label and "distribution" in label:
            data["income_dist"] = v
        elif "change in investment" in label:
            data["change_inv"] = v
        elif label == "accruals":
            data["accruals"] = v
        elif "market value with accruals" in label:
            data["mv_with_accruals"] = v
    return data


def _first_number_in_row(r: list, start_col: int = 1) -> float | None:
    """First numeric value in row from start_col onward (skip label in col 0)."""
    for i in range(start_col, len(r)):
        v = _num(r[i])
        if v is not None:
            return v
    return None


def _parse_tax_summary_block(rows: list[list]) -> dict:
    """Get dividend_income, interest_income, realized_gain, taxable_income, etc. (consolidated or first account)."""
    data = {"dividend_income": None, "interest_income": None, "realized_gain": None, "taxable_income": None}
    for r in (rows or []):
        r = r if isinstance(r, (list, tuple)) else [r]
        label = (str(r[0] or "").strip().lower() if r else "")[:60]
        v = _first_number_in_row(r, 1)
        if "dividend" in label:
            data["dividend_income"] = v or data.get("dividend_income")
        elif "interest income" in label:
            data["interest_income"] = v or data.get("interest_income")
        elif "realized" in label and "gain" in label:
            data["realized_gain"] = v or data.get("realized_gain")
        elif "taxable income" in label:
            data["taxable_income"] = v or data.get("taxable_income")
    return data


def build_plsummary_jpm_sheet(
    wb_out: Workbook,
    by_target: dict[str, list[tuple[str, list[list]]]],
    entity_name: str = "ABC Trust",
    report_title: str = "MTD PNL Per Trading Account Summary",
    broker_name: str = "J.P. Morgan Chase",
) -> None:
    """
    Add a worksheet "PLSummary J.P. Morgan Chase" to wb_out with sample-like layout.
    Fills from by_target (Account Summary, Asset Allocation, Portfolio Activity, Tax Summary).
    """
    ws = wb_out.create_sheet(title="PLSummary J.P. Morgan Chase", index=0)

    # Get first block of each type
    acc_rows = by_target.get("Account Summary", [])
    acc_block = acc_rows[0][1] if acc_rows else []
    asset_rows = by_target.get("Asset Allocation", [])
    asset_block = asset_rows[0][1] if asset_rows else []
    port_rows = by_target.get("Portfolio Activity", [])
    port_block = port_rows[0][1] if port_rows else []
    tax_rows = by_target.get("Tax Summary", [])
    tax_block = []
    for _section, rows in (tax_rows or []):
        tax_block.extend(rows or [])

    acc = _parse_account_summary_block(acc_block)
    asset = _parse_asset_allocation_block(asset_block)
    port = _parse_portfolio_activity_block(port_block)
    tax = _parse_tax_summary_block(tax_block)

    # Resolve key numbers (prefer totals, then first account)
    bom = port.get("beginning_mv") or acc.get("total_beginning") or (acc["accounts"][0]["begin"] if acc["accounts"] else None)
    eom = port.get("ending_mv") or acc.get("total_ending") or (acc["accounts"][0]["end"] if acc["accounts"] else None)
    eom_with_accruals = port.get("mv_with_accruals") or asset.get("market_value_with_accruals") or acc.get("total_ending")
    pnl = acc.get("total_change") or port.get("change_inv") or (acc["accounts"][0]["change"] if acc["accounts"] else None)
    mv_with_accruals = eom_with_accruals
    investments = asset.get("market_value") or asset.get("equity")
    cash_equiv = asset.get("cash_fixed")
    if investments is None and acc.get("total_beginning") is not None:
        investments = acc["total_beginning"]
    if cash_equiv is None:
        cash_equiv = 0
    if investments is None:
        investments = 0
    totals = mv_with_accruals or acc.get("total_ending") or eom
    accruals = port.get("accruals") or asset.get("accruals") or 0
    dividend_income = tax.get("dividend_income")
    interest_income = tax.get("interest_income")

    # Row 1–4: Header
    ws.cell(1, 1, entity_name)
    ws.cell(2, 1, report_title)
    ws.cell(3, 1, broker_name)
    acct_id = acc["accounts"][0]["id"] if acc["accounts"] else "E79271004"
    ws.cell(4, 1, acct_id)
    # Rows 5–6 blank
    # Row 7: BOM, MTD, MTD, MTD, EOM
    for col, label in enumerate(["BOM", "MTD", "MTD", "MTD", "EOM"], start=2):
        ws.cell(7, col, label)
    ws.cell(7, 2).fill = FILL_TABLE_HEADER
    ws.cell(7, 3).fill = FILL_TABLE_HEADER
    ws.cell(7, 4).fill = FILL_TABLE_HEADER
    ws.cell(7, 5).fill = FILL_TABLE_HEADER
    ws.cell(7, 6).fill = FILL_TABLE_HEADER
    # Row 8: Account Name, Market Value, Cash In, Cash Out, PNL, Market Value
    for col, label in enumerate(["Account Name", "Market Value", "Cash In", "Cash Out", "PNL", "Market Value"], start=1):
        ws.cell(8, col, label)
    _fill_row(ws, 8, 6, FILL_TABLE_HEADER)

    # Data rows (sample order). Col 2 = BOM/Market Value, 3-4 = Cash In/Out, 5 = PNL, 6 = EOM/Market Value
    row = 9
    # Investments: BOM, 0, 0, PNL, EOM
    ws.cell(row, 1, "Investments")
    ws.cell(row, 2, round(bom, 2) if bom is not None else None)
    ws.cell(row, 3, 0)
    ws.cell(row, 4, 0)
    ws.cell(row, 5, round(pnl, 2) if pnl is not None else None)
    ws.cell(row, 6, round(eom, 2) if eom is not None else None)
    row += 2
    # Cash and cash equivalents
    ws.cell(row, 1, "Cash and cash equivalents")
    ws.cell(row, 2, round(cash_equiv, 2) if cash_equiv is not None else 0)
    ws.cell(row, 3, 0)
    ws.cell(row, 4, 0)
    ws.cell(row, 5, 0)
    ws.cell(row, 6, round(cash_equiv, 2) if cash_equiv else None)
    row += 2
    # Totals
    ws.cell(row, 1, "Totals")
    ws.cell(row, 2, round(bom, 2) if bom is not None else None)
    ws.cell(row, 3, 0)
    ws.cell(row, 4, 0)
    ws.cell(row, 5, round(pnl, 2) if pnl is not None else None)
    ws.cell(row, 6, round(mv_with_accruals, 2) if mv_with_accruals is not None else None)
    ws.cell(row, 1).fill = FILL_TOTALS
    ws.cell(row, 2).fill = FILL_TOTALS
    ws.cell(row, 5).fill = FILL_TOTALS
    ws.cell(row, 6).fill = FILL_TOTALS
    row += 1
    # Accrued Dividend, Accrued Interest
    ws.cell(row, 1, "Accrued Dividend")
    ws.cell(row, 2, round(accruals, 2) if accruals else None)
    row += 1
    ws.cell(row, 1, "Accrued Interest")
    row += 1
    # Line items (sample has Accounting Fees, Broker fees, etc. - use 0 or from tax)
    for label in ["Accounting Fees", "Broker fees", "Change in unrealized gain/loss", "Dividend Income", "Franchise Tax",
                  "Interest expense", "Interest Income", "Long term realized gain/loss", "Other income",
                  "Short term realized gain/loss", "Withholding taxes"]:
        ws.cell(row, 1, label)
        if "Dividend" in label and dividend_income is not None:
            ws.cell(row, 2, round(dividend_income, 2))
        elif "Interest Income" in label and interest_income is not None:
            ws.cell(row, 2, round(interest_income, 2))
        else:
            ws.cell(row, 2, 0)
        row += 1
    row += 1
    # Balance Sheet section
    ws.cell(row, 1, "Balance Sheet")
    row += 2
    ws.cell(row, 1, "Accrued Dividend and Interest")
    ws.cell(row, 2, round(accruals, 2) if accruals else None)
    row += 1
    ws.cell(row, 1, "Cash - Due from Broker")
    ws.cell(row, 2, round(cash_equiv, 2) if cash_equiv else None)
    row += 1
    ws.cell(row, 1, "Investment in Securities")
    ws.cell(row, 2, round(investments, 2) if investments is not None else None)
    row += 2
    ws.cell(row, 1, "Total Asset")
    ws.cell(row, 2, round(totals, 2) if totals is not None else None)
    ws.cell(row, 1).fill = FILL_TOTALS
    ws.cell(row, 2).fill = FILL_TOTALS
    row += 3
    ws.cell(row, 1, "Securities sold short")
    ws.cell(row, 2, 0)
    row += 1
    ws.cell(row, 1, "Opening Capital")
    ws.cell(row, 2, 0)
    row += 1
    ws.cell(row, 1, "Capital Contribution")
    ws.cell(row, 2, round(bom, 2) if bom is not None else None)
    row += 1
    ws.cell(row, 1, "Capital Withdrawal")
    ws.cell(row, 2, 0)
    row += 1
    ws.cell(row, 1, "Retained Earning")
    row += 1
    ws.cell(row, 1, "Accrued Expense")
    ws.cell(row, 2, 0)
    row += 1
    ws.cell(row, 1, "Net Income")
    ws.cell(row, 2, round(pnl, 2) if pnl is not None else None)
    row += 1
    ws.cell(row, 1, "Total Liabilities")
    ws.cell(row, 2, round(totals, 2) if totals is not None else None)
    ws.cell(row, 1).fill = FILL_TOTALS
    ws.cell(row, 2).fill = FILL_TOTALS
    row += 3
    ws.cell(row, 1, "Difference")
    ws.cell(row, 2, 0)


def _fill_row(ws, row: int, num_cols: int, fill: PatternFill) -> None:
    for c in range(1, num_cols + 1):
        try:
            ws.cell(row=row, column=c).fill = fill
        except Exception:
            pass
