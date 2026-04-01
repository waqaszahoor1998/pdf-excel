import pytest
from openpyxl import Workbook

from qb_posting_rules import (
    map_interest_from_purchases_sales,
    extract_cash_closing_balance,
    extract_purchases_sales_accrued_interest,
    extract_unrealized_change_for_account,
    extract_change_in_accrued_interest_for_account,
    extract_change_in_accrued_dividend_for_account,
)

from plsummary_builder import build_plsummary_jpm_sheet


def test_map_interest_from_purchases_sales_negative():
    paid, income = map_interest_from_purchases_sales(-8278.06)
    assert paid == pytest.approx(8278.06)
    assert income == pytest.approx(0.0)


def test_map_interest_from_purchases_sales_positive():
    paid, income = map_interest_from_purchases_sales(1200.5)
    assert paid == pytest.approx(0.0)
    assert income == pytest.approx(1200.5)


def test_extract_cash_closing_balance_from_cash_activity():
    by_target = {
        "Cash Activity": [
            ("Cash Activity (Continued)", [
                ["Type", "X", "Y", "End of Day Balance"],
                ["CLOSING BALANCE AS OF DEC 31 25", None, None, "155206.79"],
            ])
        ]
    }
    assert extract_cash_closing_balance(by_target) == pytest.approx(155206.79)


def test_extract_purchases_sales_accrued_interest_total():
    by_target = {
        "Purchases and Sales": [
            ("Purchases & Sales", [
                ["A", "B", "Accrued Interest", "D"],
                ["TOTAL PURCHASES & SALES", None, "(8,278.06)", None],
            ])
        ]
    }
    assert extract_purchases_sales_accrued_interest(by_target) == pytest.approx(-8278.06)


def test_account_specific_latest_value_extractors():
    account_id = "366-3"
    by_target = {
        "Unrealized": [
            ("Unrealized", [
                [account_id],
                ["Change in Unrealized", "-", 10, "75,079.07"],
                ["Other row", 1],
            ])
        ],
        "Change in Interest": [
            ("Change in Interest", [
                [account_id],
                ["Change in Accrued Interest", "-", 1, "3,418.32"],
            ])
        ],
        "Change in Dividend": [
            ("Change in Dividend", [
                [account_id],
                ["Change in Accrued Dividend", "-", 2, "123.45"],
            ])
        ],
    }

    assert extract_unrealized_change_for_account(by_target, account_id) == pytest.approx(75079.07)
    assert extract_change_in_accrued_interest_for_account(by_target, account_id) == pytest.approx(3418.32)
    assert extract_change_in_accrued_dividend_for_account(by_target, account_id) == pytest.approx(123.45)


def test_plsummary_builder_applies_rules():
    account_id = "366-3"
    by_target = {
        # Minimum structures required by plsummary_builder parsing helpers
        "Account Summary": [
            ("Account Summary", [
                ["Account Name", "X", "Account ID", "Beginning", "Ending", "Change"],
                ["AB REVOCABLE TRUST", "", account_id, 1000, 1100, 100],
                ["Totals", "", "TOTAL", 1000, 1100, 100],
            ])
        ],
        "Asset Allocation": [
            ("Asset Allocation", [
                ["Market Value", "", 1000],
                ["Cash Fixed", "", 200],
                ["Accruals", "", 5],
                ["Market Value With Accruals", "", 1105],
            ])
        ],
        "Portfolio Activity": [
            ("Portfolio Activity", [
                ["Beginning Market Value", "", 1000],
                ["Ending Market Value", "", 1100],
                ["Accruals", "", 5],
                ["Market Value With Accruals", "", 1105],
            ])
        ],
        "Tax Summary": [
            ("Tax Summary", [
                ["Dividend Income", 0, 168207.23],
                ["Interest Income", 0, 9999.99],  # should be overridden by Purchases & Sales sign rule
                ["Long term realized gain/loss", 0, 14875.57],
            ])
        ],
        "Purchases and Sales": [
            ("Purchases & Sales", [
                ["A", "B", "Accrued Interest", "D"],
                ["TOTAL PURCHASES & SALES", None, "(8,278.06)", None],
            ])
        ],
        "Cash Activity": [
            ("Cash Activity (Continued)", [
                ["Type", "X", "Y", "End of Day Balance"],
                ["CLOSING BALANCE AS OF DEC 31 25", None, None, "155206.79"],
            ])
        ],
        "Unrealized": [
            ("Unrealized", [
                [account_id],
                ["Change in Unrealized", "-", 10, "75079.07"],
            ])
        ],
        "Change in Interest": [
            ("Change in Interest", [
                [account_id],
                ["Change in Accrued Interest", "-", 1, "3418.32"],
            ])
        ],
        "Change in Dividend": [
            ("Change in Dividend", [
                [account_id],
                ["Change in Accrued Dividend", "-", 2, "3184.10"],
            ])
        ],
    }

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    build_plsummary_jpm_sheet(wb, by_target)

    ws = wb["PLSummary"]

    def find_row(label: str) -> int:
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == label:
                return r
        raise AssertionError(f"Row label not found: {label}")

    cash_row = find_row("Cash - Due from Broker")
    assert ws.cell(cash_row, 2).value == pytest.approx(155206.79)

    paid_row = find_row("Interest Paid on Purchases")
    assert ws.cell(paid_row, 2).value == pytest.approx(8278.06)

    income_row = find_row("Interest Income")
    assert ws.cell(income_row, 2).value == pytest.approx(0.0)

    dui_row = find_row("Change in Accrued Interest")
    assert ws.cell(dui_row, 2).value == pytest.approx(3418.32)

    unreal_row = find_row("Change in unrealized gain/loss")
    assert ws.cell(unreal_row, 2).value == pytest.approx(75079.07)

