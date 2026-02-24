"""Unit tests for extract.py helpers (CSV parsing and Excel write). No API calls."""

import tempfile
from pathlib import Path

import pytest

from extract import extract_csv_from_response, csv_to_excel


class TestExtractCsvFromResponse:
    """Tests for extract_csv_from_response()."""

    def test_explicit_block(self):
        text = "Some preamble\n---BEGIN CSV---\na,b\n1,2\n---END CSV---\nTrailer"
        assert extract_csv_from_response(text) == "a,b\n1,2"

    def test_markdown_csv_block(self):
        text = "Here is the data:\n```csv\na,b\n1,2\n```"
        assert extract_csv_from_response(text) == "a,b\n1,2"

    def test_markdown_generic_block(self):
        text = "```\na,b\n1,2\n```"
        assert extract_csv_from_response(text) == "a,b\n1,2"

    def test_last_resort_comma_line(self):
        text = "No block here\njust,a,line\nand,another,row"
        result = extract_csv_from_response(text)
        assert "just,a,line" in result
        assert "and,another,row" in result

    def test_no_csv_raises(self):
        text = "No commas or blocks at all"
        with pytest.raises(ValueError, match="No CSV block"):
            extract_csv_from_response(text)


class TestCsvToExcel:
    """Tests for csv_to_excel()."""

    def test_basic(self):
        csv_content = "col1,col2\nval1,val2"
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            csv_to_excel(csv_content, path)
            assert Path(path).exists()
            # Quick sanity: open and check row count (openpyxl)
            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb.active
            assert ws.title == "Extracted"
            rows = list(ws.iter_rows(values_only=True))
            assert len(rows) == 2
            assert rows[0] == ("col1", "col2")
            assert rows[1] == ("val1", "val2")
        finally:
            Path(path).unlink(missing_ok=True)

    def test_empty_csv_raises(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            with pytest.raises(ValueError, match="no rows"):
                csv_to_excel("", path)
        finally:
            Path(path).unlink(missing_ok=True)

    def test_creates_parent_dir(self):
        csv_content = "a\n1"
        with tempfile.TemporaryDirectory() as d:
            out = Path(d) / "sub" / "out.xlsx"
            csv_to_excel(csv_content, str(out))
            assert out.exists()

    def test_quoted_comma_in_cell(self):
        """CSV with comma inside quoted value."""
        csv_content = 'col1,col2\n"a,b",c'
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            csv_to_excel(csv_content, path)
            from openpyxl import load_workbook
            wb = load_workbook(path)
            rows = list(wb.active.iter_rows(values_only=True))
            assert rows[1][0] == "a,b"
            assert rows[1][1] == "c"
        finally:
            Path(path).unlink(missing_ok=True)
